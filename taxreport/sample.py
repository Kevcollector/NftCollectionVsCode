import json
import os

import pandas as pd
import requests as requests
import pathlib
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime

user = 'goattude'
buyPrice = 0
sellPrice = 0
profit = 0
flippers_sell = (
    "https://proton.api.atomicassets.io/atomicmarket/v1/sales?state=3&seller={}&page=1&limit=100&order=desc&sort=created".format(user))
flippers_buys = (
    "https://proton.api.atomicassets.io/atomicmarket/v1/sales?state=3&buyer={}&page=1&limit=100&order=desc&sort=created".format(user))
flippers_sell = requests.get(flippers_sell).text
flippers_sell = json.loads(flippers_sell)
flippers_buyt = requests.get(flippers_buys).text
flippers_buy = json.loads(flippers_buyt)
flippers_List_buy = []
flippers_List_sell = []
count = 0

start = flippers_buyt.find(":")+1
end = flippers_buyt.find(",")

print(flippers_buyt[start:end])

time = 0
assitID2 = ""

while len(flippers_buy['data']) != 0:
    for data_info in flippers_buy['data']:
        fixedC = 0
        fixedX = 0
        Type = data_info["listing_symbol"]
        if Type == "XPR":
            number = data_info["listing_price"]
            fixedX = int(number) / 10000

        if Type == "XUSDC":
            number = data_info["listing_price"]
            fixedC = int(number) / 1000000
        number = data_info["listing_price"]
        buyPrice = int(number) / 1000000
        sellers = data_info['seller']
        name = data_info['assets'][0]['name']
        timef = data_info['updated_at_time']
        timeb = int(timef) / 1000
        number_of_nft = int(data_info['assets'][0]['template_mint'])
        RoR = float(data_info['assets'][0]['collection']['market_fee'])
        Collection_n = data_info['assets'][0]['collection']['name']
        author_n = data_info['assets'][0]['collection']['author']
        local_time = datetime.utcfromtimestamp(
            timeb).strftime('%d-%m-%Y %H:%M:%S')
        assitID1 = data_info['assets'][0]['asset_id']

        if sellers != user and assitID1 != assitID2:
            assitID2 = assitID1
            flippers_List_buy.append(
                [name, number_of_nft, Collection_n, author_n, sellers, fixedX, fixedC, local_time])

    flippers_buy = (
        "https://proton.api.atomicassets.io/atomicmarket/v1/sales?state=3&buyer={}&before={}&page=1&limit=100&order=desc&sort=updated".format(
            user, timef))
    flippers_buyt = requests.get(flippers_buy).text
    flippers_buy = json.loads(flippers_buyt)
    len(flippers_buy['data'])
    lent = len(flippers_buy['data'])

    # print(flippers_buy)


buyer = pd.DataFrame(data=flippers_List_buy,
                     columns=["name of nft", "# of nft", "collection", "author", "bought from", "bought for Xpr", "bought for usdc",
                              "time"])

buyer.set_index("name of nft")
paidxpr = buyer['bought for Xpr'].sum()
paidusd = buyer['bought for usdc'].sum()
while len(flippers_sell['data']) != 0:
    for data_info in flippers_sell['data']:
        number = data_info["listing_price"]
        fixed = int(number) / 1000000
        name = data_info['assets'][0]['name']
        buyers = data_info['buyer']
        time = data_info['assets'][0]['transferred_at_time']
        timef = data_info['updated_at_time']
        time = int(time)/1000
        number_of_nft = int(data_info['assets'][0]['template_mint'])
        RoR = float(data_info['assets'][0]['collection']['market_fee'])
        fixed -= fixed * RoR
        sellPrice = fixed
        Collection_n = data_info['assets'][0]['collection']['name']
        author_n = data_info['assets'][0]['collection']['author']
        local_time = datetime.utcfromtimestamp(
            time).strftime('%d-%m-%Y %H:%M:%S')
        assitID1 = data_info['assets'][0]['asset_id']
        if assitID1 == assitID2:
            print(buyPrice)
        if author_n != user and assitID1 != assitID2:
            assitID2 = assitID1
            flippers_List_sell.append(
                [name, number_of_nft, Collection_n, author_n,  buyers, sellPrice])
    flippers_sell = (
        "https://proton.api.atomicassets.io/atomicmarket/v1/sales?state=3&seller={}&before={}&page=1&limit=100&order=desc&sort=updated".format(
            user, timef))
    flippers_sell = requests.get(flippers_sell).text
    flippers_sell = json.loads(flippers_sell)

sells = pd.DataFrame(data=flippers_List_sell,
                     columns=["name of nft", "# of nft", "collection", "author",
                              "sold to ", "sold for"])
sells.set_index("name of nft")


paid = buyer['bought for usdc'].sum()
sold = sells['sold for'].sum()
mergedDf = pd.merge(buyer, sells, how='inner',
                    on='name of nft', suffixes=('', '_drop'))
mergedDf.drop([col for col in mergedDf.columns if 'drop' in col],
              axis=1, inplace=True)
mergedDf['profits'] = (mergedDf['sold for']-mergedDf['bought for usdc'])
profits = mergedDf['profits'].sum()
mbuys = mergedDf['bought for usdc'].sum()
mxbuys = mergedDf['bought for Xpr'].sum()
msells = mergedDf['sold for'].sum()
wb = Workbook()
# name_df.add(names_df)
#name_df=pd.merge([name_df, names_df])
ws = wb.active
ws2 = wb.create_sheet()
ws3 = wb.create_sheet()
ws.title = ("{} Buys".format(user))
for r in dataframe_to_rows(buyer, index=False):
    ws.append(r)
dims = {}
for row in ws.rows:
    for cell in row:
        if cell.value:
            dims[cell.column_letter] = max(
                (dims.get(cell.column_letter, 0), len(str(cell.value))))
for col, value in dims.items():
    ws.column_dimensions[col].width = value
maxrow = ws.max_row+2
ws.cell(row=maxrow, column=6, value=paid)
ws.cell(row=maxrow, column=1, value="totals")
ws2.title = ("{} Sells".format(user))
for r in dataframe_to_rows(sells, index=False):
    ws2.append(r)
dims = {}
for row in ws2.rows:
    for cell in row:
        if cell.value:
            dims[cell.column_letter] = max(
                (dims.get(cell.column_letter, 0), len(str(cell.value))))
for col, value in dims.items():
    ws2.column_dimensions[col].width = value
maxrow = ws2.max_row+2
ws2.cell(row=maxrow, column=6, value=sold)
ws2.cell(row=maxrow, column=1, value="totals")
ws3.title = ("{} Flipping ".format(user))
for r in dataframe_to_rows(mergedDf, index=False):
    ws3.append(r)
dims = {}
for row in ws3.rows:
    for cell in row:
        if cell.value:
            dims[cell.column_letter] = max(
                (dims.get(cell.column_letter, 0), len(str(cell.value))))
for col, value in dims.items():
    ws3.column_dimensions[col].width = value
maxrow = ws3.max_row+2
ws3.cell(row=maxrow, column=6, value=mbuys)
ws3.cell(row=maxrow, column=8, value=msells)
ws3.cell(row=maxrow, column=10, value=profits)
ws3.cell(row=maxrow, column=1, value="totals")

wb.save("{}.xlsx".format(user))
