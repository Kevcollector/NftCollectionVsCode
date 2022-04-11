import modules.ApiClass as Api
import time
import json
import os

import pandas as pd
import requests as requests
import pathlib
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime

user = "kevcollector"

buyPrice = 0
sellPrice = 0
profit = 0

flippers_sell = "https://proton.api.atomicassets.io/atomicmarket/v1/sales?state=3&seller={}&page=1&limit=100&order=desc&sort=created".format(
    user
)

flippers_buys = "https://proton.api.atomicassets.io/atomicmarket/v1/sales?state=3&buyer={}&page=1&limit=100&order=desc&sort=created".format(
    user
)

flippers_buy_auction = "https://proton.api.atomicassets.io/atomicmarket/v1/auctions?state=3&buyer={}&page=1&limit=100&order=desc&sort=created".format(
    user
)

flippers_sell_auction = "https://proton.api.atomicassets.io/atomicmarket/v1/auctions?state=3&seller={}&page=1&limit=100&order=desc&sort=created".format(
    user
)

flippersSellOffers = "https://proton.api.atomicassets.io/atomicmarket/v1/buyoffers?state=3&seller={}&page=1&limit=100&order=desc&sort=created".format(
    user
)

flippersBuyOffers = "https://proton.api.atomicassets.io/atomicmarket/v1/buyoffers?state=3&buyer={}&page=1&limit=100&order=desc&sort=created".format(
    user
)

print()
# TODO add in the other apis ( offers, acutions), add this into the flippers_List_buy and sells recpectivly -Y
flippers_sell = requests.get(flippers_sell).text
flippers_sell = json.loads(flippers_sell)
flippers_buyt = requests.get(flippers_buys).text
flippers_buy = json.loads(flippers_buyt)
flippers_buy_auction = requests.get(flippers_buy_auction).text
flippers_buy_auction = json.loads(flippers_buy_auction)
flippers_sell_auction = requests.get(flippers_sell_auction).text
flippers_sell_auction = json.loads(flippers_sell_auction)
flippersBuyOffers = requests.get(flippersBuyOffers).text
flippersBuyOffers = json.loads(flippersBuyOffers)
flippersSellOffers = requests.get(flippersSellOffers).text
flippersSellOffers = json.loads(flippersSellOffers)
flippers_List_buy = []
flippers_List_sell = []
count = 0


assitID2 = ""

while len(flippers_buy["data"]) != 0:
    for data_info in flippers_buy["data"]:
        fixedC = 0
        fixedX = 0
        fixedL = 0
        fixedF = 0
        Type = data_info["listing_symbol"]
        # TODO add in other coins loan and foobar (will need in all loops below too) -Y
        if Type == "XPR":
            number = data_info["listing_price"]
            fixedX = int(number) / 10000
        if Type == "XUSDC":
            number = data_info["listing_price"]
            fixedC = int(number) / 1000000
        if Type == "LOAN":
            number = data_info["listing_price"]
            fixedL = int(number) / 10000
        if Type == "FOOBAR":
            number = data_info["listing_price"]
            fixedF = int(number) / 10000
        number = data_info["listing_price"]
        buyPrice = int(number) / 1000000
        sellers = data_info["seller"]
        name = data_info["assets"][0]["name"]
        timef = data_info["updated_at_time"]
        timeU = data_info["created_at_time"]
        timeb = int(timef) / 1000
        number_of_nft = int(data_info["assets"][0]["template_mint"])
        RoR = float(data_info["assets"][0]["collection"]["market_fee"])
        Collection_n = data_info["assets"][0]["collection"]["name"]
        author_n = data_info["assets"][0]["collection"]["author"]
        local_time = datetime.utcfromtimestamp(timeb).strftime("%d-%m-%Y %H:%M:%S")
        assitID1 = data_info["assets"][0]["asset_id"]

        if sellers != user and assitID1 != assitID2:
            assitID2 = assitID1
            flippers_List_buy.append(
                [
                    name,
                    number_of_nft,
                    Collection_n,
                    author_n,
                    sellers,
                    fixedC,
                    fixedX,
                    fixedL,
                    fixedF,
                    local_time,
                ]
            )

    flippers_buy = "https://proton.api.atomicassets.io/atomicmarket/v1/sales?state=3&buyer={}&before={}&page=1&limit=100&order=desc&sort=created".format(
        user, timeU
    )
    print(flippers_buy)
    flippers_buyt = requests.get(flippers_buy)
    waitUntilReset = int(flippers_buyt.headers["X-RateLimit-Reset"])
    remainderPings = int(flippers_buyt.headers["X-RateLimit-Remaining"])
    if remainderPings < 3:
        wait = waitUntilReset - time.time()
        time.sleep(wait)
    flippers_buy = json.loads(flippers_buyt.text)
    print("getting Buys")


while len(flippers_sell["data"]) != 0:
    for data_info in flippers_sell["data"]:
        fixedC = 0
        fixedX = 0
        fixedL = 0
        fixedF = 0
        Type = data_info["listing_symbol"]
        RoR = float(data_info["assets"][0]["collection"]["market_fee"])
        if Type == "XPR":
            number = data_info["listing_price"]
            fixedX = int(number) / 10000
            fixedX -= RoR * fixedX
        if Type == "XUSDC":
            number = data_info["listing_price"]
            fixedC = int(number) / 1000000
            fixedC -= RoR * fixedC
        if Type == "LOAN":
            number = data_info["listing_price"]
            fixedL = int(number) / 10000
            fixedL -= RoR * fixedL
        if Type == "FOOBAR":
            number = data_info["listing_price"]
            fixedF = int(number) / 10000
            fixedF -= RoR * fixedF
        name = data_info["assets"][0]["name"]
        buyers = data_info["buyer"]
        timet = data_info["assets"][0]["transferred_at_time"]
        timef = data_info["updated_at_time"]
        timet = int(timet) / 1000
        number_of_nft = int(data_info["assets"][0]["template_mint"])
        Collection_n = data_info["assets"][0]["collection"]["name"]
        author_n = data_info["assets"][0]["collection"]["author"]
        local_time = datetime.utcfromtimestamp(timet).strftime("%d-%m-%Y %H:%M:%S")
        assitID1 = data_info["assets"][0]["asset_id"]
        if assitID1 == assitID2:
            print(buyPrice)
        if author_n != user and assitID1 != assitID2:
            assitID2 = assitID1
            flippers_List_sell.append(
                [
                    name,
                    number_of_nft,
                    Collection_n,
                    author_n,
                    buyers,
                    fixedC,
                    fixedX,
                    fixedL,
                    fixedF,
                    local_time,
                ]
            )
    flippers_sell = "https://proton.api.atomicassets.io/atomicmarket/v1/sales?state=3&seller={}&before={}&page=1&limit=100&order=desc&sort=updated".format(
        user, timef
    )
    flippers_sell = requests.get(flippers_sell).text
    waitUntilReset = int(flippers_buyt.headers["X-RateLimit-Reset"])
    remainderPings = int(flippers_buyt.headers["X-RateLimit-Remaining"])
    if remainderPings < 3:
        wait = waitUntilReset - time.time()
        time.sleep(wait)
    flippers_sell = json.loads(flippers_sell)
    print("getting Resells")

while len(flippers_buy_auction["data"]) != 0:
    for data_info in flippers_buy_auction["data"]:
        fixedC = 0
        fixedX = 0
        fixedL = 0
        fixedF = 0

        Type = data_info["price"]["token_symbol"]
        if Type == "XPR":
            number = data_info["price"]["amount"]
            fixedX = int(number) / 10000
        if Type == "XUSDC":
            number = data_info["price"]["amount"]
            fixedC = int(number) / 1000000
        if Type == "LOAN":
            number = data_info["price"]["amount"]
            fixedL = int(number) / 10000
        if Type == "FOOBAR":
            number = data_info["listing_price"]
            fixedF = int(number) / 10000
            fixedF -= RoR * fixedF
        name = data_info["assets"][0]["name"]
        timez = data_info["updated_at_time"]
        timeMs = data_info["created_at_time"]
        timeSec = int(timez) / 1000
        number_of_nft = int(data_info["assets"][0]["template_mint"])
        Collection_n = data_info["assets"][0]["collection"]["name"]
        author_n = data_info["assets"][0]["collection"]["author"]
        buyer = data_info["buyer"]
        seller = data_info["seller"]
        local_time = datetime.utcfromtimestamp(timeSec).strftime("%m-%d-%Y %H:%M:%S")

        flippers_List_buy.append(
            [
                name,
                number_of_nft,
                Collection_n,
                author_n,
                seller,
                fixedC,
                fixedX,
                fixedL,
                fixedF,
                local_time,
            ]
        )

    flippers_buy_auction = (
        "https://proton.api.atomicassets.io/atomicmarket/v1/auctions?state=3&seller={}"
        "&before={}&page=1&limit=100&order=desc&sort=created".format(user, timeMs)
    )
    flippers_buy_auction = requests.get(flippers_buy_auction).text
    waitUntilReset = int(flippers_buyt.headers["X-RateLimit-Reset"])
    remainderPings = int(flippers_buyt.headers["X-RateLimit-Remaining"])
    if remainderPings < 3:
        wait = waitUntilReset - time.time()
        time.sleep(wait)
    flippers_buy_auction = json.loads(flippers_buy_auction)
    print("getting buy auctions")

while len(flippers_sell_auction["data"]) != 0:
    for data_info in flippers_sell_auction["data"]:
        fixedC = 0
        fixedX = 0
        fixedL = 0
        fixedF = 0

        Type = data_info["price"]["token_symbol"]
        if Type == "XPR":
            number = data_info["price"]["amount"]
            fixedX = int(number) / 10000
        if Type == "XUSDC":
            number = data_info["price"]["amount"]
            fixedC = int(number) / 1000000
        if Type == "LOAN":
            number = data_info["price"]["amount"]
            fixedL = int(number) / 10000
        if Type == "FOOBAR":
            number = data_info["listing_price"]
            fixedF = int(number) / 10000
            fixedF -= RoR * fixedF
        name = data_info["assets"][0]["name"]
        timez = data_info["assets"][0]["transferred_at_time"]
        timeMs = data_info["created_at_time"]
        timeSec = int(timeMs) / 1000
        number_of_nft = int(data_info["assets"][0]["template_mint"])
        Collection_n = data_info["assets"][0]["collection"]["name"]
        author_n = data_info["assets"][0]["collection"]["author"]
        buyer = data_info["buyer"]
        seller = data_info["seller"]
        local_time = datetime.utcfromtimestamp(timeSec).strftime("%m-%d-%Y %H:%M:%S")

        flippers_List_sell.append(
            [
                name,
                number_of_nft,
                Collection_n,
                author_n,
                buyers,
                fixedC,
                fixedX,
                fixedL,
                fixedF,
                local_time,
            ]
        )

    flippers_sell_auction = (
        "https://proton.api.atomicassets.io/atomicmarket/v1/auctions?state=3&seller={}"
        "&before={}&page=1&limit=100&order=desc&sort=created".format(user, timeMs)
    )
    flippers_sell_auction = requests.get(flippers_sell_auction).text
    waitUntilReset = int(flippers_buyt.headers["X-RateLimit-Reset"])
    remainderPings = int(flippers_buyt.headers["X-RateLimit-Remaining"])
    if remainderPings < 3:
        wait = waitUntilReset - time.time()
        time.sleep(wait)
    flippers_sell_auction = json.loads(flippers_sell_auction)
    print("getting sell auctions")


while len(flippersSellOffers["data"]) != 0:
    for data_info in flippersSellOffers["data"]:
        fixedC = 0
        fixedX = 0
        fixedL = 0
        fixedF = 0

        Type = data_info["price"]["token_symbol"]
        if Type == "XPR":
            number = data_info["price"]["amount"]
            fixedX = int(number) / 10000
        if Type == "XUSDC":
            number = data_info["price"]["amount"]
            fixedC = int(number) / 1000000
        if Type == "LOAN":
            number = data_info["price"]["amount"]
            fixedL = int(number) / 10000
        if Type == "FOOBAR":
            number = data_info["listing_price"]
            fixedF = int(number) / 10000
            fixedF -= RoR * fixedF
        name = data_info["assets"][0]["name"]
        timez = data_info["assets"][0]["transferred_at_time"]
        timeMs = data_info["created_at_time"]
        timeSec = int(timeMs) / 1000
        number_of_nft = int(data_info["assets"][0]["template_mint"])
        Collection_n = data_info["assets"][0]["collection"]["name"]
        author_n = data_info["assets"][0]["collection"]["author"]
        buyer = data_info["buyer"]
        seller = data_info["seller"]
        local_time = datetime.utcfromtimestamp(timeSec).strftime("%m-%d-%Y %H:%M:%S")

        flippers_List_sell.append(
            [
                name,
                number_of_nft,
                Collection_n,
                author_n,
                buyers,
                fixedC,
                fixedX,
                fixedL,
                fixedF,
                local_time,
            ]
        )

    flippersSellOffers = "https://proton.api.atomicassets.io/atomicmarket/v1/buyoffers?state=3&seller={}&before={}&page=1&limit=100&order=desc&sort=created".format(
        user, timeMs
    )
    flippersSellOffers = requests.get(flippersSellOffers).text
    waitUntilReset = int(flippers_buyt.headers["X-RateLimit-Reset"])
    remainderPings = int(flippers_buyt.headers["X-RateLimit-Remaining"])
    if remainderPings < 3:
        wait = waitUntilReset - time.time()
        time.sleep(wait)
    flippersSellOffers = json.loads(flippersSellOffers)
    print("getting sell offers " + str(timeMs))


while len(flippersBuyOffers["data"]) != 0:
    for data_info in flippersBuyOffers["data"]:
        fixedC = 0
        fixedX = 0
        fixedL = 0
        fixedF = 0

        Type = data_info["price"]["token_symbol"]
        if Type == "XPR":
            number = data_info["price"]["amount"]
            fixedX = int(number) / 10000
        if Type == "XUSDC":
            number = data_info["price"]["amount"]
            fixedC = int(number) / 1000000
        if Type == "LOAN":
            number = data_info["price"]["amount"]
            fixedL = int(number) / 10000
        if Type == "FOOBAR":
            number = data_info["listing_price"]
            fixedF = int(number) / 10000

        name = data_info["assets"][0]["name"]
        timez = data_info["updated_at_time"]
        timeMs = data_info["created_at_time"]
        timeSec = int(timez) / 1000
        number_of_nft = int(data_info["assets"][0]["template_mint"])
        Collection_n = data_info["collection"]["name"]
        author_n = data_info["collection"]["author"]
        buyer = data_info["buyer"]
        sellers = data_info["seller"]
        local_time = datetime.utcfromtimestamp(timeSec).strftime("%m-%d-%Y %H:%M:%S")
        flippers_List_buy.append(
            [
                name,
                number_of_nft,
                Collection_n,
                author_n,
                sellers,
                fixedC,
                fixedX,
                fixedL,
                fixedF,
                local_time,
            ]
        )

    flippersBuyOffers = "https://proton.api.atomicassets.io/atomicmarket/v1/buyoffers?state=3&buyer={}&before={}&page=1&limit=100&order=desc&sort=created".format(
        user, timeMs
    )
    flippersBuyOffers = requests.get(flippersBuyOffers).text
    waitUntilReset = int(flippers_buyt.headers["X-RateLimit-Reset"])
    remainderPings = int(flippers_buyt.headers["X-RateLimit-Remaining"])
    if remainderPings < 3:
        wait = waitUntilReset - time.time()
        time.sleep(wait)
    flippersBuyOffers = json.loads(flippersBuyOffers)
    print("getting buy offers")

buy = pd.DataFrame(
    data=flippers_List_buy,
    columns=[
        "name of nft",
        "# of nft",
        "collection",
        "author",
        "bought from",
        "bought for usdc",
        "bought for Xpr",
        "bought for loan",
        "bought for foobar",
        "time",
    ],
)

buy.set_index("name of nft")
paidxpr = buy["bought for Xpr"].sum()
paidusd = buy["bought for usdc"].sum()
paidloan = buy["bought for loan"].sum()
paidfoob = buy["bought for foobar"].sum()
sells = pd.DataFrame(
    data=flippers_List_sell,
    columns=(
        [
            "name of nft",
            "number of nft",
            "collection",
            "author",
            "sold to",
            "sold for xusd",
            "sold for xpr",
            "sold for loan",
            "sold for foobar",
            "time",
        ]
    ),
)
sells.set_index("name of nft")

paidc = sells["sold for xusd"].sum()
paidx = sells["sold for xpr"].sum()
paidl = sells["sold for loan"].sum()
paidf = sells["sold for foobar"].sum()  # TODO sum the other coins the same way
sold = sells["sold to"].sum()
buy.to_pickle("./dummy.pkl")
sells.to_pickle("./dummy2.pkl")
mergedDf = pd.merge(buy, sells, how="inner", on="name of nft", suffixes=("", "_drop"))
mergedDf.drop([col for col in mergedDf.columns if "drop" in col], axis=1, inplace=True)
mergedDf["profits"] = mergedDf["sold for xusd"] - mergedDf["bought for usdc"]
profits = mergedDf["profits"].sum()
mbuys = mergedDf["bought for usdc"].sum()
mxbuys = mergedDf["bought for Xpr"].sum()
mlbuys = mergedDf["bought for loan"].sum()
mfbuys = mergedDf["bought for foobar"].sum()
msells = mergedDf["sold to"].sum()
# sold = sells['sold for'].sum()
wb = Workbook()
# name_df.add(names_df)
# name_df=pd.merge([name_df, names_df])
ws = wb.active
ws2 = wb.create_sheet()
ws3 = wb.create_sheet()
ws.title = "{} Buys".format(user)
for r in dataframe_to_rows(buy, index=False):
    ws.append(r)
dims = {}
for row in ws.rows:
    for cell in row:
        if cell.value:
            dims[cell.column_letter] = max(
                (dims.get(cell.column_letter, 0), len(str(cell.value)))
            )
for col, value in dims.items():
    ws.column_dimensions[col].width = value
maxrow = ws.max_row + 2
ws.cell(row=maxrow, column=6, value=paidusd)  # TODO add in more here
ws.cell(row=maxrow, column=7, value=paidx)
ws.cell(row=maxrow, column=8, value=paidl)
ws.cell(row=maxrow, column=9, value=paidf)  # TODO add in more here
ws.cell(row=maxrow, column=1, value="totals")
ws2.title = "{} Sells".format(user)
for r in dataframe_to_rows(sells, index=False):
    ws2.append(r)
dims = {}
for row in ws2.rows:
    for cell in row:
        if cell.value:
            dims[cell.column_letter] = max(
                (dims.get(cell.column_letter, 0), len(str(cell.value)))
            )
for col, value in dims.items():
    ws2.column_dimensions[col].width = value
maxrow = ws2.max_row + 2
ws2.cell(row=maxrow, column=6, value=sold)  # TODO add in more here
ws2.cell(row=maxrow, column=1, value="totals")
ws3.title = "{} Flipping ".format(user)
for r in dataframe_to_rows(mergedDf, index=False):
    ws3.append(r)
dims = {}
for row in ws3.rows:
    for cell in row:
        if cell.value:
            dims[cell.column_letter] = max(
                (dims.get(cell.column_letter, 0), len(str(cell.value)))
            )
for col, value in dims.items():
    ws3.column_dimensions[col].width = value
maxrow = ws3.max_row + 2
ws3.cell(row=maxrow, column=6, value=mbuys)
ws3.cell(row=maxrow, column=8, value=msells)
# ws3.cell(row=maxrow, column=10, value=profits)
ws3.cell(row=maxrow, column=1, value="totals")

wb.save("{}.xlsx".format(user))
