import numpy as np
import modules.ApiClass as Api
import time
import json
import os

import pandas as pd
import requests as requests
import pathlib
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime

buy = pd.read_pickle("./dummy.pkl")
sells = pd.read_pickle("./dummy2.pkl")
gainedUSD = sells["sold for xusd"].sum()
gainedXPR = sells["sold for xpr"].sum()
gainedLOAN = sells["sold for loan"].sum()
gainedFOOBAR = sells["sold for foobar"].sum()  # TODO sum the other coins the same way
paidxpr = buy["bought for Xpr"].sum()
paidusd = buy["bought for usdc"].sum()
paidloan = buy["bought for loan"].sum()
paidfoob = buy["bought for foobar"].sum()
print(buy)
print(sells)
mergedDf = pd.merge(buy, sells, how="inner", on="name of nft", suffixes=("", "_drop"))
mergedDf.drop([col for col in mergedDf.columns if "drop" in col], axis=1, inplace=True)
mergedDf["profits USDC"] = mergedDf["sold for xusd"] - mergedDf["bought for usdc"]
mergedDf["profits XPR"] = mergedDf["sold for xpr"] - mergedDf["bought for Xpr"]
mergedDf["profits LOAN"] = mergedDf["sold for loan"] - mergedDf["bought for loan"]
mergedDf["profits Foobar"] = mergedDf["sold for foobar"] - mergedDf["bought for foobar"]


profits_USD = mergedDf["profits USDC"].sum()
profits_XPR = mergedDf["profits XPR"].sum()
profits_LOAN = mergedDf["profits LOAN"].sum()
profits_Foobar = mergedDf["profits Foobar"].sum()

mbuys = mergedDf["bought for usdc"].sum()
mxbuys = mergedDf["bought for Xpr"].sum()
mlbuys = mergedDf["bought for loan"].sum()
mfbuys = mergedDf["bought for foobar"].sum()
mergedDf["sold for xusd"].sum()
# sold = sells['sold for'].sum()
mergedDf["sold for xpr"].sum()

mergedDf["sold for loan"].sum()

mergedDf["sold for foobar"].sum()
wb = Workbook()
# name_df.add(names_df)
# name_df=pd.merge([name_df, names_df])
ws = wb.active
ws2 = wb.create_sheet()
ws3 = wb.create_sheet()
ws.title = "{} Buys".format("user")
for r in dataframe_to_rows(buy, index=False):
    ws.append(r)
dims = {}
for row in ws.rows:
    for cell in row:
        if cell.value:
            dims[cell.column_letter] = max(
                (dims.get(cell.column_letter, 0), len(str(cell.value)))
            )
for col, value in dims.items():
    ws.column_dimensions[col].width = value
maxrow = ws.max_row + 2
ws.cell(row=maxrow, column=6, value=paidusd)  # TODO add in more here
ws.cell(row=maxrow, column=7, value=paidxpr)
ws.cell(row=maxrow, column=8, value=paidloan)
ws.cell(row=maxrow, column=9, value=paidfoob)  # TODO add in more here
ws.cell(row=maxrow, column=1, value="totals")
ws2.title = "{} Sells".format("user")
for r in dataframe_to_rows(sells, index=False):
    ws2.append(r)
dims = {}
for row in ws2.rows:
    for cell in row:
        if cell.value:
            dims[cell.column_letter] = max(
                (dims.get(cell.column_letter, 0), len(str(cell.value)))
            )
for col, value in dims.items():
    ws2.column_dimensions[col].width = value
maxrow = ws2.max_row + 2
ws2.cell(row=maxrow, column=6, value=gainedUSD)  # TODO add in more here
ws2.cell(row=maxrow, column=1, value="totals")
ws3.title = "{} Flipping ".format("user")
for r in dataframe_to_rows(mergedDf, index=False):
    ws3.append(r)
dims = {}
for row in ws3.rows:
    for cell in row:
        if cell.value:
            dims[cell.column_letter] = max(
                (dims.get(cell.column_letter, 0), len(str(cell.value)))
            )
for col, value in dims.items():
    ws3.column_dimensions[col].width = value
maxrow = ws3.max_row + 2
ws3.cell(row=maxrow, column=6, value=mbuys)

# ws3.cell(row=maxrow, column=10, value=profits)
ws3.cell(row=maxrow, column=1, value="totals")

wb.save("{}.xlsx".format("user"))
