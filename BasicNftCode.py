
import json
import os
import pathlib
import time
from datetime import datetime
from modules.ApiClass import ApiAuthor
import modules.SalesClass as SalesClass
import pandas as pd
import requests as requests
from openpyxl import Workbook
from openpyxl.styles import Color, Font, colors
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet import hyperlink

print(
    "Thank you for running the programme\nIt will take a few seconds to create everything\nHope you enjoy  "
    "\n-kevcollector\n\n\n\n")

# set what command you want to run here
# os.getenv('PWD')
# mac=pathlib.Path().cwd() /'Desktop'
# os.chdir(mac)

current = pathlib.Path().cwd()


def collection(author, collection_name, heading, *excelsheetname):
    collection = "".join(excelsheetname)
    collection = collection.replace('.xlsx', '')

    path = pathlib.Path().cwd() / ("{}".format(heading))
    if (current != pathlib.Path().cwd()):
        path = pathlib.Path().cwd()
        pathlib.Path(path).mkdir(parents=True, exist_ok=True)
    else:
        pathlib.Path(path).mkdir(parents=True, exist_ok=True)
    os.chdir(path)

    authors = ApiAuthor(author, collection_name)
    sales = SalesClass.Sales(authors.authors_)
    while len(authors.authors_["data"]) != 0:
        sales.add(authors.authors_)
        authors.update(sales.timeMs)
    auctions = (
        "https://proton.api.atomicassets.io/atomicmarket/v1/auctions?state=3&seller={}&collection_name={}&page=1&limit=100&order=desc&sort=created".format(author,                                                                                                                                                           collection_name))
    auctions = requests.get(auctions).text
    auctions_ = json.loads(auctions)
    auctions_list = []
    total = 0
    for data_info in auctions_['data']:
        number = data_info["price"]["amount"]
        fixed = int(number) / 1000000
        name = data_info['assets'][0]['name']
        timez = data_info['assets'][0]['transferred_at_time']
        timef = data_info['updated_at_time']
        timex = int(timez) / 1000
        number_of_nft = int(data_info['assets'][0]['template_mint'])
        buyer = data_info['buyer']
        seller = data_info['seller']

        local_time = datetime.utcfromtimestamp(
            timex).strftime('%d-%m-%Y %H:%M:%S')

        auctions_list.append(
            [seller, fixed, sales.fee, buyer, number_of_nft, name, local_time])

    auctions_df = pd.DataFrame(data=auctions_list, columns=[
                               "author ", "price listed usd", "Fee's", "buyer", "# of nft", "name", "time"])
    name_df = pd.DataFrame(data=sales.lists, columns=["author ", "price listed usd", "price listed usd",
                           "price listed usd", "price listed usd", "Fee's", "buyer", "# of nft", "name", "time"])

    wb = Workbook()
    count = 0
    ws = wb.active
    ws.title = ("FirstSale")
    ws2 = wb.create_sheet("Resales")
    ws3 = wb.create_sheet("Holders")
    for r in dataframe_to_rows(name_df, index=False):
        ws.append(r)
        count = count+1

    dims = {}
    for row in ws.rows:
        for cell in row:
            if cell.value:
                dims[cell.column_letter] = max(
                    (dims.get(cell.column_letter, 0), len(str(cell.value))))
    for col, value in dims.items():
        ws.column_dimensions[col].width = value
    maxrow = ws.max_row
    ws.cell(row=maxrow+1, column=1, value="First sale")
    ws.cell(row=maxrow + 1, column=3, value="Royalties ")
    ws.cell(row=maxrow+1, column=4).value = total
    ws.cell(row=count + 3, column=1, value="Auctions")
    count = 0
    for r in dataframe_to_rows(auctions_df, index=False):
        ws.append(r)
    for r in dataframe_to_rows(auctions_df, index=False):
        count = count + 1
    for i in range(0, count - 1):
        total = (auctions_df.loc[auctions_df.index[i], "price listed usd"] * auctions_df.loc[
            auctions_df.index[i], "Fee's"]) + total
    for row in ws.rows:
        for cell in row:
            if cell.value:
                dims[cell.column_letter] = max(
                    (dims.get(cell.column_letter, 0), len(str(cell.value))))
    for col, value in dims.items():
        ws.column_dimensions[col].width = value
        maxrow = ws.max_row
    ws.cell(row=maxrow+2, column=1, value="totals")
    asum = auctions_df['price listed usd'].sum()
    ws.cell(row=maxrow + 2, column=3).value = "Royalties"
    ws.cell(row=maxrow + 2, column=4).value = total

    auctions = ("https://proton.api.atomicassets.io/atomicmarket/v1/auctions?state=3&seller_blacklist={}&collection_name={}&page=1&limit=100&order=desc&sort=created".format(author, collection_name))
    auctions = requests.get(auctions).text
    auctions_ = json.loads(auctions)
    print(auctions_)
    auctions_list = []

    print("getting resales")
    total = 0
    while len(['data']) != 0:

        auctions_df.iloc[0:0]

        for data_info in auctions_['data']:
            number = data_info["price"]["amount"]
            fixed = int(number) / 1000000
            name = data_info['assets'][0]['name']
            timez = data_info['assets'][0]['transferred_at_time']
            timef = data_info['updated_at_time']
            number_of_nft = int(data_info['assets'][0]['template_mint'])
            Royal = data_info['collection']['market_fee']
            timefixe = int(timez) / 1000
            buyer = data_info['buyer']
            seller = data_info['seller']

            local_time = datetime.utcfromtimestamp(
                timefixe).strftime('%d-%m-%Y %H:%M:%S')
            auctions_list.append(
                [seller, buyer, fixed, Royal, name, number_of_nft, local_time])

    auctions_df = pd.DataFrame(data=auctions_list,
                               columns=["first buyer ", "next buyer", "price paid usd", "Fee's", "name", "# of nft", "time"])
    names_df = pd.DataFrame(data=resales.list,
                            columns=["first buyer ", "next buyer", "price paid usd", "Fee's", "name", "# of nft", "time"])
    names_df.drop(names_df[names_df["first buyer "]
                  == f"{author}"].index, inplace=True)
    print()
    count = 0
    total = 0
    for r in dataframe_to_rows(names_df, index=False):
        count = count + 1
    for i in range(0, count-1):
        total = (names_df.loc[names_df.index[i], "price paid usd"]
                 * names_df.loc[names_df.index[i], "Fee's"])+total

    Royalties = total
    Rows = int(names_df.index.max() + 1)
    names_df.at[Rows, 'price paid usd'] = Royalties
    count = 0
    for r in dataframe_to_rows(names_df, index=False):
        ws2.append(r)
        count = count+1

    dims = {}
    for row in ws2.rows:
        for cell in row:
            if cell.value:
                dims[cell.column_letter] = max(
                    (dims.get(cell.column_letter, 0), len(str(cell.value))))
    for col, value in dims.items():
        ws2.column_dimensions[col].width = value
    maxrow = ws2.max_row
    ws2.cell(row=maxrow, column=1, value="Royalties")
    temp = auctions_df["price paid usd"].sum()
    temp = temp*Royal
    Royalties = Royalties+temp
    ws2.cell(row=count + 3, column=1, value="Auctions")
    for r in dataframe_to_rows(auctions_df, index=False):
        ws2.append(r)
    for row in ws2.rows:
        for cell in row:
            if cell.value:
                dims[cell.column_letter] = max(
                    (dims.get(cell.column_letter, 0), len(str(cell.value))))
    for col, value in dims.items():
        ws2.column_dimensions[col].width = value
        maxrow = ws2.max_row

    ws2.cell(row=maxrow+2, column=1).value = "Total Royalties"

    ws2.cell(row=maxrow + 2, column=3, value=Royalties)

    for data_info in holders_['data']:
        holders = int(data_info['assets'])
        holder_list.append([data_info['account'], holders])
    dino_holder_df = pd.DataFrame(data=holder_list, columns=[
                                  "account ", "amount held"])
    # print(dino_holder_df)
    len(dino_holder_df) - 1
    count = 0
    rowz = 1
    Peoplelist = []
    for r in dataframe_to_rows(dino_holder_df, index=False):
        ws3.append(r)
    dims = {}

    assitID2 = " "

    for data_info in holders_['data']:
        checker = (data_info['account'])
        print("getting {}'s data".format(checker))
        people = requests.get(
            "https://proton.api.atomicassets.io/atomicmarket/v1/assets?collection_name={}&owner={}&page=1&limit=100&order=desc&sort=asset_id".format(
                collection_name, checker)).text
        people_ = json.loads(people)
        time.sleep(0.6)
        count = 0
        com = 0
        rare = 0
        epic = 0
        rowz = rowz + 1
        for data_info in people_["data"]:
            nft_NAR = (data_info["data"]["name"])
            number_of_nft = data_info['template_mint']
            assitID1 = data_info["asset_id"]
            if assitID1 != assitID2:
                assitID2 = assitID1
                nft_NAR1 = nft_NAR.replace('(', '')
                nft_NAR2 = nft_NAR1.replace(')', '')
                done = nft_NAR2
                count = count + 1
                ws3.cell(row=rowz, column=3 + count).value = done + \
                    ' (#'+number_of_nft+')'

    for row in ws3.rows:
        for cell in row:
            if cell.value:
                dims[cell.column_letter] = max(
                    (dims.get(cell.column_letter, 0), len(str(cell.value))))
    for col, value in dims.items():
        ws3.column_dimensions[col].width = value + 5

    excelsave = "".join(excelsheetname)
    wb.save(excelsave)
    print("Creating the excel file")
    wb.close()
    os.chdir(path.parent.absolute())


# collection(author,collection_name,heading,excelsheetname1)


author = 'yfnc'
universe = 'Young Fennecs'
collection1 = 'Young Fennecs'

heading = "{} Collection".format(universe)
collection_name1 = '452134422111'  # Dino the dragon


excelsheetname1 = "{}.xlsx".format(collection1)

collection(author, collection_name1, heading, excelsheetname1)
