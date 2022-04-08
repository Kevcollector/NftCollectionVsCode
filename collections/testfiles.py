from openpyxl import load_workbook
import requests
import json

wb = load_workbook(filename="saved.xlsx")
ws3 = wb["holders"]
qua = 0
ws3.cell(row=1, column=4).value = "USDC"
usd_user = requests.get(
    "https://proton.cryptolions.io/v2/state/get_tokens?limit=1000&account=pimlrp"
).text
usd_user = json.loads(usd_user)
for x in usd_user["tokens"]:
    if x["symbol"] == "XUSDC":
        USDC = x["amount"]
print(USDC)
upper5 = 0.035 * USDC
num6_20 = 0.025 * USDC
num21_40 = 0.0125 * USDC

for r in range(2, ws3.max_row):
    if r > 40 and int(ws3.cell(row=r, column=3).value) >= 100:
        qua += 1
others = USDC * (0.1 / qua)
print(qua)
for r in range(2, ws3.max_row):
    if r < 7:
        ws3.cell(row=r, column=4).value = upper5
    if r < 21 and r > 6:
        ws3.cell(row=r, column=4).value = num6_20
    if r > 20 and r < 42:
        ws3.cell(row=r, column=4).value = num21_40
    if r > 41 and r <= 40 + qua:
        ws3.cell(row=r, column=4).value = others
wb.save("temp.xlsx")
