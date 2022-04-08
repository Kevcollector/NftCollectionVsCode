import json
import os
import pandas as pd
import requests as requests
import pathlib
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import time

print(
    "Thank you for running the programe\nIt will take a few seconds to create everything\nHope you enjoy  \n-kevcollector"
)
# set what command you want to run here
# os.getenv('PWD')
# mac=pathlib.Path().cwd() /'Desktop'
# os.chdir(mac)
os.system("cls" if os.name == "nt" else "clear")
userMonth = input("Please enter the month for the rewards 1-12")
os.system("cls" if os.name == "nt" else "clear")
userYear = input("Please enter the Year for the rewards. eg 2022")
os.system("cls" if os.name == "nt" else "clear")


def writeToExcel(worksheet, data):
    for r in dataframe_to_rows(data, index=False):
        worksheet.append(r)
    dims = {}
    for row in worksheet.rows:
        for cell in row:
            if cell.value:
                dims[cell.column_letter] = max(
                    (dims.get(cell.column_letter, 0), len(str(cell.value)))
                )
    for col, value in dims.items():
        worksheet.column_dimensions[col].width = value


current = pathlib.Path().cwd()
Royal = 0.1
wb2 = Workbook()

totalBuyslist = []
totalResellslist = []
totalholderslist = []


def collection(author, collection_name, heading, userMonth, userYear, *excelsheetname):
    global Royal
    global total_df
    collecion = "".join(excelsheetname)
    collecion = collecion.replace(".xlsx", "")

    path = pathlib.Path().cwd() / ("{}".format(heading))
    if current != pathlib.Path().cwd():
        path = pathlib.Path().cwd()
        pathlib.Path(path).mkdir(parents=True, exist_ok=True)
    else:
        pathlib.Path(path).mkdir(parents=True, exist_ok=True)

    authors = " "
    all = " "
    holders = " "
    resales = "Resells"
    FirstSale = "First sale"
    Holders = "Holders"

    holder_list = []
    authors_list = []
    all_list = []
    parents_list = []

    def normalServic(
        authors, all, resales, FirstSale, Holders, holders, *excelsheetname
    ):
        global Royal
        global totalBuyslist
        global totalResellslist
        global totalholderslist
        os.chdir(path)
        authors = (
            "https://proton.api.atomicassets.io/atomicmarket/v1/sales?state=3&account={}"
            "&collection_name={}&page=1&limit=100&order=desc&sort=updated".format(
                author, collection_name
            )
        )
        authors = requests.get(authors).text
        authors_ = json.loads(authors)
        print("getting First sales")
        while len(authors_["data"]) != 0:

            for data_info in authors_["data"]:
                fixedC = 0
                fixedX = 0
                fixedL = 0
                Type = data_info["listing_symbol"]
                if Type == "XPR":
                    number = data_info["listing_price"]
                    fixedX = int(number) / 10000

                if Type == "XUSDC":
                    number = data_info["listing_price"]
                    fixedC = int(number) / 1000000

                if Type == "LOAN":
                    number = data_info["listing_price"]
                    fixedL = int(number) / 10000
                number = data_info["listing_price"]
                fixed = int(number) / 1000000
                name = data_info["assets"][0]["name"]
                timez = data_info["assets"][0]["transferred_at_time"]
                timef = data_info["updated_at_time"]
                timex = int(timez) / 1000
                number_of_nft = int(data_info["assets"][0]["template_mint"])
                buyer = data_info["buyer"]
                seller = data_info["seller"]
                local_time = datetime.utcfromtimestamp(timex).strftime(
                    "%d-%m-%Y %H:%M:%S"
                )
                # this is to check if the month is the same as the user entered amount
                timeChecker = datetime.utcfromtimestamp(timex)
                authors_list.append(
                    [
                        seller,
                        fixedC,
                        fixedX,
                        fixedL,
                        buyer,
                        number_of_nft,
                        name,
                        local_time,
                    ]
                )
                if userMonth == str(timeChecker.month) and userYear == str(
                    timeChecker.year
                ):
                    totalBuyslist.append(
                        [
                            seller,
                            fixedC,
                            fixedX,
                            fixedL,
                            buyer,
                            number_of_nft,
                            name,
                            local_time,
                        ]
                    )
            time.sleep(0.2)
            authors = (
                "https://proton.api.atomicassets.io/atomicmarket/v1/sales?state=3&account={}"
                "&collection_name={}&before={}&page=1&limit=100&order=desc&sort=updated".format(
                    author, collection_name, timef
                )
            )
            authors = requests.get(authors).text
            authors_ = json.loads(authors)

        time.sleep(4)

        authorsOffers = (
            "https://proton.api.atomicassets.io/atomicmarket/v1/buyoffers?state=3&seller={}"
            "&collection_name={}&page=1&limit=100&order=desc&sort=created".format(
                author, collection_name
            )
        )
        authorsOffers = requests.get(authorsOffers).text
        authorsOffers_ = json.loads(authorsOffers)

        print("getting sales offers")
        while len(authorsOffers_["data"]) != 0:

            for data_info in authorsOffers_["data"]:
                fixedC = 0
                fixedX = 0
                fixedL = 0
                Type = data_info["price"]["token_symbol"]
                if Type == "XPR":
                    number = data_info["price"]["amount"]
                    fixedX = int(number) / 10000

                if Type == "XUSDC":
                    number = data_info["price"]["amount"]
                    fixedC = int(number) / 1000000

                if Type == "LOAN":
                    number = data_info["price"]["amount"]
                    fixedL = int(number) / 10000
                fixed = int(number) / 1000000
                name = data_info["assets"][0]["name"]
                timez = data_info["assets"][0]["transferred_at_time"]
                timef = data_info["created_at_time"]
                timex = int(timez) / 1000
                number_of_nft = int(data_info["assets"][0]["template_mint"])
                buyer = data_info["buyer"]
                seller = data_info["seller"]
                timeChecker = datetime.utcfromtimestamp(timex)
                local_time = datetime.utcfromtimestamp(timex).strftime(
                    "%d-%m-%Y %H:%M:%S"
                )
                timeChecker = datetime.utcfromtimestamp(
                    timex
                )  # this is to check if the month is the same as the user entered amount
                authors_list.append(
                    [
                        seller,
                        fixedC,
                        fixedX,
                        fixedL,
                        buyer,
                        number_of_nft,
                        name,
                        local_time,
                    ]
                )
                if userMonth == str(timeChecker.month) and userYear == str(
                    timeChecker.year
                ):
                    totalBuyslist.append(
                        [
                            seller,
                            fixedC,
                            fixedX,
                            fixedL,
                            buyer,
                            number_of_nft,
                            name,
                            local_time,
                        ]
                    )
            time.sleep(0.2)
            authorsOffers = (
                "https://proton.api.atomicassets.io/atomicmarket/v1/buyoffers?state=3&seller={}"
                "&collection_name={}&before={}&page=1&limit=100&order=desc&sort=created".format(
                    author, collection_name, timef
                )
            )
            authorsOffers = requests.get(authorsOffers).text
            authorsOffers_ = json.loads(authorsOffers)

        auctions = "https://proton.api.atomicassets.io/atomicmarket/v1/auctions?state=3&seller={}&collection_name={}&page=1&limit=100&order=desc&sort=created".format(
            author, collection_name
        )
        auctions = requests.get(auctions).text
        auctions_ = json.loads(auctions)
        while len(auctions_["data"]) != 0:
            for data_info in auctions_["data"]:
                fixedC = 0
                fixedX = 0
                fixedL = 0
                Type = data_info["price"]["token_symbol"]
                if Type == "XPR":
                    number = data_info["price"]["amount"]
                    fixedX = int(number) / 10000

                if Type == "XUSDC":
                    number = data_info["price"]["amount"]
                    fixedC = int(number) / 1000000

                if Type == "LOAN":
                    number = data_info["price"]["amount"]
                    fixedL = int(number) / 10000

                name = data_info["assets"][0]["name"]
                timez = data_info["assets"][0]["transferred_at_time"]
                timeMs = data_info["created_at_time"]
                timeSec = int(timeMs) / 1000
                number_of_nft = int(data_info["assets"][0]["template_mint"])
                buyer = data_info["buyer"]
                seller = data_info["seller"]
                timeChecker = datetime.utcfromtimestamp((timef / 1000))
                local_time = datetime.utcfromtimestamp(timeSec).strftime(
                    "%m-%d-%Y %H:%M:%S"
                )

                authors_list.append(
                    [
                        seller,
                        fixedC,
                        fixedX,
                        fixedL,
                        buyer,
                        number_of_nft,
                        name,
                        local_time,
                    ]
                )
                if userMonth == str(timeChecker.month) and userYear == str(
                    timeChecker.year
                ):
                    totalBuyslist.append(
                        [
                            seller,
                            buyer,
                            fixedC,
                            fixedX,
                            fixedL,
                            name,
                            number_of_nft,
                            local_time,
                        ]
                    )

            auctions = (
                "https://proton.api.atomicassets.io/atomicmarket/v1/auctions?state=3&seller={}&collection_name={}"
                "&before={}&page=1&limit=100&order=desc&sort=created".format(
                    author, collection_name, timeMs
                )
            )
            auctions = requests.get(auctions).text
            auctions_ = json.loads(auctions)

        name_df = pd.DataFrame(
            data=authors_list,
            columns=[
                "author ",
                "price listed usd",
                "price listed xpr",
                "price listed loan",
                "buyer",
                "# of nft",
                "name",
                "time",
            ],
        )
        total = name_df["price listed usd"].sum()
        name_df.at["Total", "price listed usd"] = name_df["price listed usd"].sum()

        wb = Workbook()

        ws = wb.active
        ws.title = FirstSale
        ws2 = wb.create_sheet(resales)
        ws3 = wb.create_sheet(Holders)
        for r in dataframe_to_rows(name_df, index=False):
            ws.append(r)
        dims = {}
        for row in ws.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column_letter] = max(
                        (dims.get(cell.column_letter, 0), len(str(cell.value)))
                    )
        for col, value in dims.items():
            ws.column_dimensions[col].width = value
        maxrow = ws.max_row
        ws.cell(row=maxrow, column=1, value="totals")

        resales = (
            "https://proton.api.atomicassets.io/atomicmarket/v1/sales?state=1%2C3&seller_blacklist"
            "={}&buyer_blacklist={}&collection_name={}&page=1&limit=100&order"
            "=desc&sort=updated".format(author, author, collection_name)
        )
        resales = requests.get(resales).text
        resales_ = json.loads(resales)
        print("getting resales")
        while len(resales_["data"]) != 0:
            for data_info in resales_["data"]:
                fixedC = 0
                fixedX = 0
                fixedL = 0
                Type = data_info["listing_symbol"]
                if Type == "XPR":
                    number = data_info["listing_price"]
                    fixedX = int(number) / 10000

                if Type == "XUSDC":
                    number = data_info["listing_price"]
                    fixedC = int(number) / 1000000

                if Type == "LOAN":
                    number = data_info["listing_price"]
                    fixedL = int(number) / 10000

                number = data_info["listing_price"]
                fixed = int(number) / 1000000
                name = data_info["assets"][0]["name"]
                timez = data_info["created_at_time"]
                timef = data_info["updated_at_time"]
                number_of_nft = int(data_info["assets"][0]["template_mint"])
                Royal = data_info["collection"]["market_fee"]
                timefixe = int(timez) / 1000
                buyer = data_info["buyer"]
                seller = data_info["seller"]
                timeChecker = datetime.utcfromtimestamp(timefixe)
                local_time = datetime.utcfromtimestamp(timefixe).strftime(
                    "%d-%m-%Y %H:%M:%S"
                )
                all_list.append(
                    [
                        seller,
                        buyer,
                        fixedC,
                        fixedX,
                        fixedL,
                        name,
                        number_of_nft,
                        local_time,
                    ]
                )
                if userMonth == str(timeChecker.month) and userYear == str(
                    timeChecker.year
                ):
                    totalResellslist.append(
                        [
                            seller,
                            buyer,
                            fixedC,
                            fixedX,
                            fixedL,
                            name,
                            number_of_nft,
                            local_time,
                        ]
                    )
            time.sleep(2)
            resales = (
                "https://proton.api.atomicassets.io/atomicmarket/v1/sales?state=1%2C3&seller_blacklist"
                "={}&buyer_blacklist={}&collection_name={}&before={}&page=1&limit=100&order"
                "=desc&sort=updated".format(author, author, collection_name, timef)
            )
            resales = requests.get(resales).text
            resales_ = json.loads(resales)
        time.sleep(4)
        resellOffer = (
            "https://proton.api.atomicassets.io/atomicmarket/v1/buyoffers?state=3&seller_blacklist={}"
            "&collection_name={}&page=1&limit=100&order=desc&sort=created".format(
                author, collection_name
            )
        )
        resellOffer = requests.get(resellOffer).text
        resellOffer_ = json.loads(resellOffer)
        print("getting resales offers")
        while len(resellOffer_["data"]) != 0:

            for data_info in resellOffer_["data"]:
                fixedC = 0
                fixedX = 0
                fixedL = 0
                Type = data_info["price"]["token_symbol"]
                if Type == "XPR":
                    number = data_info["price"]["amount"]
                    fixedX = int(number) / 10000

                if Type == "XUSDC":
                    number = data_info["price"]["amount"]
                    fixedC = int(number) / 1000000

                if Type == "LOAN":
                    number = data_info["price"]["amount"]
                    fixedL = int(number) / 10000

                fixed = int(number) / 1000000
                name = data_info["assets"][0]["name"]
                timez = data_info["assets"][0]["transferred_at_time"]
                timef = data_info["created_at_time"]
                timex = int(timef) / 1000
                number_of_nft = int(data_info["assets"][0]["template_mint"])
                buyer = data_info["buyer"]
                seller = data_info["seller"]
                timeChecker = datetime.utcfromtimestamp(timex)
                local_time = datetime.utcfromtimestamp(timex).strftime(
                    "%d-%m-%Y %H:%M:%S"
                )
                all_list.append(
                    [
                        seller,
                        buyer,
                        fixedC,
                        fixedX,
                        fixedL,
                        name,
                        number_of_nft,
                        local_time,
                    ]
                )
                if userMonth == str(timeChecker.month) and userYear == str(
                    timeChecker.year
                ):
                    totalResellslist.append(
                        [
                            seller,
                            buyer,
                            fixedC,
                            fixedX,
                            fixedL,
                            name,
                            number_of_nft,
                            local_time,
                        ]
                    )
            time.sleep(0.2)
            resellOffer = (
                "https://proton.api.atomicassets.io/atomicmarket/v1/buyoffers?state=3&seller_blacklist={}"
                "&collection_name={}&before={}&page=1&limit=100&order=desc&sort=created".format(
                    author, collection_name, timef
                )
            )
            resellOffer = requests.get(resellOffer).text
            resellOffer_ = json.loads(resellOffer)

        auctionResale = "https://proton.api.atomicassets.io/atomicmarket/v1/auctions?state=3&seller_blacklist={}&collection_name={}&page=1&limit=100&order=desc&sort=created".format(
            author, collection_name
        )
        auctionResale = requests.get(auctionResale).text
        auctionResale_ = json.loads(auctionResale)
        print("getting auction resales")
        while len(auctionResale_["data"]) != 0:

            for data_info in auctionResale_["data"]:
                fixedC = 0
                fixedX = 0
                fixedL = 0
                Type = data_info["price"]["token_symbol"]
                if Type == "XPR":
                    number = data_info["price"]["amount"]
                    fixedX = int(number) / 10000

                if Type == "XUSDC":
                    number = data_info["price"]["amount"]
                    fixedC = int(number) / 1000000

                if Type == "LOAN":
                    number = data_info["price"]["amount"]
                    fixedL = int(number) / 10000
                name = data_info["assets"][0]["name"]
                timez = data_info["assets"][0]["transferred_at_time"]
                timeMs = data_info["updated_at_time"]
                number_of_nft = int(data_info["assets"][0]["template_mint"])
                timeSec = int(timeMs) / 1000
                buyer = data_info["buyer"]
                seller = data_info["seller"]
                timeChecker = datetime.utcfromtimestamp(timeSec)
                local_time = datetime.utcfromtimestamp(timeSec).strftime(
                    "%d-%m-%Y %H:%M:%S"
                )
                all_list.append(
                    [
                        seller,
                        buyer,
                        fixedC,
                        fixedX,
                        fixedL,
                        name,
                        number_of_nft,
                        local_time,
                    ]
                )
                if userMonth == str(timeChecker.month) and userYear == str(
                    timeChecker.year
                ):
                    totalResellslist.append(
                        [
                            seller,
                            buyer,
                            fixedC,
                            fixedX,
                            fixedL,
                            name,
                            number_of_nft,
                            local_time,
                        ]
                    )
            time.sleep(0.6)
            auctionResale = (
                "https://proton.api.atomicassets.io/atomicmarket/v1/auctions?state=3&seller_blacklist={}&collection_name={}"
                "&before={}&page=1&limit=100&order=desc&sort=updated".format(
                    author, collection_name, timeMs
                )
            )
            auctionResale = requests.get(auctionResale).text
            auctionResale_ = json.loads(auctionResale)
        names_df = pd.DataFrame(
            data=all_list,
            columns=[
                "first buyer ",
                "next buyer",
                "price paid usd",
                "price paid xpr",
                "price paid loans",
                "name",
                "# of nft",
                "time",
            ],
        )
        names_df.drop(
            names_df[names_df["first buyer "] == f"{author}"].index, inplace=True
        )
        totals = names_df["price paid usd"].sum()
        Royalties = totals * Royal
        Rows = int(names_df.index.max() + 1)
        names_df.at[Rows, "price paid usd"] = Royalties

        for r in dataframe_to_rows(names_df, index=False):
            ws2.append(r)
        dims = {}
        for row in ws2.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column_letter] = max(
                        (dims.get(cell.column_letter, 0), len(str(cell.value)))
                    )
        for col, value in dims.items():
            ws2.column_dimensions[col].width = value
        maxrow = ws2.max_row
        ws2.cell(row=maxrow, column=1, value="Royalties")
        points = 0
        holders = (
            "https://proton.api.atomicassets.io/atomicassets/v1/accounts?collection_name={}"
            "&page=1&limit=100&order=desc".format(collection_name)
        )
        holders = requests.get(holders).text
        holders_ = json.loads(holders)
        pages = 1
        while len(holders_["data"]) != 0:
            pages = pages + 1
            for data_info in holders_["data"]:
                holders = int(data_info["assets"])
                authorsName = data_info["account"]
                if authorsName != "delaneycb":
                    holder_list.append([data_info["account"], holders])
            holders = (
                "https://proton.api.atomicassets.io/atomicassets/v1/accounts?collection_name={}"
                "&page={}&limit=100&order=desc".format(collection_name, pages)
            )
            holders = requests.get(holders).text
            holders_ = json.loads(holders)
        dino_holder_df = pd.DataFrame(
            data=holder_list, columns=["account ", "amount held"]
        )
        len(dino_holder_df) - 1
        count = 0
        rowz = 1
        Peoplelist = []
        for r in dataframe_to_rows(dino_holder_df, index=False):
            ws3.append(r)
        for r in dataframe_to_rows(dino_holder_df, index=False):
            count = count + 1
            if count == 1:
                """"""
            else:
                ws3.cell(row=count, column=3).value = 0

        temp1 = 0
        temp3 = 0
        temp2 = 0
        totalz = 0
        pages = 0
        holders = (
            "https://proton.api.atomicassets.io/atomicassets/v1/accounts?collection_name={}"
            "&page=1&limit=100&order=desc".format(collection_name)
        )
        holders = requests.get(holders).text
        holders_ = json.loads(holders)
        amount = 1
        skcondition1 = 0
        skcondition2 = 0
        skcondition3 = 0
        skcondition4 = 0
        skcondition5 = 0
        Skirmisher = ""
        Imp = ""
        SE = 0
        while len(holders_["data"]) != 0:
            pages = pages + 1
            amount = amount + 1
            for data_info in holders_["data"]:
                if data_info["account"] != "delaneycb":
                    skcondition1 = 0
                    skcondition2 = 0
                    skcondition3 = 0
                    skcondition4 = 0
                    skcondition5 = 0
                    checker = data_info["account"]
                    if rowz != 1:
                        temp2 = ws3.cell(row=rowz, column=3).value
                    print("value = " + str(temp2))
                    print("getting {}'s data".format(checker))
                    totalz = 0
                    people = "https://proton.api.atomicassets.io/atomicmarket/v1/assets?collection_name={}&owner={}&page=1&limit=100&order=desc&sort=asset_id".format(
                        collection_name, checker
                    )
                    test = requests.get(people)
                    next = test.headers["X-RateLimit-Reset"]
                    resset = test.headers["X-RateLimit-Remaining"]
                    resset = int(resset)
                    next = int(next)
                    wait = next - time.time()
                    if resset < 3:
                        time.sleep(wait)
                    people_ = json.loads((test.text))
                    time.sleep(0.2)
                    count = 0
                    pages = 1
                    rowz = rowz + 1
                    assitID2 = " "
                    ImpCount = 0
                    ImpsDict = {
                        "1": 0,
                        "2": 0,
                        "3": 0,
                        "4": 0,
                        "5": 0,
                        "6": 0,
                        "7": 0,
                        "8": 0,
                        "9": 0,
                        "10": 0,
                        "11": 0,
                        "12": 0,
                    }  # fill in until 12
                    truthDict = {
                        "1": 0,
                        "2": 0,
                        "3": 0,
                        "4": 0,
                        "5": 0,
                        "6": 0,
                        "7": 0,
                        "8": 0,
                        "9": 0,
                        "10": 0,
                        "11": 0,
                        "12": 0,
                    }

                    while len(people_["data"]) != 0:
                        pages = pages + 1
                        if pages == 2:
                            count = 0
                        if pages == 3:
                            count = 100
                        if pages == 4:
                            count = 200
                        for data_info in people_["data"]:
                            word = data_info["data"]["desc"]
                            nft_name = data_info["data"]["name"]
                            number_of_nft = data_info["template_mint"]
                            assitID1 = data_info["asset_id"]
                            if assitID1 != assitID2:
                                assitID2 = assitID1
                                print(ws3.cell(row=rowz, column=3).value)
                                done = nft_name
                                count = count + 1
                                lookup = {
                                    "Skirmisher": 5,
                                    "Imp": 5,
                                    "Mounted": 10,
                                    "Grunt": 15,
                                    "Nimrod": 15,
                                    "Volant": 15,
                                    "Gorgon": 15,
                                    "Gremlin": 15,
                                    "Rarog": 15,
                                    "PecuBeast": 30,
                                    "Battle": 30,
                                    "Chimera": 30,
                                    "Ink": 35,
                                    "Augury Ogre": 40,
                                    "Jinn": 40,
                                    "Augury Troll": 50,
                                    "Cherufe": 50,
                                    "Mage": 60,
                                    "Anomaly": 60,
                                    "Shango": 60,
                                    "Augury Dragon": 75,
                                    "Typhon": 75,
                                    "Panda": 75,
                                    "PANDA": 75,
                                    "Peculiar Soldier": 100,
                                    "Peculiar Lieutenant": 125,
                                    "Peculiar Captain": 150,
                                }
                                ws3.cell(row=1, column=3).value = "Points"
                                temp1 = ws3.cell(row=rowz, column=3).value
                                if (
                                    " SE" in nft_name
                                    or "SE" in nft_name
                                    or "SE " in nft_name
                                    or "Wildcard" in nft_name
                                ):
                                    ws3.cell(row=rowz, column=3 + count).value = (
                                        done + " (#" + number_of_nft + ")"
                                    )
                                else:
                                    if collection_name == "241115151314":
                                        gorilaPoints = 0
                                        word = word.lower()
                                        word = word.split()
                                        placehodle = ""
                                        for i in word:

                                            if i == "common":
                                                gorilaPoints = gorilaPoints + 1
                                            if i == "uncommon":
                                                gorilaPoints = gorilaPoints + 3
                                            if i == "rare" and placehodle != "ultra":
                                                gorilaPoints = gorilaPoints + 5
                                            if i == "ultra" and placehodle != "rare":
                                                gorilaPoints = gorilaPoints + 7
                                            if i == "epic":
                                                gorilaPoints = gorilaPoints + 8
                                            if i == "legendary":
                                                gorilaPoints = gorilaPoints + 10
                                            if i == "unique":
                                                gorilaPoints = gorilaPoints + 12
                                            placehodle = i
                                        print(people)
                                        print(gorilaPoints)
                                        temp1 = ws3.cell(row=rowz, column=3).value
                                        ws3.cell(row=rowz, column=3).value = temp1 + (
                                            gorilaPoints * 4
                                        )

                                    if collection_name == "521533225213":
                                        SE = 0
                                        s = done.split(" ")
                                        for i in s:
                                            if i == "PIM":
                                                holders_amount = ws3.cell(
                                                    row=rowz, column=2
                                                ).value
                                                holders_amount = holders_amount - 1
                                                ws3.cell(
                                                    row=rowz, column=2
                                                ).value = holders_amount
                                            else:
                                                try:
                                                    if i == "SE":
                                                        temp1 = ws3.cell(
                                                            row=rowz, column=3
                                                        ).value
                                                        ws3.cell(
                                                            row=rowz, column=3
                                                        ).value = (int(temp1) - SE)
                                                        print("removed")
                                                    if i == lookup[i]:
                                                        SE = int(lookup[i])
                                                        ws3.cell(
                                                            row=rowz, column=3
                                                        ).value = int(lookup[i]) + int(
                                                            temp1
                                                        )

                                                except KeyError:
                                                    ws3.cell(
                                                        row=rowz, column=3
                                                    ).value = int(0) + int(temp1)

                                        ws3.cell(row=rowz, column=3 + count).value = (
                                            done + " (#" + number_of_nft + ")" + word
                                        )

                                    if (
                                        collection_name == "312124133135"
                                        or collection_name == "451243333513"
                                        or collection_name == "135115145544"
                                        or collection_name == "132423131521"
                                        or collection_name == "534133213533"
                                        or collection_name == "234141453513"
                                    ):
                                        ws3.cell(row=rowz, column=3 + count).value = (
                                            done + " (#" + number_of_nft + ")" + word
                                        )
                                        ws3.cell(row=rowz, column=3).value = int(
                                            10
                                        ) + int(temp1)
                                        s = done.split(" ")
                                        for i in s:
                                            if i == "SE":
                                                temp1 = ws3.cell(
                                                    row=rowz, column=3
                                                ).value
                                                ws3.cell(row=rowz, column=3).value = (
                                                    int(temp1) - 10
                                                )
                                                print("removed")

                                    else:
                                        if collection_name == "241115151314":
                                            ws3.cell(
                                                row=rowz, column=3 + count
                                            ).value = (
                                                done + " (#" + number_of_nft + ")"
                                            )
                                        else:
                                            ws3.cell(
                                                row=rowz, column=3 + count
                                            ).value = (
                                                done
                                                + " (#"
                                                + number_of_nft
                                                + ")"
                                                + word
                                            )
                                        if "Imp " in nft_name:
                                            i = nft_name.replace("Imp ", "")
                                            i = i.replace("#", "")

                                            keys_list = list(ImpsDict)
                                            key = keys_list[0]
                                            # so we want to step through the dict's key to find where its the same as the vaule in newString. when it is we will then plus that keys vaule
                                            for x in ImpsDict:
                                                key = keys_list[int(x) - 1]
                                                if key == i:
                                                    ImpsDict[x] = ImpsDict[x] + 1
                                            print(str(ImpsDict) + " imps")
                                            for x in ImpsDict:
                                                if ImpsDict[x] == 1:
                                                    ImpCount += 1
                                                    if truthDict[x] == 0:
                                                        truthDict[x] += 1
                                                if ImpCount == 5:
                                                    ImpCount = 0
                                                    print(str(truthDict) + " truth")
                                                    print(str(ImpsDict) + " before")
                                                    for i in truthDict:
                                                        if truthDict[i] == 1:
                                                            truthDict[i] = 0
                                                            ImpsDict[i] -= 1
                                                    temp1 = ws3.cell(
                                                        row=rowz, column=3
                                                    ).value
                                                    ws3.cell(
                                                        row=rowz, column=3
                                                    ).value = int(5) + int(temp1)
                                                    temp1 = ws3.cell(
                                                        row=rowz, column=3
                                                    ).value
                                                    print("5 imps!")

                                        s = done.split(" ")

                                        ogreplaceholder = ""
                                        placehodler = ""
                                        for i in s:
                                            if (
                                                i == "Grunt"
                                                or i == "Volant"
                                                or i == "Nimrod"
                                                or i == "PecuBeast"
                                            ):
                                                ogreplaceholder = "true"
                                            try:
                                                if collection_name == "413424453454":
                                                    if i == "Skirmisher":
                                                        Skirmisher = i
                                                    if (
                                                        i != Skirmisher
                                                        and Skirmisher != ""
                                                    ):
                                                        i = i.replace("#", "")
                                                        if i == "1":
                                                            skcondition1 = (
                                                                skcondition1 + 1
                                                            )
                                                        if i == "2":
                                                            skcondition2 = (
                                                                skcondition2 + 1
                                                            )
                                                        if i == "3":
                                                            skcondition3 = (
                                                                skcondition3 + 1
                                                            )
                                                        if i == "4":
                                                            skcondition4 = (
                                                                skcondition4 + 1
                                                            )
                                                        if i == "5":
                                                            skcondition5 = (
                                                                skcondition5 + 1
                                                            )
                                                        if (
                                                            skcondition1 > 0
                                                            and skcondition2 > 0
                                                            and skcondition3 > 0
                                                            and skcondition4 > 0
                                                            and skcondition5 > 0
                                                        ):
                                                            skcondition1 = (
                                                                skcondition1 - 1
                                                            )
                                                            skcondition2 = (
                                                                skcondition2 - 1
                                                            )
                                                            skcondition3 = (
                                                                skcondition3 - 1
                                                            )
                                                            skcondition4 = (
                                                                skcondition4 - 1
                                                            )
                                                            skcondition5 = (
                                                                skcondition5 - 1
                                                            )
                                                            temp1 = ws3.cell(
                                                                row=rowz, column=3
                                                            ).value
                                                            ws3.cell(
                                                                row=rowz, column=3
                                                            ).value = int(5) + int(
                                                                temp1
                                                            )
                                                            temp1 = ws3.cell(
                                                                row=rowz, column=3
                                                            ).value

                                                            print(
                                                                "5 different Skirmisher "
                                                                + " total now = "
                                                                + str(temp1)
                                                            )

                                                            ws3.cell(
                                                                row=rowz, column=3
                                                            ).value = int(5) + int(
                                                                temp1
                                                            )
                                                if i in lookup or placehodler == "":
                                                    if placehodler == "Augury":
                                                        i = placehodler + " " + i
                                                        print(i + " " + str(lookup[i]))
                                                        ws3.cell(
                                                            row=rowz, column=3
                                                        ).value = int(lookup[i]) + int(
                                                            temp1
                                                        )
                                                    if i == "Augury":
                                                        placehodler = i

                                                    else:
                                                        if i == lookup[i]:
                                                            SE = int(lookup[i])
                                                            ws3.cell(
                                                                row=rowz, column=3
                                                            ).value = int(
                                                                lookup[i]
                                                            ) + int(
                                                                temp1
                                                            )
                                                        if i == "SE":
                                                            temp1 = ws3.cell(
                                                                row=rowz, column=3
                                                            ).value
                                                            ws3.cell(
                                                                row=rowz, column=3
                                                            ).value = (int(temp1) - SE)
                                                            print("removed")
                                                        temp1 = ws3.cell(
                                                            row=rowz, column=3
                                                        ).value
                                                        print(
                                                            i
                                                            + " "
                                                            + str(lookup[i])
                                                            + " a"
                                                        )
                                                        ws3.cell(
                                                            row=rowz, column=3
                                                        ).value = int(lookup[i]) + int(
                                                            temp1
                                                        )
                                                else:
                                                    if placehodler == "Augury":
                                                        i = placehodler + " " + i
                                                        print(i + " " + str(lookup[i]))
                                                        ws3.cell(
                                                            row=rowz, column=3
                                                        ).value = int(lookup[i]) + int(
                                                            temp1
                                                        )
                                                    if i == "Augury":
                                                        placehodler = i

                                                    else:
                                                        print(i + " " + str(lookup[i]))
                                                        ws3.cell(
                                                            row=rowz, column=3
                                                        ).value = int(lookup[i]) + int(
                                                            temp1
                                                        )
                                            except KeyError:
                                                if collection_name == "413424453454":
                                                    if (
                                                        i == "Ogre"
                                                        and placehodler == ""
                                                        and ogreplaceholder != "true"
                                                    ):
                                                        i = "Augury" + " " + i
                                                        points = int(lookup[i])
                                                        print(done + str(lookup[i]))
                                                        ws3.cell(
                                                            row=rowz, column=3
                                                        ).value = int(lookup[i]) + int(
                                                            temp1
                                                        )
                                                    if i == "Peculiar" or i == "The":
                                                        placehodler = i
                                                    if i == placehodler:
                                                        if placehodler == "The":
                                                            print(
                                                                placehodler
                                                                + " "
                                                                + i
                                                                + " = "
                                                                + str(225)
                                                            )
                                                            ws3.cell(
                                                                row=rowz, column=3
                                                            ).value = 225 + int(temp1)
                                                        if (
                                                            i != placehodler
                                                            and placehodler
                                                            == "Peculiar"
                                                        ):
                                                            i = i.replace("#", "")
                                                            if int(i) < 31:
                                                                print(200)
                                                                ws3.cell(
                                                                    row=rowz, column=3
                                                                ).value = 200 + int(
                                                                    temp1
                                                                )

                                                        string = word
                                                        nft_NAR1 = string.replace(
                                                            "(", ""
                                                        )
                                                        nft_NAR2 = nft_NAR1.replace(
                                                            ")", ""
                                                        )
                                                        string = nft_NAR2
                                                        string = string.split(" ")

                                                        for i in string:
                                                            if i == "General":
                                                                ws3.cell(
                                                                    row=rowz, column=3
                                                                ).value = 200 + int(
                                                                    temp1
                                                                )
                                                                print(i + str(200))
                                                            elif (
                                                                i == "Soldier"
                                                                or i == "Captain"
                                                                or i == "Lieutenant"
                                                            ):
                                                                i = (
                                                                    placehodler
                                                                    + " "
                                                                    + i
                                                                )
                                                                ws3.cell(
                                                                    row=rowz, column=3
                                                                ).value = int(
                                                                    lookup[i]
                                                                ) + int(
                                                                    temp1
                                                                )
                                                                print(
                                                                    i + str(lookup[i])
                                                                )
                        people = "https://proton.api.atomicassets.io/atomicmarket/v1/assets?collection_name={}&owner={}&page={}&limit=100&order=desc&sort=asset_id".format(
                            collection_name, checker, pages
                        )
                        test = requests.get(people)
                        next = test.headers["X-RateLimit-Reset"]
                        resset = test.headers["X-RateLimit-Remaining"]
                        resset = int(resset)
                        next = int(next)
                        wait = next - time.time()
                        if resset < 3:
                            time.sleep(wait)
                        people_ = json.loads((test.text))
                        time.sleep(0.2)

                    holders_amount = ws3.cell(row=rowz, column=2).value

                    Rpoints = ws3.cell(row=rowz, column=3).value

                    totalholderslist.append([checker, holders_amount, Rpoints])

            holders = (
                "https://proton.api.atomicassets.io/atomicassets/v1/accounts?collection_name={}"
                "&page={}&limit=100&order=desc".format(collection_name, amount)
            )
            holders = requests.get(holders).text
            holders_ = json.loads(holders)

        count = 0

        for row in ws3.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column_letter] = max(
                        (dims.get(cell.column_letter, 0), len(str(cell.value)))
                    )
        for col, value in dims.items():
            ws3.column_dimensions[col].width = value + 5

        excelsave = "".join(excelsheetname)
        wb.save(excelsave)
        print("Creating the excel file")
        wb.close()
        os.chdir(path.parent.absolute())

    normalServic(authors, all, resales, FirstSale, Holders, holders, *excelsheetname)


author = "delaneycb"
universe = "Pecuilar inks & GGIP"
heading = "{} Collection".format(universe)
collection_name = "413424453454"
collection1 = "Pecuilar"
excelsheetname1 = "{}.xlsx".format(collection1)
collection(author, collection_name, heading, userMonth, userYear, excelsheetname1)
collection_name = "521533225213"
collection1 = "Panda Boy Multiverse"
excelsheetname1 = "{}.xlsx".format(collection1)
collection(author, collection_name, heading, userMonth, userYear, excelsheetname1)
collection_name = "312124133135"
collection1 = "Bomboy"
time.sleep(6)
excelsheetname1 = "{}.xlsx".format(collection1)
collection(author, collection_name, heading, userMonth, userYear, excelsheetname1)
collection_name = "451243333513"
collection1 = "Crypto Gorilla"
time.sleep(6)
excelsheetname1 = "{}.xlsx".format(collection1)
collection(author, collection_name, heading, userMonth, userYear, excelsheetname1)
collection_name = "135115145544"
collection1 = "Crypto Panda"
time.sleep(6)
excelsheetname1 = "{}.xlsx".format(collection1)
collection(author, collection_name, heading, userMonth, userYear, excelsheetname1)
collection_name = "132423131521"
collection1 = "Crypto Kevin"
time.sleep(4)
excelsheetname1 = "{}.xlsx".format(collection1)
collection(author, collection_name, heading, userMonth, userYear, excelsheetname1)
collection_name = "534133213533"
collection1 = "Crypto Steve "
excelsheetname1 = "{}.xlsx".format(collection1)
time.sleep(10)
collection(author, collection_name, heading, userMonth, userYear, excelsheetname1)
collection_name = "234141453513"
collection1 = "Crypto owls "
excelsheetname1 = "{}.xlsx".format(collection1)
time.sleep(10)
collection(author, collection_name, heading, userMonth, userYear, excelsheetname1)

buys_df = pd.DataFrame(
    data=totalBuyslist,
    columns=[
        "author ",
        "price listed usd",
        "price listed xpr",
        "price listed loans",
        "buyer",
        "# of nft",
        "name",
        "time",
    ],
)
resale_df = pd.DataFrame(
    data=totalResellslist,
    columns=[
        "first buyer ",
        "next buyer",
        "price listed usd",
        "price listed xpr",
        "price listed loans",
        "name",
        "# of nft",
        "time",
    ],
)
holders_df = pd.DataFrame(
    data=totalholderslist, columns=["account", "amount held", "points"]
)
ws1 = wb2.active
ws1 = wb2.create_sheet("sales")
ws2 = wb2.create_sheet("resales")
ws3 = wb2.create_sheet("holders")
print(totalholderslist)
# holders_df=holders_df.groupby('account','amount held','points').sum().reset_index()
holders_df = holders_df.groupby(["account"]).agg(
    {"account": "min", "amount held": "sum", "points": "sum"}
)
holders_df = holders_df.sort_values(by=["points"], ascending=False)


writeToExcel(ws1, buys_df)
writeToExcel(ws2, resale_df)
writeToExcel(ws3, holders_df)

for r in range(2, ws3.max_row + 2):
    count = ws3.cell(row=r, column=1).value
    holders = (
        "https://proton.api.atomicassets.io/atomicassets/v1/accounts/{}?collection_whitelist="
        "451243333513%2C312124133135%2C135115145544%2C132423131521%2C534133213533%2C413424453454%"
        "2C521533225213%2C234141453513".format(count)
    )
    print(holders)
    holders = requests.get(holders).text
    holders_ = json.loads(holders)
    holders = holders_["data"]["assets"]
    ws3.cell(row=r, column=2).value = holders
    time.sleep(1)
qua = 0
ws3.cell(row=1, column=4).value = "USDC"
usd_user = requests.get(
    "https://proton.cryptolions.io/v2/state/get_tokens?limit=1000&account=pimlrp"
).text
usd_user = json.loads(usd_user)
for x in usd_user["tokens"]:
    if x["symbol"] == "XUSDC":
        USDC = x["amount"]
upper5 = 0.035 * USDC
num6_20 = 0.025 * USDC
num21_40 = 0.0125 * USDC

for r in range(2, ws3.max_row):
    if r > 40 and ws3.cell(row=r, column=3).value >= 100:
        qua += 1
others = USDC * (0.1 / qua)

for r in range(2, ws3.max_row):
    if r < 7:
        ws3.cell(row=r, column=4).value = upper5
    if r < 21 and r > 6:
        ws3.cell(row=r, column=4).value = num6_20
    if r > 21 and r < 42:
        ws3.cell(row=r, column=4).value = num21_40
    if r > 41:
        ws3.cell(row=r, column=4).value = others


wb2.save("saved.xlsx")
print("Creating the excel file")
wb2.close()


author = "ggip"
collection_name = "241115151314"
collection6 = "GGIP"
excelsheetname1 = "{}.xlsx".format(collection6)
time.sleep(6)
collection(author, collection_name, heading, userMonth, userYear, excelsheetname1)
