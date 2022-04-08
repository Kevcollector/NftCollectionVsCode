import json
import os
import pandas as pd
import requests as requests
import pathlib
from openpyxl import Workbook
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import time

print(
    "Thank you for running the programe\nIt will take a few seconds to create everything\nHope you enjoy  \n-kevcollector"
)
# set what command you want to run here
# os.getenv('PWD')
# mac=pathlib.Path().cwd() /'Desktop'
# os.chdir(mac)


def writeToExcel(worksheet, data):
    for r in dataframe_to_rows(data, index=False):
        worksheet.append(r)
    dims = {}
    for row in worksheet.rows:
        for cell in row:
            if cell.value:
                dims[cell.column_letter] = max(
                    (dims.get(cell.column_letter, 0), len(str(cell.value)))
                )
    for col, value in dims.items():
        worksheet.column_dimensions[col].width = value + 20


current = pathlib.Path().cwd()
Royal = 0.1
wb2 = Workbook()

totalBuyslist = []
totalResellslist = []
totalholderslist = []


def collection(auther, collection_name, heading, *excelsheetname):
    global Royal
    global total_df
    collecion = "".join(excelsheetname)
    collecion = collecion.replace(".xlsx", "")

    path = pathlib.Path().cwd() / ("{}".format(heading))
    if current != pathlib.Path().cwd():
        path = pathlib.Path().cwd()
        pathlib.Path(path).mkdir(parents=True, exist_ok=True)
    else:
        pathlib.Path(path).mkdir(parents=True, exist_ok=True)

    authers = " "
    all = " "
    holders = " "
    resales = "Resells"
    FirstSale = "First sale"
    Holders = "Holders"

    holder_list = []
    authers_list = []
    all_list = []
    parents_list = []

    def normalServic(
        authers, all, resales, FirstSale, Holders, holders, *excelsheetname
    ):
        global Royal
        global totalBuyslist
        global totalResellslist
        global totalholderslist
        os.chdir(path)
        authers = (
            "https://proton.api.atomicassets.io/atomicmarket/v1/sales?state=3&seller={}"
            "&collection_name={}&page=1&limit=100&order=desc&sort=updated".format(
                auther, collection_name
            )
        )
        authers = requests.get(authers).text
        authers_ = json.loads(authers)
        print("getting First sales")
        while len(authers_["data"]) != 0:

            for data_info in authers_["data"]:
                fixedC = 0
                fixedX = 0
                fixedL = 0
                Type = data_info["listing_symbol"]
                if Type == "XPR":
                    number = data_info["listing_price"]
                    fixedX = int(number) / 10000

                if Type == "XUSDC":
                    number = data_info["listing_price"]
                    fixedC = int(number) / 1000000

                if Type == "LOAN":
                    number = data_info["listing_price"]
                    fixedL = int(number) / 10000
                number = data_info["listing_price"]
                fixed = int(number) / 1000000
                name = data_info["assets"][0]["name"]
                timez = data_info["assets"][0]["transferred_at_time"]
                timef = data_info["updated_at_time"]
                timex = int(timez) / 1000
                number_of_nft = int(data_info["assets"][0]["template_mint"])
                buyer = data_info["buyer"]
                seller = data_info["seller"]
                # this is to check if the month is the same as the user entered amount
                local_time = datetime.utcfromtimestamp(timex).strftime(
                    "%d-%m-%Y %H:%M:%S"
                )
                authers_list.append(
                    [
                        seller,
                        fixedC,
                        fixedX,
                        fixedL,
                        buyer,
                        number_of_nft,
                        name,
                        local_time,
                    ]
                )

            time.sleep(0.2)
            authers = (
                "https://proton.api.atomicassets.io/atomicmarket/v1/sales?state=3&seller={}"
                "&collection_name={}&before={}&page=1&limit=100&order=desc&sort=updated".format(
                    auther, collection_name, timef
                )
            )
            authers = requests.get(authers).text
            authers_ = json.loads(authers)

        time.sleep(4)

        authersOffers = (
            "https://proton.api.atomicassets.io/atomicmarket/v1/buyoffers?state=3&seller={}"
            "&collection_name={}&page=1&limit=100&order=desc&sort=created".format(
                auther, collection_name
            )
        )
        authersOffers = requests.get(authersOffers).text
        authersOffers_ = json.loads(authersOffers)

        print("getting sales offers")
        while len(authersOffers_["data"]) != 0:

            for data_info in authersOffers_["data"]:
                fixedC = 0
                fixedX = 0
                fixedL = 0
                Type = data_info["price"]["token_symbol"]
                if Type == "XPR":
                    number = data_info["price"]["amount"]
                    fixedX = int(number) / 10000

                if Type == "XUSDC":
                    number = data_info["price"]["amount"]
                    fixedC = int(number) / 1000000

                if Type == "LOAN":
                    number = data_info["price"]["amount"]
                    fixedL = int(number) / 10000
                fixed = int(number) / 1000000
                name = data_info["assets"][0]["name"]
                timez = data_info["assets"][0]["transferred_at_time"]
                timef = data_info["created_at_time"]
                timex = int(timez) / 1000
                number_of_nft = int(data_info["assets"][0]["template_mint"])
                buyer = data_info["buyer"]
                seller = data_info["seller"]
                local_time = datetime.utcfromtimestamp(timex).strftime(
                    "%d-%m-%Y %H:%M:%S"
                )
                authers_list.append(
                    [
                        seller,
                        fixedC,
                        fixedX,
                        fixedL,
                        buyer,
                        number_of_nft,
                        name,
                        local_time,
                    ]
                )

            time.sleep(0.2)
            authersOffers = (
                "https://proton.api.atomicassets.io/atomicmarket/v1/buyoffers?state=3&seller={}"
                "&collection_name={}&before={}&page=1&limit=100&order=desc&sort=created".format(
                    auther, collection_name, timef
                )
            )
            authersOffers = requests.get(authersOffers).text
            authersOffers_ = json.loads(authersOffers)

        auctions = "https://proton.api.atomicassets.io/atomicmarket/v1/auctions?state=3&seller={}&collection_name={}&page=1&limit=100&order=desc&sort=created".format(
            auther, collection_name
        )
        auctions = requests.get(auctions).text
        auctions_ = json.loads(auctions)
        while len(auctions_["data"]) != 0:
            for data_info in auctions_["data"]:
                fixedC = 0
                fixedX = 0
                fixedL = 0
                Type = data_info["price"]["token_symbol"]
                if Type == "XPR":
                    number = data_info["price"]["amount"]
                    fixedX = int(number) / 10000

                if Type == "XUSDC":
                    number = data_info["price"]["amount"]
                    fixedC = int(number) / 1000000

                if Type == "LOAN":
                    number = data_info["price"]["amount"]
                    fixedL = int(number) / 10000

                name = data_info["assets"][0]["name"]
                timez = data_info["assets"][0]["transferred_at_time"]
                timeMs = data_info["created_at_time"]
                timeSec = int(timeMs) / 1000
                number_of_nft = int(data_info["assets"][0]["template_mint"])
                buyer = data_info["buyer"]
                seller = data_info["seller"]
                local_time = datetime.utcfromtimestamp(timeSec).strftime(
                    "%m-%d-%Y %H:%M:%S"
                )

                authers_list.append(
                    [
                        seller,
                        fixedC,
                        fixedX,
                        fixedL,
                        buyer,
                        number_of_nft,
                        name,
                        local_time,
                    ]
                )

            auctions = (
                "https://proton.api.atomicassets.io/atomicmarket/v1/auctions?state=3&seller={}&collection_name={}"
                "&before={}&page=1&limit=100&order=desc&sort=created".format(
                    auther, collection_name, timeMs
                )
            )
            auctions = requests.get(auctions).text
            auctions_ = json.loads(auctions)

        name_df = pd.DataFrame(
            data=authers_list,
            columns=[
                "author ",
                "price listed usd",
                "price listed xpr",
                "price listed loan",
                "buyer",
                "# of nft",
                "name",
                "time",
            ],
        )
        total = name_df["price listed usd"].sum()
        name_df.at["Total", "price listed usd"] = name_df["price listed usd"].sum()
        xprtotal = name_df["price listed xpr"].sum()
        loantotal = name_df["price listed loan"].sum()

        wb = Workbook()

        ws = wb.active
        ws.title = FirstSale
        ws2 = wb.create_sheet(resales)
        ws3 = wb.create_sheet(Holders)
        for r in dataframe_to_rows(name_df, index=False):
            ws.append(r)
        dims = {}
        for row in ws.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column_letter] = max(
                        (dims.get(cell.column_letter, 0), len(str(cell.value)))
                    )
        for col, value in dims.items():
            ws.column_dimensions[col].width = value
        maxrow = ws.max_row
        ws.cell(row=maxrow, column=1, value="totals")
        ws.cell(row=maxrow, column=3).value = xprtotal
        ws.cell(row=maxrow, column=4).value = loantotal

        resales = (
            "https://proton.api.atomicassets.io/atomicmarket/v1/sales?state=1%2C3&seller_blacklist"
            "={}&buyer_blacklist={}&collection_name={}&page=1&limit=100&order"
            "=desc&sort=updated".format(auther, auther, collection_name)
        )
        resales = requests.get(resales).text
        resales_ = json.loads(resales)
        print("getting resales")
        while len(resales_["data"]) != 0:
            for data_info in resales_["data"]:
                fixedC = 0
                fixedX = 0
                fixedL = 0
                Type = data_info["listing_symbol"]
                if Type == "XPR":
                    number = data_info["listing_price"]
                    fixedX = int(number) / 10000

                if Type == "XUSDC":
                    number = data_info["listing_price"]
                    fixedC = int(number) / 1000000

                if Type == "LOAN":
                    number = data_info["listing_price"]
                    fixedL = int(number) / 10000

                number = data_info["listing_price"]
                fixed = int(number) / 1000000
                name = data_info["assets"][0]["name"]
                timez = data_info["created_at_time"]
                timef = data_info["updated_at_time"]
                number_of_nft = int(data_info["assets"][0]["template_mint"])
                Royal = data_info["collection"]["market_fee"]
                timefixe = int(timez) / 1000
                buyer = data_info["buyer"]
                seller = data_info["seller"]
                local_time = datetime.utcfromtimestamp(timefixe).strftime(
                    "%d-%m-%Y %H:%M:%S"
                )
                all_list.append(
                    [
                        seller,
                        buyer,
                        fixedC,
                        fixedX,
                        fixedL,
                        name,
                        number_of_nft,
                        local_time,
                    ]
                )

            time.sleep(2)
            resales = (
                "https://proton.api.atomicassets.io/atomicmarket/v1/sales?state=1%2C3&seller_blacklist"
                "={}&buyer_blacklist={}&collection_name={}&before={}&page=1&limit=100&order"
                "=desc&sort=updated".format(auther, auther, collection_name, timef)
            )
            resales = requests.get(resales).text
            resales_ = json.loads(resales)
        time.sleep(4)
        resellOffer = (
            "https://proton.api.atomicassets.io/atomicmarket/v1/buyoffers?state=3&seller_blacklist={}"
            "&collection_name={}&page=1&limit=100&order=desc&sort=created".format(
                auther, collection_name
            )
        )
        resellOffer = requests.get(resellOffer).text
        resellOffer_ = json.loads(resellOffer)
        print("getting resales offers")
        while len(resellOffer_["data"]) != 0:

            for data_info in resellOffer_["data"]:
                fixedC = 0
                fixedX = 0
                fixedL = 0
                Type = data_info["price"]["token_symbol"]
                if Type == "XPR":
                    number = data_info["price"]["amount"]
                    fixedX = int(number) / 10000

                if Type == "XUSDC":
                    number = data_info["price"]["amount"]
                    fixedC = int(number) / 1000000

                if Type == "LOAN":
                    number = data_info["price"]["amount"]
                    fixedL = int(number) / 10000

                fixed = int(number) / 1000000
                name = data_info["assets"][0]["name"]
                timez = data_info["assets"][0]["transferred_at_time"]
                timef = data_info["created_at_time"]
                timex = int(timef) / 1000
                number_of_nft = int(data_info["assets"][0]["template_mint"])
                buyer = data_info["buyer"]
                seller = data_info["seller"]
                local_time = datetime.utcfromtimestamp(timex).strftime(
                    "%d-%m-%Y %H:%M:%S"
                )
                all_list.append(
                    [
                        seller,
                        buyer,
                        fixedC,
                        fixedX,
                        fixedL,
                        name,
                        number_of_nft,
                        local_time,
                    ]
                )

            time.sleep(0.2)
            resellOffer = (
                "https://proton.api.atomicassets.io/atomicmarket/v1/buyoffers?state=3&seller_blacklist={}"
                "&collection_name={}&before={}&page=1&limit=100&order=desc&sort=created".format(
                    auther, collection_name, timef
                )
            )
            resellOffer = requests.get(resellOffer).text
            resellOffer_ = json.loads(resellOffer)

        auctionResale = "https://proton.api.atomicassets.io/atomicmarket/v1/auctions?state=3&seller_blacklist={}&collection_name={}&page=1&limit=100&order=desc&sort=created".format(
            auther, collection_name
        )
        auctionResale = requests.get(auctionResale).text
        auctionResale_ = json.loads(auctionResale)
        print("getting auction resales")
        while len(auctionResale_["data"]) != 0:

            for data_info in auctionResale_["data"]:
                fixedC = 0
                fixedX = 0
                fixedL = 0
                Type = data_info["price"]["token_symbol"]
                if Type == "XPR":
                    number = data_info["price"]["amount"]
                    fixedX = int(number) / 10000

                if Type == "XUSDC":
                    number = data_info["price"]["amount"]
                    fixedC = int(number) / 1000000

                if Type == "LOAN":
                    number = data_info["price"]["amount"]
                    fixedL = int(number) / 10000
                name = data_info["assets"][0]["name"]
                timez = data_info["assets"][0]["transferred_at_time"]
                timeMs = data_info["updated_at_time"]
                number_of_nft = int(data_info["assets"][0]["template_mint"])
                timeSec = int(timeMs) / 1000
                buyer = data_info["buyer"]
                seller = data_info["seller"]
                local_time = datetime.utcfromtimestamp(timeSec).strftime(
                    "%d-%m-%Y %H:%M:%S"
                )
                all_list.append(
                    [
                        seller,
                        buyer,
                        fixedC,
                        fixedX,
                        fixedL,
                        name,
                        number_of_nft,
                        local_time,
                    ]
                )

            time.sleep(0.6)
            auctionResale = (
                "https://proton.api.atomicassets.io/atomicmarket/v1/auctions?state=3&seller_blacklist={}&collection_name={}"
                "&before={}&page=1&limit=100&order=desc&sort=updated".format(
                    auther, collection_name, timeMs
                )
            )
            auctionResale = requests.get(auctionResale).text
            auctionResale_ = json.loads(auctionResale)
        names_df = pd.DataFrame(
            data=all_list,
            columns=[
                "first buyer ",
                "next buyer",
                "price paid usd",
                "price paid xpr",
                "price paid loans",
                "name",
                "# of nft",
                "time",
            ],
        )
        names_df.drop(
            names_df[names_df["first buyer "] == f"{auther}"].index, inplace=True
        )
        totals = names_df["price paid usd"].sum()
        totalsX = names_df["price paid xpr"].sum()
        totalsL = names_df["price paid loans"].sum()
        Royalties = totals * Royal
        RoyaltiesX = totalsX * Royal
        RoyaltiesL = totalsL * Royal
        Rows = int(names_df.index.max() + 1)
        names_df.at[Rows, "price paid usd"] = Royalties
        names_df.at[Rows, "price paid xpr"] = RoyaltiesX
        names_df.at[Rows, "price paid loans"] = RoyaltiesL

        for r in dataframe_to_rows(names_df, index=False):
            ws2.append(r)
        dims = {}
        for row in ws2.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column_letter] = max(
                        (dims.get(cell.column_letter, 0), len(str(cell.value)))
                    )
        for col, value in dims.items():
            ws2.column_dimensions[col].width = value
        maxrow = ws2.max_row
        ws2.cell(row=maxrow, column=1, value="Royalties")
        points = 0
        holders = (
            "https://proton.api.atomicassets.io/atomicassets/v1/accounts?collection_name={}"
            "&page=1&limit=100&order=desc".format(collection_name)
        )
        holders = requests.get(holders).text
        holders_ = json.loads(holders)
        pages = 1
        while len(holders_["data"]) != 0:
            pages = pages + 1
            for data_info in holders_["data"]:
                holders = int(data_info["assets"])
                authersName = data_info["account"]
                if authersName != "crystalfrogs":
                    holder_list.append([data_info["account"], holders])
            holders = (
                "https://proton.api.atomicassets.io/atomicassets/v1/accounts?collection_name={}"
                "&page={}&limit=100&order=desc".format(collection_name, pages)
            )
            holders = requests.get(holders).text
            holders_ = json.loads(holders)
        crystal_holder_df = pd.DataFrame(
            data=holder_list, columns=["account", "amount held"]
        )
        len(crystal_holder_df) - 1
        count = 0
        rowz = 1
        Peoplelist = []
        for r in dataframe_to_rows(crystal_holder_df, index=False):
            ws3.append(r)
        for r in dataframe_to_rows(crystal_holder_df, index=False):
            count = count + 1
            if count == 1:
                """"""
            else:
                ws3.cell(row=count, column=3).value = 0
                ws3.cell(row=count, column=4).value = 0
                ws3.cell(row=count, column=5).value = 0
                ws3.cell(row=count, column=6).value = 0
                ws3.cell(row=count, column=7).value = 0

        temp1 = 0
        temp3 = 0
        temp2 = 0
        totalz = 0
        pages = 0
        holders = (
            "https://proton.api.atomicassets.io/atomicassets/v1/accounts?collection_name={}"
            "&page=1&limit=100&order=desc".format(collection_name)
        )
        holders = requests.get(holders).text
        holders_ = json.loads(holders)
        amount = 1

        while len(holders_["data"]) != 0:
            pages = pages + 1
            amount = amount + 1
            for data_info in holders_["data"]:
                if data_info["account"] != "crystalfrogs":
                    checker = data_info["account"]
                    if rowz != 1:
                        temp2 = ws3.cell(row=rowz, column=3).value
                    print("getting {}'s data".format(checker))
                    totalz = 0
                    people = "https://proton.api.atomicassets.io/atomicmarket/v1/assets?collection_name={}&owner={}&page=1&limit=100&order=desc&sort=asset_id".format(
                        collection_name, checker
                    )
                    test = requests.get(people)
                    next = test.headers["X-RateLimit-Reset"]
                    resset = test.headers["X-RateLimit-Remaining"]
                    resset = int(resset)
                    next = int(next)
                    wait = next - time.time()
                    if resset < 3:
                        time.sleep(wait)
                    people_ = json.loads((test.text))
                    time.sleep(0.2)
                    count = 0
                    pages = 1
                    rowz = rowz + 1
                    assitID2 = " "
                    ImpCount = 0

                    while len(people_["data"]) != 0:
                        pages = pages + 1
                        if pages == 2:
                            count = 0
                        if pages == 3:
                            count = 100
                        if pages == 4:
                            count = 200
                        for data_info in people_["data"]:
                            word = data_info["data"]["desc"]
                            nft_name = data_info["data"]["name"]
                            number_of_nft = data_info["template_mint"]
                            assitID1 = data_info["asset_id"]
                            if assitID1 != assitID2:
                                assitID2 = assitID1
                                count = count + 1
                                ws3.cell(row=1, column=3).value = "crystals per week"
                                ws3.cell(row=1, column=4).value = "total crystals"
                                ws3.cell(row=1, column=5).value = "miner"
                                ws3.cell(row=1, column=6).value = "wizard"
                                ws3.cell(row=1, column=7).value = "lord"

                                points = ws3.cell(row=rowz, column=3).value
                                crystal = ws3.cell(row=rowz, column=4).value
                                miner = ws3.cell(row=rowz, column=5).value
                                wizard = ws3.cell(row=rowz, column=6).value
                                lord = ws3.cell(row=rowz, column=7).value
                                ws3.cell(row=rowz, column=7 + count).value = (
                                    nft_name + " (#" + number_of_nft + ")"
                                )
                                ws3.cell(row=1, column=7 + count).value = "nft " + str(
                                    count
                                )
                                nft_name = nft_name.lower()
                                if not "token" in nft_name:
                                    if "mysterious crystal" in nft_name:
                                        word = nft_name
                                        start = word.find("(")
                                        start += 1
                                        end = word.find(")")
                                        new = int(word[start:end])
                                        ws3.cell(row=rowz, column=4).value = (
                                            crystal + new
                                        )

                                    if "miner" in nft_name:
                                        ws3.cell(row=rowz, column=3).value = points + 5
                                        ws3.cell(row=rowz, column=5).value = miner + 1
                                    if "wizard" in nft_name:
                                        ws3.cell(row=rowz, column=3).value = points + 10
                                        ws3.cell(row=rowz, column=6).value = wizard + 1
                                    if "lord" in nft_name:
                                        ws3.cell(row=rowz, column=3).value = points + 20
                                        ws3.cell(row=rowz, column=7).value = lord + 1
                        people = "https://proton.api.atomicassets.io/atomicmarket/v1/assets?collection_name={}&owner={}&page={}&limit=100&order=desc&sort=asset_id".format(
                            collection_name, checker, pages
                        )
                        test = requests.get(people)
                        next = test.headers["X-RateLimit-Reset"]
                        resset = test.headers["X-RateLimit-Remaining"]
                        resset = int(resset)
                        next = int(next)
                        wait = next - time.time()
                        if resset < 3:
                            time.sleep(wait)
                        people_ = json.loads((test.text))
                        time.sleep(0.2)
                    holders_amount = ws3.cell(row=rowz, column=2).value
                    totalholderslist.append([checker, holders_amount])

            holders = (
                "https://proton.api.atomicassets.io/atomicassets/v1/accounts?collection_name={}"
                "&page={}&limit=100&order=desc".format(collection_name, amount)
            )
            holders = requests.get(holders).text
            holders_ = json.loads(holders)

        count = 0

        for row in ws3.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column_letter] = max(
                        (dims.get(cell.column_letter, 0), len(str(cell.value)))
                    )
        for col, value in dims.items():
            ws3.column_dimensions[col].width = value + 5
        excelsave = "".join(excelsheetname)
        wb.save(excelsave)
        print("Creating the excel file")
        wb.close()
        edit = pd.read_excel(excelsave, 2, index_col=False)
        holdersTab = edit
        holdersTab = holdersTab.loc[
            :, ~holdersTab.columns.str.contains("crystals per week")
        ]
        holdersTab = holdersTab.loc[:, ~holdersTab.columns.str.contains("crystals")]
        holdersTab = holdersTab.loc[:, ~holdersTab.columns.str.contains("miner")]
        holdersTab = holdersTab.loc[:, ~holdersTab.columns.str.contains("wizard")]
        holdersTab = holdersTab.loc[:, ~holdersTab.columns.str.contains("lord")]
        crystalsTab = edit
        crystalsTab.sort_values(by=["crystals per week"], ascending=False, inplace=True)
        crystalsTab = crystalsTab.loc[:, ~crystalsTab.columns.str.contains("^nft")]
        crystalsTab = crystalsTab.loc[
            :, ~crystalsTab.columns.str.contains("amount held")
        ]
        wb3 = openpyxl.load_workbook(excelsave)
        ws = wb3["Holders"]
        wb3.remove(ws)
        ws5 = wb3.create_sheet("Holders")
        ws6 = wb3.create_sheet("Crystals")
        writeToExcel(ws5, holdersTab)
        writeToExcel(ws6, crystalsTab)
        os.remove(excelsave)
        wb3.save(excelsave)
        wb3.close()
        os.chdir(path.parent.absolute())

    normalServic(authers, all, resales, FirstSale, Holders, holders, *excelsheetname)


auther = "crystalfrogs"
universe = "Crystal Frogs"
heading = "{} Collection".format(universe)
collection_name = "542514111454"
collection1 = "Crystal Frogs"
excelsheetname1 = "{}.xlsx".format(collection1)
collection(auther, collection_name, heading, excelsheetname1)
