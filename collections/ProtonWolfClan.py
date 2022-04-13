from ast import Pass
import json
import os
from wsgiref.simple_server import WSGIServer
import pandas as pd
import requests as requests
import pathlib
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import time

print(
    "Thank you for running the program\nIt will take a few seconds to create everything\nHope you enjoy  \n-kevcollector"
)
# set what command you want to run here
# os.getenv('PWD')
# mac=pathlib.Path().cwd() /'Desktop'
# os.chdir(mac)
# os.system('cls' if os.name == 'nt' else 'clear')
# userCollection = input("Please enter the collection you want to scan")
# os.system('cls' if os.name == 'nt' else 'clear')


def writeToExcel(worksheet, data):
    for r in dataframe_to_rows(data, index=False):
        worksheet.append(r)
    dims = {}
    for row in worksheet.rows:
        for cell in row:
            if cell.value:
                dims[cell.column_letter] = max(
                    (dims.get(cell.column_letter, 0), len(str(cell.value)))
                )
    for col, value in dims.items():
        worksheet.column_dimensions[col].width = value


current = pathlib.Path().cwd()

totalholderslist = []


def collection(author, collection_name, heading, *excelsheetname):
    collecion = "".join(excelsheetname)
    collecion = collecion.replace(".xlsx", "")

    path = pathlib.Path().cwd() / ("{}".format(heading))
    if current != pathlib.Path().cwd():
        path = pathlib.Path().cwd()
        pathlib.Path(path).mkdir(parents=True, exist_ok=True)
    else:
        pathlib.Path(path).mkdir(parents=True, exist_ok=True)

    holders = " "

    holder_list = []

    def normalServic(author, holders, *excelsheetname):
        global totalholderslist
        os.chdir(path)

        wb1 = Workbook()
        ws1 = wb1.active
        ws1.title = "holders"

        pages = 0
        holders = (
            "https://proton.api.atomicassets.io/atomicassets/v1/accounts?collection_name={}"
            "&page=1&limit=100&order=desc".format(collection_name)
        )
        holders = requests.get(holders).text
        holders_ = json.loads(holders)
        amount = 1
        rowz = 1

        while len(holders_["data"]) != 0:
            pages = pages + 1
            amount = amount + 1
            for data_info in holders_["data"]:
                if data_info["account"] != author:
                    checker = data_info["account"]
                    print("getting {}'s data".format(checker))
                    people = "https://proton.api.atomicassets.io/atomicmarket/v1/assets?collection_name={}&owner={}&page=1&limit=100&order=desc&sort=asset_id".format(
                        collection_name, checker
                    )
                    test = requests.get(people)
                    next = test.headers["X-RateLimit-Reset"]
                    resset = test.headers["X-RateLimit-Remaining"]
                    resset = int(resset)
                    next = int(next)
                    wait = next - time.time()
                    if resset < 3:
                        time.sleep(wait)
                    people_ = json.loads((test.text))
                    count = 0
                    wolfPoints = 0
                    rowz = rowz + 1
                    pages = 1
                    assitID2 = " "

                    while len(people_["data"]) != 0:
                        pages = pages + 1
                        if pages >= 3:
                            count = pages * 100

                        for data_info in people_["data"]:
                            nft_name = data_info["data"]["name"]
                            assitID1 = data_info["asset_id"]
                            if assitID1 != assitID2:
                                assitID2 = assitID1
                                done = nft_name

                                if collection_name == "524211545444":
                                    count = count + 1

                                    if nft_name == "Proton Wolf Clan #1":
                                        ws1.cell(
                                            row=rowz, column=3
                                        ).value = wolfPoints = (wolfPoints + 3)
                                    if nft_name == "Proton Wolf Clan #2":
                                        ws1.cell(
                                            row=rowz, column=3
                                        ).value = wolfPoints = (wolfPoints + 7)
                                    if nft_name == "Proton Wolf Clan #3":
                                        ws1.cell(
                                            row=rowz, column=3
                                        ).value = wolfPoints = (wolfPoints + 3)
                                    if nft_name == "Proton Wolf Clan #4":
                                        ws1.cell(
                                            row=rowz, column=3
                                        ).value = wolfPoints = (wolfPoints + 3)
                                    if nft_name == "Proton Wolf Clan #5":
                                        ws1.cell(
                                            row=rowz, column=3
                                        ).value = wolfPoints = (wolfPoints + 3)
                                    if nft_name == "Proton Wolf Clan #6":
                                        ws1.cell(
                                            row=rowz, column=3
                                        ).value = wolfPoints = (wolfPoints + 7)
                                    if nft_name == "Proton Wolf Clan #7":
                                        ws1.cell(
                                            row=rowz, column=3
                                        ).value = wolfPoints = (wolfPoints + 3)
                                    if nft_name == "Proton Wolf Clan #8":
                                        ws1.cell(
                                            row=rowz, column=3
                                        ).value = wolfPoints = (wolfPoints + 5)
                                    if nft_name == "Proton Wolf Clan #9":
                                        ws1.cell(
                                            row=rowz, column=3
                                        ).value = wolfPoints = (wolfPoints + 3)
                                    if nft_name == "Proton Wolf Clan #10":
                                        ws1.cell(
                                            row=rowz, column=3
                                        ).value = wolfPoints = (wolfPoints + 7)
                                    if nft_name == "Proton Wolf Clan #11":
                                        ws1.cell(
                                            row=rowz, column=3
                                        ).value = wolfPoints = (wolfPoints + 3)
                                    if nft_name == "Proton Wolf Clan #12":
                                        ws1.cell(
                                            row=rowz, column=3
                                        ).value = wolfPoints = (wolfPoints + 3)
                                    if nft_name == "Proton Wolf Clan #13":
                                        ws1.cell(
                                            row=rowz, column=3
                                        ).value = wolfPoints = (wolfPoints + 3)
                                    if nft_name == "Proton Wolf Clan #14":
                                        ws1.cell(
                                            row=rowz, column=3
                                        ).value = wolfPoints = (wolfPoints + 3)
                                    if nft_name == "Proton Wolf Clan #15":
                                        ws1.cell(
                                            row=rowz, column=3
                                        ).value = wolfPoints = (wolfPoints + 5)
                                    if nft_name == "Proton Wolf Clan #16":
                                        ws1.cell(
                                            row=rowz, column=3
                                        ).value = wolfPoints = (wolfPoints + 3)
                                    if nft_name == "Proton Wolf Clan #17":
                                        ws1.cell(
                                            row=rowz, column=3
                                        ).value = wolfPoints = (wolfPoints + 3)
                                    if nft_name == "Proton Wolf Clan #18":
                                        ws1.cell(
                                            row=rowz, column=3
                                        ).value = wolfPoints = (wolfPoints + 3)
                                    if nft_name == "Proton Wolf Clan #19":
                                        ws1.cell(
                                            row=rowz, column=3
                                        ).value = wolfPoints = (wolfPoints + 3)
                                    if nft_name == "Proton Wolf Clan #20":
                                        ws1.cell(
                                            row=rowz, column=3
                                        ).value = wolfPoints = (wolfPoints + 3)
                                    if nft_name == "Proton Wolf Clan #21":
                                        ws1.cell(
                                            row=rowz, column=3
                                        ).value = wolfPoints = (wolfPoints + 3)
                                    if nft_name == "Proton Wolf Clan #22":
                                        ws1.cell(
                                            row=rowz, column=3
                                        ).value = wolfPoints = (wolfPoints + 7)
                                    if nft_name == "Proton Wolf Clan #23":
                                        ws1.cell(
                                            row=rowz, column=3
                                        ).value = wolfPoints = (wolfPoints + 3)
                                    if nft_name == "Proton Wolf Clan #24":
                                        ws1.cell(
                                            row=rowz, column=3
                                        ).value = wolfPoints = (wolfPoints + 3)
                                    if nft_name == "Proton Wolf Clan #25":
                                        ws1.cell(
                                            row=rowz, column=3
                                        ).value = wolfPoints = (wolfPoints + 3)
                                    if nft_name == "Proton Wolf Clan #26":
                                        ws1.cell(
                                            row=rowz, column=3
                                        ).value = wolfPoints = (wolfPoints + 3)
                                    if nft_name == "Proton Wolf Clan #27":
                                        ws1.cell(
                                            row=rowz, column=3
                                        ).value = wolfPoints = (wolfPoints + 5)
                                    if nft_name == "Proton Wolf Clan #28":
                                        ws1.cell(
                                            row=rowz, column=3
                                        ).value = wolfPoints = (wolfPoints + 3)
                                    if nft_name == "Proton Wolf Clan #29":
                                        ws1.cell(
                                            row=rowz, column=3
                                        ).value = wolfPoints = (wolfPoints + 5)
                                    if nft_name == "Proton Wolf Clan #30":
                                        ws1.cell(
                                            row=rowz, column=3
                                        ).value = wolfPoints = (wolfPoints + 3)
                                    if nft_name == "Proton Wolf Clan #31":
                                        ws1.cell(
                                            row=rowz, column=3
                                        ).value = wolfPoints = (wolfPoints + 3)
                                    if nft_name == "Proton Wolf Clan #32":
                                        ws1.cell(
                                            row=rowz, column=3
                                        ).value = wolfPoints = (wolfPoints + 3)
                                    if nft_name == "Proton Wolf Clan #33":
                                        ws1.cell(
                                            row=rowz, column=3
                                        ).value = wolfPoints = (wolfPoints + 7)
                                    if nft_name == "Proton Wolf Clan #34":
                                        ws1.cell(
                                            row=rowz, column=3
                                        ).value = wolfPoints = (wolfPoints + 3)
                                    if nft_name == "Proton Wolf Clan #35":
                                        ws1.cell(
                                            row=rowz, column=3
                                        ).value = wolfPoints = (wolfPoints + 3)
                                    if nft_name == "Proton Wolf Clan #36":
                                        ws1.cell(
                                            row=rowz, column=3
                                        ).value = wolfPoints = (wolfPoints + 5)
                                    if nft_name == "Proton Wolf Clan #37":
                                        ws1.cell(
                                            row=rowz, column=3
                                        ).value = wolfPoints = (wolfPoints + 3)
                                    if nft_name == "Proton Wolf Clan #38":
                                        ws1.cell(
                                            row=rowz, column=3
                                        ).value = wolfPoints = (wolfPoints + 3)
                                    if nft_name == "Proton Wolf Clan #39":
                                        ws1.cell(
                                            row=rowz, column=3
                                        ).value = wolfPoints = (wolfPoints + 3)
                                    if nft_name == "Proton Wolf Clan #40":
                                        ws1.cell(
                                            row=rowz, column=3
                                        ).value = wolfPoints = (wolfPoints + 3)
                                    if nft_name == "Proton Wolf Clan #41":
                                        ws1.cell(
                                            row=rowz, column=3
                                        ).value = wolfPoints = (wolfPoints + 3)
                                    if nft_name == "Proton Wolf Clan #42":
                                        ws1.cell(
                                            row=rowz, column=3
                                        ).value = wolfPoints = (wolfPoints + 3)
                                    if nft_name == "Proton Wolf Clan #43":
                                        ws1.cell(
                                            row=rowz, column=3
                                        ).value = wolfPoints = (wolfPoints + 3)
                                    if nft_name == "Proton Wolf Clan #44":
                                        ws1.cell(
                                            row=rowz, column=3
                                        ).value = wolfPoints = (wolfPoints + 5)
                                    if nft_name == "Proton Wolf Clan #45":
                                        ws1.cell(
                                            row=rowz, column=3
                                        ).value = wolfPoints = (wolfPoints + 3)
                                    if nft_name == "Proton Wolf Clan #46":
                                        ws1.cell(
                                            row=rowz, column=3
                                        ).value = wolfPoints = (wolfPoints + 3)
                                    if nft_name == "Proton Wolf Clan #47":
                                        ws1.cell(
                                            row=rowz, column=3
                                        ).value = wolfPoints = (wolfPoints + 3)
                                    if nft_name == "Proton Wolf Clan #48":
                                        ws1.cell(
                                            row=rowz, column=3
                                        ).value = wolfPoints = (wolfPoints + 3)
                                    if nft_name == "Proton Wolf Clan #49":
                                        ws1.cell(
                                            row=rowz, column=3
                                        ).value = wolfPoints = (wolfPoints + 3)
                                    if nft_name == "Proton Wolf Clan #50":
                                        ws1.cell(
                                            row=rowz, column=3
                                        ).value = wolfPoints = (wolfPoints + 5)
                                    if nft_name == "Proton Wolf Clan #51":
                                        ws1.cell(
                                            row=rowz, column=3
                                        ).value = wolfPoints = (wolfPoints + 3)
                                    if nft_name == "Proton Wolf Clan #52":
                                        ws1.cell(
                                            row=rowz, column=3
                                        ).value = wolfPoints = (wolfPoints + 3)
                                    if nft_name == "Proton Wolf Clan #53":
                                        ws1.cell(
                                            row=rowz, column=3
                                        ).value = wolfPoints = (wolfPoints + 3)
                                    if nft_name == "Proton Wolf Clan 54":
                                        ws1.cell(
                                            row=rowz, column=3
                                        ).value = wolfPoints = (wolfPoints + 3)
                                    if nft_name == "Proton Wolf Clan #55":
                                        ws1.cell(
                                            row=rowz, column=3
                                        ).value = wolfPoints = (wolfPoints + 7)
                                    if nft_name == "Proton Wolf Clan #56":
                                        ws1.cell(
                                            row=rowz, column=3
                                        ).value = wolfPoints = (wolfPoints + 3)
                                    if nft_name == "Proton Wolf Clan #57":
                                        ws1.cell(
                                            row=rowz, column=3
                                        ).value = wolfPoints = (wolfPoints + 3)
                                    if nft_name == "Proton Wolf Clan #58":
                                        ws1.cell(
                                            row=rowz, column=3
                                        ).value = wolfPoints = (wolfPoints + 3)
                                    if nft_name == "Proton Wolf Clan #59":
                                        ws1.cell(
                                            row=rowz, column=3
                                        ).value = wolfPoints = (wolfPoints + 3)
                                    if nft_name == "Proton Wolf Clan #60":
                                        ws1.cell(
                                            row=rowz, column=3
                                        ).value = wolfPoints = (wolfPoints + 3)
                                    if nft_name == "Proton Wolf Clan #61":
                                        ws1.cell(
                                            row=rowz, column=3
                                        ).value = wolfPoints = (wolfPoints + 3)
                                    if nft_name == "Proton Wolf Clan #62":
                                        ws1.cell(
                                            row=rowz, column=3
                                        ).value = wolfPoints = (wolfPoints + 3)
                                    if nft_name == "Proton Wolf Clan #63":
                                        ws1.cell(
                                            row=rowz, column=3
                                        ).value = wolfPoints = (wolfPoints + 3)
                                    if nft_name == "Proton Wolf Clan #64":
                                        ws1.cell(
                                            row=rowz, column=3
                                        ).value = wolfPoints = (wolfPoints + 3)
                                    if nft_name == "Proton Wolf Clan #65":
                                        ws1.cell(
                                            row=rowz, column=3
                                        ).value = wolfPoints = (wolfPoints + 3)
                                    if nft_name == "Proton Wolf Clan #66":
                                        ws1.cell(
                                            row=rowz, column=3
                                        ).value = wolfPoints = (wolfPoints + 5)
                                    if nft_name == "Proton Wolf Clan #67":
                                        ws1.cell(
                                            row=rowz, column=3
                                        ).value = wolfPoints = (wolfPoints + 3)
                                    if nft_name == "Proton Wolf Clan #68":
                                        ws1.cell(
                                            row=rowz, column=3
                                        ).value = wolfPoints = (wolfPoints + 3)
                                    if nft_name == "Proton Wolf Clan #69":
                                        ws1.cell(
                                            row=rowz, column=3
                                        ).value = wolfPoints = (wolfPoints + 5)
                                    if nft_name == "Proton Wolf Clan #70":
                                        ws1.cell(
                                            row=rowz, column=3
                                        ).value = wolfPoints = (wolfPoints + 3)
                                    if nft_name == "Proton Wolf Clan #71":
                                        ws1.cell(
                                            row=rowz, column=3
                                        ).value = wolfPoints = (wolfPoints + 3)
                                    if nft_name == "Proton Wolf Clan #72":
                                        ws1.cell(
                                            row=rowz, column=3
                                        ).value = wolfPoints = (wolfPoints + 3)
                                    if nft_name == "Proton Wolf Clan #73":
                                        ws1.cell(
                                            row=rowz, column=3
                                        ).value = wolfPoints = (wolfPoints + 3)
                                    if nft_name == "Proton Wolf Clan #74":
                                        ws1.cell(
                                            row=rowz, column=3
                                        ).value = wolfPoints = (wolfPoints + 3)
                                    if nft_name == "Proton Wolf Clan #75":
                                        ws1.cell(
                                            row=rowz, column=3
                                        ).value = wolfPoints = (wolfPoints + 5)
                                    if nft_name == "Proton Wolf Clan #76":
                                        ws1.cell(
                                            row=rowz, column=3
                                        ).value = wolfPoints = (wolfPoints + 3)
                                    if nft_name == "Proton Wolf Clan #77":
                                        ws1.cell(
                                            row=rowz, column=3
                                        ).value = wolfPoints = (wolfPoints + 7)
                                    if nft_name == "Proton Wolf Clan #78":
                                        ws1.cell(
                                            row=rowz, column=3
                                        ).value = wolfPoints = (wolfPoints + 3)
                                    if nft_name == "Proton Wolf Clan #79":
                                        ws1.cell(
                                            row=rowz, column=3
                                        ).value = wolfPoints = (wolfPoints + 3)
                                    if nft_name == "Proton Wolf Clan #80":
                                        ws1.cell(
                                            row=rowz, column=3
                                        ).value = wolfPoints = (wolfPoints + 5)
                                    if nft_name == "Proton Wolf Clan #81":
                                        ws1.cell(
                                            row=rowz, column=3
                                        ).value = wolfPoints = (wolfPoints + 3)
                                    if nft_name == "Proton Wolf Clan #82":
                                        ws1.cell(
                                            row=rowz, column=3
                                        ).value = wolfPoints = (wolfPoints + 3)
                                    if nft_name == "Proton Wolf Clan #83":
                                        ws1.cell(
                                            row=rowz, column=3
                                        ).value = wolfPoints = (wolfPoints + 3)
                                    if nft_name == "Proton Wolf Clan #84":
                                        ws1.cell(
                                            row=rowz, column=3
                                        ).value = wolfPoints = (wolfPoints + 7)
                                    if nft_name == "Proton Wolf Clan #85":
                                        ws1.cell(
                                            row=rowz, column=3
                                        ).value = wolfPoints = (wolfPoints + 5)
                                    if nft_name == "Proton Wolf Clan #86":
                                        ws1.cell(
                                            row=rowz, column=3
                                        ).value = wolfPoints = (wolfPoints + 5)
                                    if nft_name == "Proton Wolf Clan #87":
                                        ws1.cell(
                                            row=rowz, column=3
                                        ).value = wolfPoints = (wolfPoints + 3)
                                    if nft_name == "Proton Wolf Clan #88":
                                        ws1.cell(
                                            row=rowz, column=3
                                        ).value = wolfPoints = (wolfPoints + 7)
                                    if nft_name == "Proton Wolf Clan #89":
                                        ws1.cell(
                                            row=rowz, column=3
                                        ).value = wolfPoints = (wolfPoints + 3)
                                    if nft_name == "Proton Wolf Clan #90":
                                        ws1.cell(
                                            row=rowz, column=3
                                        ).value = wolfPoints = (wolfPoints + 5)
                                    if nft_name == "Proton Wolf Clan #91":
                                        ws1.cell(
                                            row=rowz, column=3
                                        ).value = wolfPoints = (wolfPoints + 5)
                                    if nft_name == "Proton Wolf Clan #92":
                                        ws1.cell(
                                            row=rowz, column=3
                                        ).value = wolfPoints = (wolfPoints + 3)
                                    if nft_name == "Proton Wolf Clan #93":
                                        ws1.cell(
                                            row=rowz, column=3
                                        ).value = wolfPoints = (wolfPoints + 3)
                                    if nft_name == "Proton Wolf Clan #94":
                                        ws1.cell(
                                            row=rowz, column=3
                                        ).value = wolfPoints = (wolfPoints + 5)
                                    if nft_name == "Proton Wolf Clan #95":
                                        ws1.cell(
                                            row=rowz, column=3
                                        ).value = wolfPoints = (wolfPoints + 5)
                                    if nft_name == "Proton Wolf Clan #96":
                                        ws1.cell(
                                            row=rowz, column=3
                                        ).value = wolfPoints = (wolfPoints + 3)
                                    if nft_name == "Proton Wolf Clan #97":
                                        ws1.cell(
                                            row=rowz, column=3
                                        ).value = wolfPoints = (wolfPoints + 5)
                                    if (
                                        nft_name
                                        == "Proton Wolf Clan #98 | Special Rare"
                                    ):
                                        ws1.cell(
                                            row=rowz, column=3
                                        ).value = wolfPoints = (wolfPoints + 6)
                                    if (
                                        nft_name
                                        == "Proton Wolf Clan #99 | Special EPIC"
                                    ):
                                        ws1.cell(
                                            row=rowz, column=3
                                        ).value = wolfPoints = (wolfPoints + 8)
                                    if nft_name == "Proton Wolf Clan #100 | Legendary":
                                        ws1.cell(
                                            row=rowz, column=3
                                        ).value = wolfPoints = (wolfPoints + 15)
                                    else:
                                        ws1.cell(row=rowz, column=3).value = wolfPoints

                                    ws1.cell(row=rowz, column=3 + count).value = done

                                if collection_name == "3drfwzczslri":
                                    count = count + 1

                                    if "Proton Wolf Clan 2" in nft_name:
                                        if "COMMON" in nft_name:
                                            ws1.cell(
                                                row=rowz, column=3
                                            ).value = wolfPoints = (wolfPoints + 2)

                                        if "RARE" in nft_name:
                                            ws1.cell(
                                                row=rowz, column=3
                                            ).value = wolfPoints = (wolfPoints + 5)

                                        if "EPIC" in nft_name:
                                            ws1.cell(
                                                row=rowz, column=3
                                            ).value = wolfPoints = (wolfPoints + 7)

                                        if "HEROIC" in nft_name:
                                            ws1.cell(
                                                row=rowz, column=3
                                            ).value = wolfPoints = (wolfPoints + 9)

                                        if "ULTRA RARE" in nft_name:
                                            ws1.cell(
                                                row=rowz, column=3
                                            ).value = wolfPoints = (wolfPoints + 10)

                                        if "ULTRA EPIC" in nft_name:
                                            ws1.cell(
                                                row=rowz, column=3
                                            ).value = wolfPoints = (wolfPoints + 14)

                                        if "LEGENDARY" in nft_name:
                                            ws1.cell(
                                                row=rowz, column=3
                                            ).value = wolfPoints = (wolfPoints + 15)

                                    if "FUSION" in nft_name:
                                        if "ULTRA RARE" in nft_name:
                                            ws1.cell(
                                                row=rowz, column=3
                                            ).value = wolfPoints = (wolfPoints + 10)

                                        if "ULTRA EPIC" in nft_name:
                                            ws1.cell(
                                                row=rowz, column=3
                                            ).value = wolfPoints = (wolfPoints + 14)
                                    else:
                                        ws1.cell(row=rowz, column=3).value = wolfPoints

                                    ws1.cell(row=rowz, column=3 + count).value = done

                                if count > 0:
                                    ws1.cell(
                                        row=1, column=3 + count
                                    ).value = "NFT " + str(count)
                                    ws1.cell(row=1, column=1).value = "Account"
                                    ws1.cell(row=1, column=2).value = "Amount"
                                    ws1.cell(row=1, column=3).value = "Points"
                                    ws1.cell(row=rowz, column=1).value = checker
                                    ws1.cell(row=rowz, column=2).value = count

                        people = "https://proton.api.atomicassets.io/atomicmarket/v1/assets?collection_name={}&owner={}&page={}&limit=100&order=desc&sort=asset_id".format(
                            collection_name, checker, pages
                        )
                        test = requests.get(people)
                        next = test.headers["X-RateLimit-Reset"]
                        resset = test.headers["X-RateLimit-Remaining"]
                        resset = int(resset)
                        next = int(next)
                        wait = next - time.time()
                        if resset < 3:
                            time.sleep(wait)
                        people_ = json.loads((test.text))

                    holders_amount = ws1.cell(row=rowz, column=2).value

                    totalPoints = ws1.cell(row=rowz, column=3).value

                    totalholderslist.append([checker, holders_amount, totalPoints])

            holders = (
                "https://proton.api.atomicassets.io/atomicassets/v1/accounts?collection_name={}"
                "&page={}&limit=100&order=desc".format(collection_name, amount)
            )
            holders = requests.get(holders).text
            holders_ = json.loads(holders)

        count = 0
        dims = {}
        for row in ws1.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column_letter] = max(
                        (dims.get(cell.column_letter, 0), len(str(cell.value)))
                    )
        for col, value in dims.items():
            ws1.column_dimensions[col].width = value + 5

        excelsave = "".join(excelsheetname)
        wb1.save(excelsave)
        print("Creating the excel file")

        wb1.close()

        os.chdir(path.parent.absolute())

    normalServic(author, holders, *excelsheetname)


author = "redbrush"
universe = "Proton Wolf Clan"
heading = "{} Collection".format(universe)
collection_name = "524211545444"
collection1 = "PWC Gen I"
excelsheetname1 = "{}.xlsx".format(collection1)
time.sleep(6)
collection(author, collection_name, heading, excelsheetname1)
collection_name = "3drfwzczslri"
collection2 = "PWC Gen II"
excelsheetname1 = "{}.xlsx".format(collection2)
time.sleep(6)
collection(author, collection_name, heading, excelsheetname1)
wb2 = Workbook()
ws1 = wb2.active
ws1 = wb2.create_sheet("holders")
holders_df = pd.DataFrame(
    data=totalholderslist, columns=["account", "amount held", "points"]
)

print(totalholderslist)

holders_df = holders_df.sort_values(by=["points"], ascending=False)


writeToExcel(ws1, holders_df)

wb2.save("testfile.xlsx")
