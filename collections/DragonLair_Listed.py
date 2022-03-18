
import json
import os

import pandas as pd
import requests as requests
import pathlib
from openpyxl import Workbook
from openpyxl.styles import Font, Color, colors
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import time

from openpyxl.worksheet import hyperlink

print(
    "Thank you for running the programe\nIt will take a few seconds to create everything\nHope you enjoy  "
    "\n-kevcollector\n\n\n\n")


# set what command you want to run here
# os.getenv('PWD')
# mac=pathlib.Path().cwd() /'Desktop'
# os.chdir(mac)

author = 'dragontm'
universe = 'DRAGONtm Creations'
collection1 = 'DRAGONtm'
collection2 = 'Baby Dino'
collection3 = 'Baby Dino (Fusions)'
collection4 = 'Sprout'
collection5 = 'Dino'
collection6 = 'serpent'
heading = "{} Collection".format(universe)
collection_name = '445323453535'  # dragontm
collection_name1 = '423534451325'  # Dino the dragon
collection_name2 = '513555243311'  # baby dino
collection_name3 = '133523522522'  # fussion
collection_name4 = '334431125343'  # sprouts
collection_name5 = '133552212214'  # serpent

excelsheetname1 = "{}.xlsx".format(collection1)  # dra
excelsheetname2 = "{}.xlsx".format(collection2)  # bab
excelsheetname3 = "{}.xlsx".format(collection3)  # fus
excelsheetname4 = "{}.xlsx".format(collection4)
excelsheetname5 = "{}.xlsx".format(collection5)  # dino
excelsheetname6 = "{}.xlsx".format(collection6)  # dino
current = pathlib.Path().cwd()


def collection(author, collection_name, heading, *excelsheetname):
    collecion = "".join(excelsheetname)
    collecion = collecion.replace('.xlsx', '')

    path = pathlib.Path().cwd() / ("{}".format(heading))
    if (current != pathlib.Path().cwd()):
        path = pathlib.Path().cwd()
        pathlib.Path(path).mkdir(parents=True, exist_ok=True)
    else:
        pathlib.Path(path).mkdir(parents=True, exist_ok=True)

    listed = ("https://proton.api.atomicassets.io/atomicmarket/v1/sales?state=1&collection_name={}&page=1&limit=100&order=asc&sort=price".format(collection_name))
    listed = requests.get(listed)
    listed_ = json.loads(listed.text)
    listed_done = []

    def normalServic(*excelsheetname):
        os.chdir(path)

        listed = (
            "https://proton.api.atomicassets.io/atomicmarket/v1/sales?state=1&collection_name={}&page=1&limit=100&order=asc&sort=updated".format(
                collection_name))
        listed = requests.get(listed)
        listed_ = json.loads(listed.text)
        wb2 = Workbook()

        wsL = wb2.active
        wsL.title = "Cheapest"
        count = 0
        while len(listed_['data']) != 0:
            for data_info in listed_['data']:
                seller = data_info['seller']
                number = data_info["listing_price"]
                fixed = int(number) / 1000000
                name = data_info['assets'][0]['name']
                ids = data_info['assets'][0]['template']['template_id']
                number_of_nft = int(data_info['assets'][0]['template_mint'])
                timef = data_info['updated_at_time']
                amount = data_info['assets'][0]['template']['max_supply']
                amount = amount+"/"+amount
                link = "https://protonmint.com/{}/{}".format(
                    collection_name, ids)
                listed_done.append(
                    [seller, fixed, name, number_of_nft, amount, link])
                count = count + 1

            time.sleep(0.4)
            listed = (
                "https://proton.api.atomicassets.io/atomicmarket/v1/sales?state=1&collection_name={}&after={}&page=1&limit=100&order=asc&sort=updated".format(
                    collection_name, timef))
            listed = requests.get(listed)
            listed_ = json.loads(listed.text)
        listed_df = pd.DataFrame(data=listed_done,
                                 columns=["Seller", "Price", "Name", "# of NFT", "Edition Size", "Link"])
        listed_df = listed_df.sort_values(by=['Price'], ascending=True)
        count = 0

        for r in dataframe_to_rows(listed_df, index=False):
            count = count+1
            wsL.append(r)
            if count > 2:
                link = wsL.cell(row=count-1, column=6).value
                wsL.cell(
                    row=count-1, column=6).value = '=HYPERLINK("{}", "{}")'.format(link, link)
                wsL.cell(row=count-1, column=6).style = 'Hyperlink'
        count = count+1
        link = wsL.cell(row=count - 1, column=6).value
        wsL.cell(row=count - 1,
                 column=6).value = '=HYPERLINK("{}", "{}")'.format(link, link)
        wsL.cell(row=count - 1, column=6).style = 'Hyperlink'

        dims = {}
        for row in wsL.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column_letter] = max(
                        (dims.get(cell.column_letter, 0), len(str(cell.value))))
        for col, value in dims.items():
            wsL.column_dimensions[col].width = value

        excelsave = "".join(excelsheetname)
        wb2.save('Listed'+excelsave)
        print("Creating the excel file")
        wb2.close()

        os.chdir(path.parent.absolute())

    normalServic(*excelsheetname)


# collection(author,collection_name,heading,excelsheetname1)


author = 'anderson22'
universe = 'Dragonslair'
heading = "{} Collection".format(universe)
collection_name = collection_name1
collection(author, collection_name, heading, excelsheetname5)
collection_name = collection_name2
time.sleep(1)
collection(author, collection_name, heading, excelsheetname2)
collection_name = collection_name3
time.sleep(1)
collection(author, collection_name, heading, excelsheetname3)
collection_name = collection_name5
time.sleep(1)
collection(author, collection_name, heading, excelsheetname6)


'''
author = 'protoverse21'
universe = 'Havas'
heading = "{} Collection".format(universe)
collection_name = '311251121121'
collection6 = 'Havas'
excelsheetname1 = "{}.xlsx".format(collection6)

collection(author, collection_name, heading, excelsheetname1)
'''
