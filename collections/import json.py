import json
import os
import pandas as pd
import requests as requests
import pathlib
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import time

print(
    "Thank you for running the programe\nIt will take a few seconds to create everything\nHope you enjoy  \n-kevcollector"
)
# set what command you want to run here
# os.getenv('PWD')
# mac=pathlib.Path().cwd() /'Desktop'
# os.chdir(mac)
# os.system('cls' if os.name == 'nt' else 'clear')
# userCollection = input("Please enter the collection you want to scan")
# os.system('cls' if os.name == 'nt' else 'clear')


def writeToExcel(worksheet, data):
    for r in dataframe_to_rows(data, index=False):
        worksheet.append(r)
    dims = {}
    for row in worksheet.rows:
        for cell in row:
            if cell.value:
                dims[cell.column_letter] = max(
                    (dims.get(cell.column_letter, 0), len(str(cell.value)))
                )
    for col, value in dims.items():
        worksheet.column_dimensions[col].width = value


current = pathlib.Path().cwd()

totalholderslist = []


def collection(author, collection_name, heading, *excelsheetname):
    collecion = "".join(excelsheetname)
    collecion = collecion.replace(".xlsx", "")

    path = pathlib.Path().cwd() / ("{}".format(heading))
    if current != pathlib.Path().cwd():
        path = pathlib.Path().cwd()
        pathlib.Path(path).mkdir(parents=True, exist_ok=True)
    else:
        pathlib.Path(path).mkdir(parents=True, exist_ok=True)

    holders = " "

    holder_list = []

    def normalServic(holders, *excelsheetname):
        global totalholderslist
        os.chdir(path)

        wb1 = Workbook()
        ws1 = wb1.active
        ws1.title = "holders"

        holders = (
            "https://proton.api.atomicassets.io/atomicassets/v1/accounts?collection_name={}"
            "&page=1&limit=100&order=desc".format(collection_name)
        )
        holders = requests.get(holders).text
        holders_ = json.loads(holders)
        pages = 1
        while len(holders_["data"]) != 0:
            pages = pages + 1
            for data_info in holders_["data"]:
                holders = int(data_info["assets"])
                authorsName = data_info["account"]
                if authorsName != :
                    holder_list.append([data_info["account"], holders])
            holders = (
                "https://proton.api.atomicassets.io/atomicassets/v1/accounts?collection_name={}"
                "&page={}&limit=100&order=desc".format(collection_name, pages)
            )
            holders = requests.get(holders).text
            holders_ = json.loads(holders)
        holder_df = pd.DataFrame(data=holder_list, columns=["account ", "amount held"])
        len(holder_df) - 1
        count = 0
        rowz = 1
        for r in dataframe_to_rows(holder_df, index=False):
            ws1.append(r)
        
        pages = 0
        holders = (
            "https://proton.api.atomicassets.io/atomicassets/v1/accounts?collection_name={}"
            "&page=1&limit=100&order=desc".format(collection_name)
        )
        holders = requests.get(holders).text
        holders_ = json.loads(holders)
        amount = 1

        gorillaCount = 0
        while len(holders_["data"]) != 0:
            pages = pages + 1
            amount = amount + 1
            for data_info in holders_["data"]:
                if data_info["account"] != "delaneycb":
                    checker = data_info["account"]
                    if rowz != 1:
                        temp2 = ws1.cell(row=rowz, column=3).value
                    print("getting {}'s data".format(checker))
                    people = "https://proton.api.atomicassets.io/atomicmarket/v1/assets?collection_name={}&owner={}&page=1&limit=100&order=desc&sort=asset_id".format(
                        collection_name, checker
                    )
                    test = requests.get(people)
                    next = test.headers["X-RateLimit-Reset"]
                    resset = test.headers["X-RateLimit-Remaining"]
                    resset = int(resset)
                    next = int(next)
                    wait = next - time.time()
                    if resset < 3:
                        time.sleep(wait)
                    people_ = json.loads((test.text))
                    time.sleep(0.2)
                    count = 0
                    pages = 1
                    rowz = rowz + 1
                    assitID2 = " "

                    while len(people_["data"]) != 0:
                        pages = pages + 1
                        if pages >= 3:
                            count =  pages * 100
                        for data_info in people_["data"]:
                            word = data_info["data"]["desc"]
                            nft_name = data_info["data"]["name"]
                            number_of_nft = data_info["template_mint"]
                            assitID1 = data_info["asset_id"]
                            if assitID1 != assitID2:
                                assitID2 = assitID1
                                done = nft_name
                                if not "GGIP Proton Gorilla" in nft_name:
                                    count = count + 1

                                    ws1.cell(row=rowz, column=3 + count).value = (
                                        done + " (#" + number_of_nft + ")"
                                    )

                                else:
                                    gorillaCount += 1
                                    totalnftnumber = ws1.cell(row=rowz, column=2).value
                                    totalnftnumber = int(totalnftnumber) - 1
                                    ws1.cell(row=rowz, column=2).value = int(
                                        totalnftnumber
                                    )

                        people = "https://proton.api.atomicassets.io/atomicmarket/v1/assets?collection_name={}&owner={}&page={}&limit=100&order=desc&sort=asset_id".format(
                            collection_name, checker, pages
                        )
                        test = requests.get(people)
                        next = test.headers["X-RateLimit-Reset"]
                        resset = test.headers["X-RateLimit-Remaining"]
                        resset = int(resset)
                        next = int(next)
                        wait = next - time.time()
                        if resset < 3:
                            time.sleep(wait)
                        people_ = json.loads((test.text))
                        time.sleep(0.2)

                    holders_amount = ws1.cell(row=rowz, column=2).value

                    totalholderslist.append([checker, holders_amount])

            holders = (
                "https://proton.api.atomicassets.io/atomicassets/v1/accounts?collection_name={}"
                "&page={}&limit=100&order=desc".format(collection_name, amount)
            )
            holders = requests.get(holders).text
            holders_ = json.loads(holders)

        count = 0
        dims = {}
        for row in ws1.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column_letter] = max(
                        (dims.get(cell.column_letter, 0), len(str(cell.value)))
                    )
        for col, value in dims.items():
            ws1.column_dimensions[col].width = value + 5

        excelsave = "".join(excelsheetname)
        wb1.save(excelsave)
        print("Creating the excel file")
        wb1.close()
        os.chdir(path.parent.absolute())

    normalServic(holders, *excelsheetname)


author = "ggip"
universe = "GGIP + PLC"
heading = "{} Collection".format(universe)
collection_name = "241115151314"
collection1 = "GGIP"
excelsheetname1 = "{}.xlsx".format(collection1)
time.sleep(6)
collection(author, collection_name, heading, excelsheetname1)
collection_name = "kxlulfrvzsdd"
collection2 = "PLC"
excelsheetname1 = "{}.xlsx".format(collection2)
time.sleep(6)
collection(author, collection_name, heading, excelsheetname1)
wb2 = Workbook()
ws1 = wb2.active
ws1 = wb2.create_sheet("holders")
holders_df = pd.DataFrame(data=totalholderslist, columns=["account", "amount held"])

print(totalholderslist)
holders_df = holders_df.groupby(["account"]).agg(
    {"account": "min", "amount held": "sum"}
)
holders_df = holders_df.sort_values(by=["amount held"], ascending=False)


writeToExcel(ws1, holders_df)
