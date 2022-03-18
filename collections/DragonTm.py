
import json
import os

import pandas as pd
import requests as requests
import pathlib
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import time
print(
    "Thank you for running the programe\nIt will take a few seconds to create everything\nHope you enjoy  \n-kevcollector")
# set what command you want to run here
# os.getenv('PWD')
# mac=pathlib.Path().cwd() /'Desktop'
# os.chdir(mac)

author = 'dragontm'
universe = 'DRAGONtm Creations'
collection1 = 'DRAGONtm'

heading = "{} Collection".format(universe)
collection_name = '445323453535'  # dragontm

excelsheetname1 = "{}.xlsx".format(collection1)  # dra

current = pathlib.Path().cwd()
Royal = 0.1


def collection(author, collection_name, heading, *excelsheetname):
    global Royal
    collecion = "".join(excelsheetname)
    collecion = collecion.replace('.xlsx', '')

    path = pathlib.Path().cwd() / ("{}".format(heading))
    if (current != pathlib.Path().cwd()):
        path = pathlib.Path().cwd()
        pathlib.Path(path).mkdir(parents=True, exist_ok=True)
    else:
        pathlib.Path(path).mkdir(parents=True, exist_ok=True)

    resells = ("https://proton.api.atomicassets.io/atomicmarket/v1/sales?state=1%2C3&seller_blacklist"
               "={}&buyer_blacklist={}&collection_name={}&page=1&limit=100&order"
               "=desc&sort=updated".format(author, author, collection_name))
    print("getting the resells {}".format(collecion))
    authors = ("https://proton.api.atomicassets.io/atomicmarket/v1/sales?state=3&account={}"
               "&collection_name={}&page=1&limit=100&order=desc&sort=updated".format(author, collection_name))
    print("getting the first sales {}".format(collecion))
    holders = ("https://proton.api.atomicassets.io/atomicassets/v1/accounts?collection_name={}"
               "&page=1&limit=100&order=desc".format(collection_name))
    parents = requests.get(
        "https://proton.api.atomicassets.io/atomicassets/v1/assets?collection_name={}&page=1&limit=100&order"
        "=desc&sort=asset_id".format(collection_name)).text
    authors = requests.get(authors).text
    authors_ = json.loads(authors)
    name = []
    all = requests.get(resells).text

    holders = requests.get(holders).text

    parents_ = json.loads(parents)
    all_ = json.loads(all)
    holders_ = json.loads(holders)

    resales = "Resells"
    FirstSale = "First sale"
    Holders = "Holders"

    holder_list = []
    authors_list = []
    all_list = []
    parents_list = []

    def normalServic(authors, all, resales, FirstSale, Holders, holders, *excelsheetname):
        global Royal
        os.chdir(path)
        if collection_name != '445323453535':
            authors = ("https://proton.api.atomicassets.io/atomicmarket/v1/sales?state=3&account={}"
                       "&collection_name={}&page=1&limit=100&order=desc&sort=updated".format(author, collection_name))
            authors = requests.get(authors).text
            authors_ = json.loads(authors)
            time.sleep(4)
            print("getting First sales")
            while len(authors_['data']) != 0:

                for data_info in authors_['data']:

                    number = data_info["listing_price"]
                    fixed = int(number) / 1000000
                    name = data_info['assets'][0]['name']
                    timez = data_info['assets'][0]['transferred_at_time']
                    timef = data_info['updated_at_time']
                    timex = int(timez) / 1000
                    number_of_nft = int(
                        data_info['assets'][0]['template_mint'])
                    buyer = data_info['buyer']
                    seller = data_info['seller']

                    local_time = datetime.utcfromtimestamp(
                        timex).strftime('%d-%m-%Y %H:%M:%S')

                    authors_list.append(
                        [seller, fixed, buyer, number_of_nft, name, local_time])
                time.sleep(0.2)
                authors = ("https://proton.api.atomicassets.io/atomicmarket/v1/sales?state=3&account={}"
                           "&collection_name={}&before={}&page=1&limit=100&order=desc&sort=updated".format(author,
                                                                                                           collection_name,
                                                                                                           timef))
                authors = requests.get(authors).text
                authors_ = json.loads(authors)
            name_df = pd.DataFrame(data=authors_list,
                                   columns=["author ", "price listed usd", "buyer", "# of nft", "name", "time"])
            total = name_df['price listed usd'].sum()
            name_df.at['Total',
                       'price listed usd'] = name_df['price listed usd'].sum()

            wb = Workbook()

            ws = wb.active
            ws.title = (FirstSale)
            ws2 = wb.create_sheet(resales)
            ws3 = wb.create_sheet(Holders)
            for r in dataframe_to_rows(name_df, index=False):
                ws.append(r)
            dims = {}
            for row in ws.rows:
                for cell in row:
                    if cell.value:
                        dims[cell.column_letter] = max(
                            (dims.get(cell.column_letter, 0), len(str(cell.value))))
            for col, value in dims.items():
                ws.column_dimensions[col].width = value
            maxrow = ws.max_row
            ws.cell(row=maxrow, column=1, value="totals")

            resells = ("https://proton.api.atomicassets.io/atomicmarket/v1/sales?state=1%2C3&seller_blacklist"
                       "={}&buyer_blacklist={}&collection_name={}&page=1&limit=100&order"
                       "=desc&sort=updated".format(author, author, collection_name))
            all = requests.get(resells).text
            all_ = json.loads(all)
            print("getting resales")
            while len(all_['data']) != 0:
                for data_info in all_['data']:
                    number = data_info["listing_price"]
                    fixed = int(number) / 1000000
                    name = data_info['assets'][0]['name']
                    timez = data_info['assets'][0]['transferred_at_time']
                    timef = timef = data_info['updated_at_time']
                    number_of_nft = int(
                        data_info['assets'][0]['template_mint'])
                    Royal = data_info['collection']['market_fee']
                    timefixe = int(timez) / 1000
                    buyer = data_info['buyer']
                    seller = data_info['seller']
                    local_time = datetime.utcfromtimestamp(
                        timefixe).strftime('%d-%m-%Y %H:%M:%S')
                    all_list.append(
                        [seller, buyer, fixed, name, number_of_nft, local_time])
                time.sleep(.6)

                resells = ("https://proton.api.atomicassets.io/atomicmarket/v1/sales?state=1%2C3&seller_blacklist"
                           "={}&buyer_blacklist={}&collection_name={}&before={}&page=1&limit=100&order"
                           "=desc&sort=updated".format(author, author, collection_name, timef))
                all = requests.get(resells).text
                all_ = json.loads(all)

            if len(all_list) != 0:
                names_df = pd.DataFrame(data=all_list,
                                        columns=["first buyer ", "next buyer", "price paid usd", "name", "# of nft", "time"])
                names_df.drop(
                    names_df[names_df["first buyer "] == f"{author}"].index, inplace=True)
                totals = names_df['price paid usd'].sum()
                Royalties = (totals * Royal)
                Rows = int(names_df.index.max() + 1)
                names_df.at[Rows, 'price paid usd'] = Royalties

                for r in dataframe_to_rows(names_df, index=False):
                    ws2.append(r)
                dims = {}
                for row in ws2.rows:
                    for cell in row:
                        if cell.value:
                            dims[cell.column_letter] = max(
                                (dims.get(cell.column_letter, 0), len(str(cell.value))))
                for col, value in dims.items():
                    ws2.column_dimensions[col].width = value
                maxrow = ws2.max_row
                ws2.cell(row=maxrow, column=1, value="Royalties")
            if len(holders_) != 0:
                for data_info in holders_['data']:
                    holders = int(data_info['assets'])
                    holder_list.append([data_info['account'], holders])
                dino_holder_df = pd.DataFrame(data=holder_list, columns=[
                                              "account ", "amount held"])
                # print(dino_holder_df)
                len(dino_holder_df) - 1
                count = 0
                rowz = 1
                Peoplelist = []
                for r in dataframe_to_rows(dino_holder_df, index=False):
                    ws3.append(r)
                dims = {}
                com = 0
                rare = 0
                epic = 0
                assitID2 = " "
                for data_info in holders_['data']:
                    checker = (data_info['account'])
                    print("getting {}'s data".format(checker))
                    people = "https://proton.api.atomicassets.io/atomicmarket/v1/assets?collection_name={}&owner={}&page=1&limit=100&order=desc&sort=asset_id".format(
                        collection_name, checker)
                    test = requests.get(people)
                    next = test.headers['X-RateLimit-Reset']
                    resset = test.headers['X-RateLimit-Remaining']
                    resset = int(resset)
                    next = int(next)
                    wait = next-time.time()
                    if resset < 3:
                        time.sleep(wait)
                    people_ = json.loads((test.text))
                    time.sleep(0.2)
                    count = 0
                    com = 0
                    rare = 0
                    epic = 0
                    rowz = rowz + 1
                    for data_info in people_["data"]:
                        nft_NAR = (data_info["data"]["name"])
                        assitID1 = data_info["asset_id"]
                        number_of_nft = data_info['template_mint']
                        if assitID1 != assitID2:
                            assitID2 = assitID1
                            nft_NAR1 = nft_NAR.replace('(', '')
                            nft_NAR2 = nft_NAR1.replace(')', '')
                            done = nft_NAR2
                            count = count + 1
                            ws3.cell(row=rowz, column=3 +
                                     count).value = done+' (#'+number_of_nft+')'
                            if "common" in done:
                                com = com + 1
                            if "rare" in done:
                                rare = rare + 1
                            if "epic" in done:
                                epic = epic + 1

                for row in ws3.rows:
                    for cell in row:
                        if cell.value:
                            dims[cell.column_letter] = max(
                                (dims.get(cell.column_letter, 0), len(str(cell.value))))
                for col, value in dims.items():
                    ws3.column_dimensions[col].width = value + 5

            if (collection_name == '133523522522'):

                ws4 = wb.create_sheet("Parents")
                for data_info in parents_['data']:

                    word = data_info['data']['desc']
                    if word.find('Parent 1:') != -1:
                        start = word.find('Parent 1')
                        start += 10
                        end = word.find('Parent 2')
                        end -= 1
                        new = word[start:end]
                        start2 = word.find('Parent 2')
                        start2 += 10
                        end1 = len(word)
                        new1 = word[start2:end1]

                        parents_list.append(
                            [data_info['data']['name'], new, new1])
                s_df = pd.DataFrame(data=parents_list, columns=[
                                    "Name ", "first parent", "second parent"])

                for r in dataframe_to_rows(s_df, index=False):
                    ws4.append(r)
                dimsz = {}
                for row in ws4.rows:
                    for cell in row:
                        if cell.value:
                            dimsz[cell.column_letter] = max(
                                (dimsz.get(cell.column_letter, 0), len(str(cell.value))))
                    for col, value in dimsz.items():
                        ws4.column_dimensions[col].width = value
                excelsave = "".join(excelsheetname)
            excelsave = "".join(excelsheetname)
            wb.save(excelsave)
            print("Creating the excel file")
            wb.close()
            os.chdir(path.parent.absolute())
        else:
            wb = Workbook()

            ws = wb.active
            ws.title = (resales)
            ws2 = wb.create_sheet(Holders)
            resells = ("https://proton.api.atomicassets.io/atomicmarket/v1/sales?state=1%2C3&seller_blacklist"
                       "={}&buyer_blacklist={}&collection_name={}&page=1&limit=100&order"
                       "=desc&sort=updated".format(author, author, collection_name))
            all = requests.get(resells).text
            all_ = json.loads(all)
            print("getting resales")
            while len(all_['data']) != 0:
                for data_info in all_['data']:
                    number = data_info["listing_price"]
                    fixed = int(number) / 1000000
                    name = data_info['assets'][0]['name']
                    timez = data_info['assets'][0]['transferred_at_time']
                    timef = timef = data_info['updated_at_time']
                    number_of_nft = int(
                        data_info['assets'][0]['template_mint'])
                    Royal = data_info['collection']['market_fee']
                    timefixe = int(timez) / 1000
                    buyer = data_info['buyer']
                    seller = data_info['seller']
                    local_time = datetime.utcfromtimestamp(
                        timefixe).strftime('%d-%m-%Y %H:%M:%S')
                    all_list.append(
                        [seller, buyer, fixed, name, number_of_nft, local_time])
                time.sleep(.6)

                resells = ("https://proton.api.atomicassets.io/atomicmarket/v1/sales?state=1%2C3&seller_blacklist"
                           "={}&buyer_blacklist={}&collection_name={}&before={}&page=1&limit=100&order"
                           "=desc&sort=updated".format(author, author, collection_name, timef))
                all = requests.get(resells).text
                all_ = json.loads(all)

            names_df = pd.DataFrame(data=all_list,
                                    columns=["first buyer ", "next buyer", "price paid usd", "name", "# of nft",
                                             "time"])
            names_df.drop(names_df[names_df["first buyer "]
                          == f"{author}"].index, inplace=True)
            totals = names_df['price paid usd'].sum()
            Royalties = (totals * Royal)
            Rows = int(names_df.index.max() + 1)
            names_df.at[Rows, 'price paid usd'] = Royalties

            for r in dataframe_to_rows(names_df, index=False):
                ws.append(r)
            dims = {}
            for row in ws.rows:
                for cell in row:
                    if cell.value:
                        dims[cell.column_letter] = max(
                            (dims.get(cell.column_letter, 0), len(str(cell.value))))
            for col, value in dims.items():
                ws.column_dimensions[col].width = value
            maxrow = ws2.max_row
            ws.cell(row=maxrow, column=1, value="Royalties")
            if len(holders_) != 0:
                for data_info in holders_['data']:
                    holders = int(data_info['assets'])
                    holder_list.append([data_info['account'], holders])
                dino_holder_df = pd.DataFrame(data=holder_list, columns=[
                                              "account ", "amount held"])
                # print(dino_holder_df)
                len(dino_holder_df) - 1
                count = 0
                rowz = 1
                Peoplelist = []
                for r in dataframe_to_rows(dino_holder_df, index=False):
                    ws2.append(r)
                dims = {}
                com = 0
                rare = 0
                epic = 0
                assitID2 = " "
            if (collection_name == '133523522522'):

                ws4 = wb.create_sheet("Parents")
                for data_info in parents_['data']:

                    word = data_info['data']['desc']
                    if word.find('Parent 1:') != -1:
                        start = word.find('Parent 1')
                        start += 10
                        end = word.find('Parent 2')
                        end -= 1
                        new = word[start:end]
                        start2 = word.find('Parent 2')
                        start2 += 10
                        end1 = len(word)
                        new1 = word[start2:end1]

                        parents_list.append(
                            [data_info['data']['name'], new, new1])
                s_df = pd.DataFrame(data=parents_list, columns=[
                                    "Name ", "first parent", "second parent"])

                for r in dataframe_to_rows(s_df, index=False):
                    ws4.append(r)
                dimsz = {}
                for row in ws4.rows:
                    for cell in row:
                        if cell.value:
                            dimsz[cell.column_letter] = max(
                                (dimsz.get(cell.column_letter, 0), len(str(cell.value))))
                    for col, value in dimsz.items():
                        ws4.column_dimensions[col].width = value
                excelsave = "".join(excelsheetname)
            excelsave = "".join(excelsheetname)
            wb.save(excelsave)
            print("Creating the excel file")
            wb.close()
            os.chdir(path.parent.absolute())

    normalServic(authors, all, resales, FirstSale,
                 Holders, holders, *excelsheetname)


author = 'dragontm'
universe = 'dragontm'
heading = "{} Collection".format(universe)
collection_name = '445323453535'
collection6 = 'DRAGONtm Creations'
excelsheetname1 = "{}.xlsx".format(collection6)

collection(author, collection_name, heading, excelsheetname1)
collection_name = '144321315411'
collection6 = 'Mythical Creatures'
excelsheetname1 = "{}.xlsx".format(collection6)
time.sleep(4)
collection(author, collection_name, heading, excelsheetname1)

collection_name = "113243453515"
collection6 = "Angery Dragons"
excelsheetname1 = "{}.xlsx".format(collection6)
time.sleep(4)
collection(author, collection_name, heading, excelsheetname1)
