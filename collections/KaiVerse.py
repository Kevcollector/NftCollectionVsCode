
import json
import os

import pandas as pd
import requests as requests
import pathlib
from openpyxl import Workbook
from openpyxl.styles import Font, Color, colors
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import time

from openpyxl.worksheet import hyperlink

print(
    "Thank you for running the programe\nIt will take a few seconds to create everything\nHope you enjoy  "
    "\n-kevcollector\n\n\n\n")

os.system('cls' if os.name == 'nt' else 'clear')
print("\t\tIn the intrest of speeding up the code you please select from the following list\n\t\tenter the number for the collection you want\n")
choice = input("1.\tDino\n2.\tBaby Dino\n3.\tBaby Dino Fussion\n4.\tSerpents\n5.\tKrok\n9. for all :")
os.system('cls' if os.name == 'nt' else 'clear')
print("\n\n\t\tenter the number for how holders are managed\n")
select = input("1.\tjust holders\n2.\tHolders with the nfts names :")
os.system('cls' if os.name == 'nt' else 'clear')
# set what command you want to run here
# os.getenv('PWD')
# mac=pathlib.Path().cwd() /'Desktop'
# os.chdir(mac)

auther = 'dragontm'
universe = 'DRAGONtm Creations'
collection1 = 'DRAGONtm'
collection2 = 'Baby Dino'
collection3 = 'Baby Dino (Fusions)'
collection4 = 'Sprout'
collection5 = 'Dino'
collection6 = 'serpent'
collection7 = 'Krok'
heading = "{} Collection".format(universe)
collection_name = '445323453535'  # dragontm
collection_name1 = '423534451325'  # Dino the dragon
collection_name2 = '513555243311'  # baby dino
collection_name3 = '133523522522'  # fussion
collection_name4 = '334431125343'  # sprouts
collection_name5 = '133552212214'  # serpent

excelsheetname1 = "{}.xlsx".format(collection1)  # dra
excelsheetname2 = "{}.xlsx".format(collection2)  # bab
excelsheetname3 = "{}.xlsx".format(collection3)  # fus
excelsheetname4 = "{}.xlsx".format(collection4)
excelsheetname5 = "{}.xlsx".format(collection5)  # dino
excelsheetname6 = "{}.xlsx".format(collection6)
excelsheetname7 = "{}.xlsx".format(collection7)

current = pathlib.Path().cwd()

global Royal
Royal=0.1

def collection(auther, collection_name, heading,select, *excelsheetname):
    global Royal

    collecion = "".join(excelsheetname)
    collecion = collecion.replace('.xlsx', '')

    path = pathlib.Path().cwd() / ("{}".format(heading))
    if (current != pathlib.Path().cwd()):
        path = pathlib.Path().cwd()
        pathlib.Path(path).mkdir(parents=True, exist_ok=True)
    else:
        pathlib.Path(path).mkdir(parents=True, exist_ok=True)

    resells = ("https://proton.api.atomicassets.io/atomicmarket/v1/sales?state=1%2C3&seller_blacklist"
               "={}&buyer_blacklist={}&collection_name={}&page=1&limit=100&order"
               "=desc&sort=updated".format(auther, auther, collection_name))
    print("getting the resells {}".format(collecion))
    authers = ("https://proton.api.atomicassets.io/atomicmarket/v1/sales?state=3&account={}"
               "&collection_name={}&page=1&limit=100&order=desc&sort=updated".format(auther, collection_name))
    print("getting the first sales {}".format(collecion))
    holders = ("https://proton.api.atomicassets.io/atomicassets/v1/accounts?collection_name={}"
               "&page=1&limit=100&order=desc".format(collection_name))
    parents = requests.get(
        "https://proton.api.atomicassets.io/atomicassets/v1/assets?collection_name={}&page=1&limit=100&order"
        "=desc&sort=asset_id".format(collection_name)).text
    listed=("https://proton.api.atomicassets.io/atomicmarket/v1/sales?state=1&collection_name={}&page=1&limit=100&order=asc&sort=price".format(collection_name))
    listed=requests.get(listed)
    listed_=json.loads(listed.text)
    listed_done=[]


    authers = requests.get(authers).text
    authers_ = json.loads(authers)
    name = []
    all = requests.get(resells).text

    holders = requests.get(holders).text

    parents_ = json.loads(parents)
    all_ = json.loads(all)
    holders_ = json.loads(holders)

    resales = "Resells"
    FirstSale = "First sale"
    Holders = "Holders"

    holder_list = []
    authers_list = []
    all_list = []
    parents_list = []

    def normalServic(authers, all, resales, FirstSale, Holders, holders, select,*excelsheetname):
        global Royal
        os.chdir(path)
        authers = ("https://proton.api.atomicassets.io/atomicmarket/v1/sales?state=3&account={}"
                   "&collection_name={}&page=1&limit=100&order=desc&sort=updated".format(auther, collection_name))
        authers = requests.get(authers).text
        authers_ = json.loads(authers)
        print("getting First sales")
        while len(authers_['data']) !=0:

            for data_info in authers_['data']:

                number = data_info["listing_price"]
                fixed = int(number) / 1000000
                name = data_info['assets'][0]['name']
                timez = data_info['assets'][0]['transferred_at_time']
                timef=data_info['updated_at_time']
                timex = int(timez) / 1000
                number_of_nft = int(data_info['assets'][0]['template_mint'])
                buyer = data_info['buyer']
                seller = data_info['seller']

                local_time = datetime.utcfromtimestamp(timex).strftime('%d-%m-%Y %H:%M:%S')

                authers_list.append([seller, fixed, buyer, number_of_nft, name, local_time])
            time.sleep(0.4)
            authers = ("https://proton.api.atomicassets.io/atomicmarket/v1/sales?state=3&account={}"
                       "&collection_name={}&before={}&page=1&limit=100&order=desc&sort=updated".format(auther,
                                                                                                      collection_name,
                                                                                                      timef))
            authers = requests.get(authers).text
            authers_ = json.loads(authers)
        name_df = pd.DataFrame(data=authers_list,
                               columns=["author ", "price listed usd", "buyer", "# of nft", "name", "time"])
        total = name_df['price listed usd'].sum()
        name_df.at['Total', 'price listed usd'] = name_df['price listed usd'].sum()

        wb = Workbook()

        ws = wb.active
        ws.title = (FirstSale)
        ws2 = wb.create_sheet(resales)
        ws3 = wb.create_sheet(Holders)
        for r in dataframe_to_rows(name_df, index=False):
            ws.append(r)
        dims = {}
        for row in ws.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
        for col, value in dims.items():
            ws.column_dimensions[col].width = value
        maxrow = ws.max_row
        ws.cell(row=maxrow, column=1, value="totals")

        resells = ("https://proton.api.atomicassets.io/atomicmarket/v1/sales?state=1%2C3&seller_blacklist"
                   "={}&buyer_blacklist={}&collection_name={}&page=1&limit=100&order"
                   "=desc&sort=updated".format(auther, auther, collection_name))
        auctions=("https://proton.api.atomicassets.io/atomicmarket/v1/auctions?state=3&collection_name={}&page=1&limit=100&order=desc&sort=created".format(collection_name))
        acutionss=requests.get(auctions).text
        acutions_=json.loads(acutionss)
        auctions_list=[]
        all = requests.get(resells).text
        all_ = json.loads(all)
        print("getting resales")
        while len(all_['data']) != 0:
            for data_info in all_['data']:
                number = data_info["listing_price"]
                fixed = int(number) / 1000000
                name = data_info['assets'][0]['name']
                timez = data_info['assets'][0]['transferred_at_time']
                timef=data_info['updated_at_time']
                number_of_nft = int(data_info['assets'][0]['template_mint'])
                Royal=data_info['collection']['market_fee']
                timefixe = int(timez) / 1000
                buyer = data_info['buyer']
                seller = data_info['seller']
                local_time = datetime.utcfromtimestamp(timefixe).strftime('%d-%m-%Y %H:%M:%S')
                all_list.append([seller, buyer, fixed, name, number_of_nft, local_time])
            time.sleep(.6)

            resells = ("https://proton.api.atomicassets.io/atomicmarket/v1/sales?state=1%2C3&seller_blacklist"
                       "={}&buyer_blacklist={}&collection_name={}&before={}&page=1&limit=100&order"
                       "=desc&sort=updated".format(auther, auther, collection_name,timef))
            all = requests.get(resells).text
            all_ = json.loads(all)




            for data_info in acutions_['data']:
                number = data_info["price"]["amount"]
                fixed = int(number) / 1000000
                name = data_info['assets'][0]['name']
                timez = data_info['assets'][0]['transferred_at_time']
                timef = data_info['updated_at_time']
                number_of_nft = int(data_info['assets'][0]['template_mint'])
                Royal = data_info['collection']['market_fee']
                timefixe = int(timez) / 1000
                buyer = data_info['buyer']
                seller = data_info['seller']
                local_time = datetime.utcfromtimestamp(timefixe).strftime('%d-%m-%Y %H:%M:%S')
                auctions_list.append([seller, buyer, fixed, name, number_of_nft, local_time])



        auctions_df=pd.DataFrame(data=auctions_list,
                                columns=["first buyer ", "next buyer", "price paid usd", "name", "# of nft", "time"])
        names_df = pd.DataFrame(data=all_list,
                                columns=["first buyer ", "next buyer", "price paid usd", "name", "# of nft", "time"])
        names_df.drop(names_df[names_df["first buyer "] == f"{auther}"].index, inplace=True)
        totals = names_df['price paid usd'].sum()
        Royalties = (totals * Royal)
        try:
            Rows = int(names_df.index.max() + 1)
        except:
            Rows=1
        names_df.at[Rows, 'price paid usd'] = Royalties
        count=0
        for r in dataframe_to_rows(names_df, index=False):
            ws2.append(r)
            count=count+1

        dims = {}
        for row in ws2.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
        for col, value in dims.items():
            ws2.column_dimensions[col].width = value
        maxrow = ws2.max_row
        ws2.cell(row=maxrow, column=1, value="Royalties")

        ws2.cell(row=count + 3, column=1 ,value="Auctions")
        for r in dataframe_to_rows(auctions_df, index=False):
            ws2.append(r)
        for row in ws2.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
        for col, value in dims.items():
            ws2.column_dimensions[col].width = value


        for data_info in holders_['data']:
            holders = int(data_info['assets'])
            holder_list.append([data_info['account'], holders])
        dino_holder_df = pd.DataFrame(data=holder_list, columns=["account ", "amount held"])
        # print(dino_holder_df)
        len(dino_holder_df) - 1
        count = 0
        rowz = 1
        Peoplelist = []
        for r in dataframe_to_rows(dino_holder_df, index=False):
            ws3.append(r)
        dims = {}
        com = 0
        rare = 0
        epic = 0
        assitID2 = " "
        if select=='2':
            for data_info in holders_['data']:
                checker = (data_info['account'])
                print("getting {}'s data".format(checker))
                people = requests.get(
                    "https://proton.api.atomicassets.io/atomicmarket/v1/assets?collection_name={}&owner={}&page=1&limit=100&order=desc&sort=asset_id".format(
                        collection_name,checker)).text
                people_ = json.loads(people)
                time.sleep(0.6)
                count = 0
                com = 0
                rare = 0
                epic = 0
                rowz = rowz + 1
                for data_info in people_["data"]:
                    nft_NAR = (data_info["data"]["name"])
                    number_of_nft = data_info['template_mint']
                    assitID1 = data_info["asset_id"]
                    if assitID1 != assitID2:
                        assitID2 = assitID1
                        nft_NAR1 = nft_NAR.replace('(', '')
                        nft_NAR2 = nft_NAR1.replace(')', '')
                        done = nft_NAR2
                        count = count + 1
                        ws3.cell(row=rowz, column=3 + count).value = done+' (#'+number_of_nft+')'
                        if "common" in done:
                            com = com + 1
                        if "rare" in done:
                            rare = rare + 1
                        if "epic" in done:
                            epic = epic + 1

        for row in ws3.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
        for col, value in dims.items():
            ws3.column_dimensions[col].width = value + 5

        listed = (
            "https://proton.api.atomicassets.io/atomicmarket/v1/sales?state=1&collection_name={}&page=1&limit=100&order=asc&sort=updated".format(
                collection_name))
        listed = requests.get(listed)
        listed_ = json.loads(listed.text)
        wb2 = Workbook()

        wsL = wb2.active
        wsL.title="Cheapest"
        count=0
        while len(listed_['data']) !=0:
            for data_info in listed_['data']:
                seller = data_info['seller']
                number = data_info["listing_price"]
                fixed = int(number) / 1000000
                name = data_info['assets'][0]['name']
                ids = data_info['assets'][0]['template']['template_id']
                number_of_nft = int(data_info['assets'][0]['template_mint'])
                timef = data_info['updated_at_time']
                amount=data_info['assets'][0]['template']['max_supply']
                amount=amount+"/"+amount
                link="https://protonmint.com/{}/{}".format(collection_name,ids)
                listed_done.append([seller, fixed, name, number_of_nft,amount,link])
                count = count + 1

            time.sleep(0.4)
            listed = (
                "https://proton.api.atomicassets.io/atomicmarket/v1/sales?state=1&collection_name={}&after={}&page=1&limit=100&order=asc&sort=updated".format(
                    collection_name,timef))
            listed = requests.get(listed)
            listed_ = json.loads(listed.text)
        listed_df = pd.DataFrame(data=listed_done,
                                columns=["Seller", "Price", "Name", "# of NFT","Edition Size","Link"])
        listed_df=listed_df.sort_values(by=['Price'], ascending=True)
        count=0

        for r in dataframe_to_rows(listed_df, index=False):
            count=count+1
            wsL.append(r)
            if count > 2:
                link=wsL.cell(row=count-1, column=6).value
                wsL.cell(row=count-1, column=6).value = '=HYPERLINK("{}", "{}")'.format(link,link)
                wsL.cell(row=count-1, column=6).style = 'Hyperlink'
        count=count+1
        link = wsL.cell(row=count - 1, column=6).value
        wsL.cell(row=count - 1, column=6).value = '=HYPERLINK("{}", "{}")'.format(link, link)
        wsL.cell(row=count - 1, column=6).style = 'Hyperlink'

        dims = {}
        for row in wsL.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
        for col, value in dims.items():
            wsL.column_dimensions[col].width = value



        excelsave = "".join(excelsheetname)
        wb2.save('Listed '+excelsave)
        print("Creating the excel file")
        wb2.close()




        if (collection_name == '133523522522'):
            count=0
            ws5 = wb.create_sheet("Parents")
            parents = requests.get(
                "https://proton.api.atomicassets.io/atomicassets/v1/assets?collection_name={}&page=1&limit=100&order"
                "=desc&sort=updated".format(collection_name)).text
            parents_ = json.loads(parents)
            while len(parents_['data']) !=0:
                count=1+count
                for data_info in parents_['data']:
                    timef = data_info['updated_at_time']

                    word = data_info['data']['desc']
                    if word.find('Parent 1:') != -1:
                        start = word.find('Parent 1')
                        start += 10
                        end = word.find('Parent 2')
                        end -= 1
                        new = word[start:end]
                        start2 = word.find('Parent 2')
                        start2 += 10
                        end1 = len(word)
                        new1 = word[start2:end1]

                        parents_list.append([data_info['data']['name'], new, new1])
                time.sleep(0.4)
                parents = requests.get(
                    "https://proton.api.atomicassets.io/atomicassets/v1/assets?collection_name={}&before={}&page=1&limit=100&order"
                    "=desc&sort=updated".format(collection_name, timef)).text
                parents_ = json.loads(parents)
            s_df = pd.DataFrame(data=parents_list, columns=["Name ", "first parent", "second parent"])


            for r in dataframe_to_rows(s_df, index=False):
                ws5.append(r)
            dimsz = {}
            for row in ws5.rows:
                for cell in row:
                    if cell.value:
                        dimsz[cell.column_letter] = max((dimsz.get(cell.column_letter, 0), len(str(cell.value))))
                for col, value in dimsz.items():
                    ws5.column_dimensions[col].width = value
            excelsave = "".join(excelsheetname)
        excelsave = "".join(excelsheetname)
        wb.save(excelsave)
        print("Creating the excel file")
        wb.close()
        os.chdir(path.parent.absolute())

    normalServic(authers, all, resales, FirstSale, Holders, holders,select, *excelsheetname)


# collection(auther,collection_name,heading,excelsheetname1)


auther = 'anderson22'
universe = 'Dragonslair'
heading = "{} Collection".format(universe)
collection_name = collection_name1
if choice=='1' or choice== '9':
    collection(auther, collection_name, heading,select, excelsheetname5)
if choice=='2' or choice== '9':
    collection_name = collection_name2
    if choice == '9':
        time.sleep(4)
    collection(auther, collection_name, heading, select,excelsheetname2)
if choice=='3' or choice== '9':
    collection_name = collection_name3
    if choice == '9':
        time.sleep(4)
    collection(auther, collection_name, heading,select, excelsheetname3)
if choice=='4' or choice== '9':
    collection_name = collection_name5
    if choice == '9':
        time.sleep(4)
    collection(auther, collection_name, heading,select, excelsheetname6)
if choice=='5' or choice== '9':
    collection_name = '355422121445'
    if choice == '9':
        time.sleep(4)
    collection(auther, collection_name, heading,select, excelsheetname7)












'''
auther = 'protoverse21'
universe = 'Havas'
heading = "{} Collection".format(universe)
collection_name = '311251121121'
collection6 = 'Havas'
excelsheetname1 = "{}.xlsx".format(collection6)

collection(auther, collection_name, heading, excelsheetname1)
'''