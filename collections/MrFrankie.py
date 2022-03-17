
import json
import os

import pandas as pd
import requests as requests
import pathlib
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import time
print(
    "Thank you for running the programme\nIt will take a few seconds to create everything\nHope you enjoy  \n-kevcollector")
# set what command you want to run here
# os.getenv('PWD')
# mac=pathlib.Path().cwd() /'Desktop'
# os.chdir(mac)


current = pathlib.Path().cwd()


def collection(auther, collection_name, heading, *excelsheetname):
    collecion = "".join(excelsheetname)
    collecion = collecion.replace('.xlsx', '')

    path = pathlib.Path().cwd() / ("{}".format(heading))
    if (current != pathlib.Path().cwd()):
        path = pathlib.Path().cwd()
        pathlib.Path(path).mkdir(parents=True, exist_ok=True)
    else:
        pathlib.Path(path).mkdir(parents=True, exist_ok=True)

    resells = ("https://proton.api.atomicassets.io/atomicmarket/v1/sales?state=1%2C3&seller_blacklist"
               "={}&buyer_blacklist={}&collection_name={}&page=1&limit=100&order"
               "=desc&sort=updated".format(auther, auther, collection_name))
    print("getting the resells {}".format(collecion))
    authers = ("https://proton.api.atomicassets.io/atomicmarket/v1/sales?state=3&account={}"
               "&collection_name={}&page=1&limit=100&order=desc&sort=updated".format(auther, collection_name))
    print("getting the first sales {}".format(collecion))
    holders = ("https://proton.api.atomicassets.io/atomicassets/v1/accounts?collection_name={}"
               "&page=1&limit=100&order=desc".format(collection_name))
    parents = requests.get(
        "https://proton.api.atomicassets.io/atomicassets/v1/assets?collection_name={}&page=1&limit=100&order"
        "=desc&sort=asset_id".format(collection_name)).text
    authers = requests.get(authers).text
    authers_ = json.loads(authers)
    name = []
    all = requests.get(resells).text

    holders = requests.get(holders).text

    parents_ = json.loads(parents)
    all_ = json.loads(all)
    holders_ = json.loads(holders)

    resales = "Resells"
    FirstSale = "First sale"
    Holders = "Holders"

    holder_list = []
    authers_list = []
    all_list = []
    parents_list = []

    def normalServic(authers, all, resales, FirstSale, Holders, holders, *excelsheetname):
        os.chdir(path)
        time.sleep(3)
        authers = ("https://proton.api.atomicassets.io/atomicmarket/v1/sales?state=3&account={}"
                   "&collection_name={}&page=1&limit=100&order=desc&sort=updated".format(auther, collection_name))
        authers = requests.get(authers).text
        authers_ = json.loads(authers)
        print("getting First sales")
        fixedX=0
        fixedC=0
        while len(authers_['data']) !=0:

            for data_info in authers_['data']:
                fixedC = 0
                fixedX = 0
                Type=data_info["listing_symbol"]
                if Type =="XPR":
                    number = data_info["listing_price"]
                    fixedX = int(number) / 10000

                if Type=="XUSDC":
                    number = data_info["listing_price"]
                    fixedC = int(number) / 1000000
                name = data_info['assets'][0]['name']
                timez = data_info['assets'][0]['transferred_at_time']
                timef=data_info['updated_at_time']
                timex = int(timez) / 1000
                number_of_nft = int(data_info['assets'][0]['template_mint'])
                buyer = data_info['buyer']
                seller = data_info['seller']

                local_time = datetime.utcfromtimestamp(timex).strftime('%d-%m-%Y %H:%M:%S')

                authers_list.append([seller, fixedC,fixedX, buyer, number_of_nft, name, local_time])
            time.sleep(0.4)
            authers = ("https://proton.api.atomicassets.io/atomicmarket/v1/sales?state=3&account={}"
                       "&collection_name={}&before={}&page=1&limit=100&order=desc&sort=updated".format(auther,
                                                                                                      collection_name,
                                                                                                      timef))
            authers = requests.get(authers).text
            authers_ = json.loads(authers)
        name_df = pd.DataFrame(data=authers_list,
                               columns=["author ", "price listed usd","price listed XPR", "buyer", "# of nft", "name", "time"])
        total = name_df['price listed usd'].sum()
        name_df.at['Total', 'price listed usd'] = name_df['price listed usd'].sum()

        wb = Workbook()

        ws = wb.active
        ws.title = (FirstSale)
        ws2 = wb.create_sheet(resales)
        ws3 = wb.create_sheet(Holders)
        for r in dataframe_to_rows(name_df, index=False):
            ws.append(r)
        dims = {}
        for row in ws.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
        for col, value in dims.items():
            ws.column_dimensions[col].width = value
        maxrow = ws.max_row
        ws.cell(row=maxrow, column=1, value="totals")

        resells = ("https://proton.api.atomicassets.io/atomicmarket/v1/sales?state=1%2C3&seller_blacklist"
                   "={}&buyer_blacklist={}&collection_name={}&page=1&limit=100&order"
                   "=desc&sort=updated".format(auther, auther, collection_name))
        all = requests.get(resells).text
        all_ = json.loads(all)
        print("getting resales")
        while len(all_['data']) != 0:
            for data_info in all_['data']:
                fixedC=0
                fixedX=0
                Type = data_info["listing_symbol"]
                if Type == "XPR":
                    number = data_info["listing_price"]
                    fixedX = int(number) / 10000

                if Type == "XUSDC":
                    number = data_info["listing_price"]
                    fixedC = int(number) / 1000000
                name = data_info['assets'][0]['name']
                timez = data_info['assets'][0]['transferred_at_time']
                timef=timef=data_info['updated_at_time']
                number_of_nft = int(data_info['assets'][0]['template_mint'])
                Royal=data_info['collection']['market_fee']
                timefixe = int(timez) / 1000
                buyer = data_info['buyer']
                seller = data_info['seller']
                local_time = datetime.utcfromtimestamp(timefixe).strftime('%d-%m-%Y %H:%M:%S')
                all_list.append([seller, buyer, fixedC,fixedX, name, number_of_nft, local_time])
            time.sleep(.6)

            resells = ("https://proton.api.atomicassets.io/atomicmarket/v1/sales?state=1%2C3&seller_blacklist"
                       "={}&buyer_blacklist={}&collection_name={}&before={}&page=1&limit=100&order"
                       "=desc&sort=updated".format(auther, auther, collection_name,timef))
            all = requests.get(resells).text
            all_ = json.loads(all)



        names_df = pd.DataFrame(data=all_list,
                                columns=["first buyer ", "next buyer", "price paid usd","price paid XPR", "name", "# of nft", "time"])
        names_df.drop(names_df[names_df["first buyer "] == f"{auther}"].index, inplace=True)
        totals = names_df['price paid usd'].sum()
        Royalties = (totals * Royal)
        Rows = int(names_df.index.max() + 1)
        names_df.at[Rows, 'price paid usd'] = Royalties

        for r in dataframe_to_rows(names_df, index=False):
            ws2.append(r)
        dims = {}
        for row in ws2.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
        for col, value in dims.items():
            ws2.column_dimensions[col].width = value
        maxrow = ws2.max_row
        ws2.cell(row=maxrow, column=1, value="Royalties")
        holders = ("https://proton.api.atomicassets.io/atomicassets/v1/accounts?collection_name={}"
                   "&page=1&limit=100&order=desc".format(collection_name))
        holders = requests.get(holders).text
        holders_ = json.loads(holders)
        pages=1
        while len(holders_['data']) != 0:
            pages=pages+1
            for data_info in holders_['data']:
                holders = int(data_info['assets'])
                holder_list.append([data_info['account'], holders])
            holders = ("https://proton.api.atomicassets.io/atomicassets/v1/accounts?collection_name={}"
                       "&page={}&limit=100&order=desc".format(collection_name, pages))
            holders = requests.get(holders).text
            holders_ = json.loads(holders)
        dino_holder_df = pd.DataFrame(data=holder_list, columns=["account ", "amount held"])
        # print(dino_holder_df)
        len(dino_holder_df) - 1
        count = 0
        rowz = 1
        Peoplelist = []
        for r in dataframe_to_rows(dino_holder_df, index=False):
            ws3.append(r)
        for r in dataframe_to_rows(dino_holder_df,index=False):
            count=count+1
            if count==1:
                ""
            else:
                ws3.cell(row=count, column=3).value = 0
                if collection_name == "432543233152":
                    ws3.cell(row=count, column=4).value = 0
        dims = {}
        com = 0
        rare = 0
        epic = 0
        assitID2 = " "
        temp1=0
        temp3 = 0
        temp2 = 0
        totalz=0
        pages=0
        holders = ("https://proton.api.atomicassets.io/atomicassets/v1/accounts?collection_name={}"
                   "&page=1&limit=100&order=desc".format(collection_name))
        holders = requests.get(holders).text
        holders_ = json.loads(holders)
        amount=1
        while len(holders_['data']) != 0:
            pages=pages+1
            amount=amount+1
            for data_info in holders_['data']:
                checker = (data_info['account'])
                print("getting {}'s data".format(checker))
                totalz=0
                people = "https://proton.api.atomicassets.io/atomicmarket/v1/assets?collection_name={}&owner={}&page=1&limit=100&order=desc&sort=asset_id".format(
                    collection_name, checker)
                test = requests.get(people)
                next = test.headers['X-RateLimit-Reset']
                resset = test.headers['X-RateLimit-Remaining']
                resset = int(resset)
                next = int(next)
                wait = next - time.time()
                if resset < 3:
                    time.sleep(wait)
                people_ = json.loads((test.text))
                time.sleep(0.2)
                count = 0
                com = 0
                rare = 0
                epic = 0
                pages=1
                rowz = rowz + 1
                Legendary =0
                while len(people_['data']) != 0:
                    pages = pages + 1
                    if pages==2:
                        count=0
                    if pages==3:
                        count=100
                    if pages==4:
                        count=200
                    for data_info in people_["data"]:
                        word = (data_info["data"]["desc"])
                        nft_name= (data_info["data"]["name"])
                        if checker=="gyasi" or checker=="stachu21" or checker=="install4u222":
                            Legendary=Legendary

                        if collection_name=="432543233152":
                            try:
                                prices=data_info["prices"][0]["min"]
                                prices=int(prices)/1000000
                            except IndexError:
                                prices=5
                        number_of_nft=data_info['template_mint']
                        assitID1 = data_info["asset_id"]
                        if assitID1 != assitID2:
                            assitID2 = assitID1
                            wordT=word
                            words=''
                            if 'Midnight Zombie' in nft_name:
                                words = word[:5]
                                words = ":" + words
                            else:
                                words = ''
                            done = word.lower()
                            count = count + 1

                            ws3.cell(row=rowz, column=6 + count).value = nft_name+'  (#'+number_of_nft+')   '+words
                            s = done.split(" ")
                            if collection_name == "432543233152":
                                lookup = {'Lynxy The Logo':0,'Lynxy #1':191.87979,'Lynxy #2':112.3730397,'Lynxy #3':252.406883,'Lynxy #4':141.2895725,'Lynxy #5':107.0504444,'Lynxy #6':129.5646309,'Lynxy #7':125.6088229,'Lynxy #8':333.0472039,'Lynxy #9':438.0535717,'Lynxy #10':128.841524,'Lynxy #11':173.7796963,'Lynxy #12':177.6369396,'Lynxy #13':557.931838,'Lynxy #14':135.3498394,'Lynxy #15':121.3654573,'Lynxy #16':165.8102899,'Lynxy #17':115.9394095,'Lynxy #18':155.7932061,'Lynxy #19':1332.34563,'Lynxy #20':183.3359316,'Lynxy #21':255.9750108,'Lynxy #22':339.7798271,'Lynxy #23':188.5663036,'Lynxy #24':146.6329134,'Lynxy #25':176.9453603,'Lynxy #26':178.88304,'Lynxy #27':124.1289047,'Lynxy #28':121.4328035,'Lynxy #29':95.47710633,'Lynxy #30':135.7452743,'Lynxy #31':126.3377376,'Lynxy #32':157.5056246,'Lynxy #33':117.6409814,'Lynxy #34':132.5679834,'Lynxy #35':134.3953783,'Lynxy #36':162.3925709,'Lynxy #37':177.6519831,'Lynxy #38':197.2459715,'Lynxy #39':239.674895,'Lynxy #40':173.7276352,'Lynxy #41':1021.507065,'Lynxy #42':141.6173283,'Lynxy #43':156.371177,'Lynxy #44':182.6136235,'Lynxy #45':157.0646871,'Lynxy #46':160.049057,'Lynxy #47':136.484193,'Lynxy #48':140.0434077,'Lynxy #49':167.5823802,'Lynxy #50':210.0032545,'Lynxy #51':220.793036,'Lynxy #52':261.7326654,'Lynxy #53':1186.223512,'Lynxy #54':151.3403755,'Lynxy #55':532.3689809,'Lynxy #56':98.98034201,'Lynxy #57':92.16590641,'Lynxy #58':172.5139803,'Lynxy #59':162.6402327,'Lynxy #60':173.82818,'Lynxy #61':180.2615931,'Lynxy #62':168.5655583,'Lynxy #63':139.2154854,'Lynxy #64':112.8491121,'Lynxy #65':124.9125987,'Lynxy #66':150.8644529,'Lynxy #67':183.4248487,'Lynxy #68':161.625539,'Lynxy #69':191.0107531,'Lynxy #70':108.7237095,'Lynxy #71':133.135592,'Lynxy #72':208.2586638,'Lynxy #73':123.6972877,'Lynxy #74':146.0264212,'Lynxy #75':202.8972395,'Lynxy #76':420.0093342,'Lynxy #77':110.3134589,'Lynxy #78':139.6230538,'Lynxy #79':127.8822481,'Lynxy #80':89.841379,'Lynxy #81':101.8065516,'Lynxy #82':160.2215267,'Lynxy #83':149.1958031,'Lynxy #84':129.7749068,'Lynxy #85':161.461613,'Lynxy #86':138.1981325,'Lynxy #87':129.9357561,'Lynxy #88':153.7237264,'Lynxy #89':216.2479378,'Lynxy #90':294.8253507,'Lynxy #91':125.8469585,'Lynxy #92':139.9755732,'Lynxy #93':167.7466726,'Lynxy #94':134.1657536,'Lynxy #95':121.5247292,'Lynxy #96':84.66685867,'Lynxy #97':308.0047025,'Lynxy #98':190.8457234,'Lynxy #99':181.5732221,'Lynxy #100':182.0732667,'Lynxy #101':804.0214119,'Lynxy #102':138.1883325,'Lynxy #103':134.8538038,'Lynxy #104':262.7468552,'Lynxy #105':334.3678685,'Lynxy #106':131.0484068,'Lynxy #107':91.77874721,'Lynxy #108':126.0326558,'Lynxy #109':189.3536825,'Lynxy #110':159.1447954,'Lynxy #111':137.0721571,'Lynxy #112':132.8288139,'Lynxy #113':300.7629345,'Lynxy #114':139.9548469,'Lynxy #115':115.9782294,'Lynxy #116':207.4845131,'Lynxy #117':133.2947583,'Lynxy #118':149.628211,'Lynxy #119':257.5632932,'Lynxy #120':131.2422702,'Lynxy #121':125.1279046,'Lynxy #122':145.1785792,'Lynxy #123':220.6034082,'Lynxy #124':188.8374058,'Lynxy #125':229.539203,'Lynxy #126':171.6194511,'Lynxy #127':783.2016974,'Lynxy #128':165.0699265,'Lynxy #129':137.9404714,'Lynxy #130':117.7573042,'Lynxy #131':175.0534972,'Lynxy #132':132.1554699,'Lynxy #133':202.8341285,'Lynxy #134':107.8864959,'Lynxy #135':148.7578447,'Lynxy #136':168.4205395,'Lynxy #137':202.3459646,'Lynxy #138':143.326309,'Lynxy #139':193.3436724,'Lynxy #140':144.5077923,'Lynxy #141':107.6621205,'Lynxy #142':170.7704789,'Lynxy #143':124.8739769,'Lynxy #144':135.8452973,'Lynxy #145':128.6124551,'Lynxy #146':154.5117299,'Lynxy #147':172.0445192,'Lynxy #148':153.2584323,'Lynxy #149':218.9131321,'Lynxy #150':126.751998,'Lynxy #151':150.9643863,'Lynxy #152':135.2878661,'Lynxy #153':180.0928405,'Lynxy #154':200.0424165,'Lynxy #155':543.4967755,'Lynxy #156':129.7140332,'Lynxy #157':115.3234756,'Lynxy #158':125.3607042,'Lynxy #159':114.8660745,'Lynxy #160':186.0480124,'Lynxy #161':146.8513543,'Lynxy #162':138.0087379,'Lynxy #163':95.78495462,'Lynxy #164':144.951786,'Lynxy #165':93.29209489,'Lynxy #166':114.6743771,'Lynxy #167':175.7264088,'Lynxy #168':88.57883582,'Lynxy #169':413.2037461,'Lynxy #170':108.5829337,'Lynxy #171':133.5107394,'Lynxy #172':171.0175527,'Lynxy #173':190.9541712,'Lynxy #174':537.8474388,'Lynxy #175':116.7754657,'Lynxy #176':154.6906758,'Lynxy #177':257.8865482,'Lynxy 178':158.8006014,'Lynxy #179':88.44983655,'Lynxy #180':196.6156369,'Lynxy #181':179.3830944,'Lynxy #182':153.9545729,'Lynxy #183':145.9429296,'Lynxy #184':119.3295213,'Lynxy #185':121.9984201,'Lynxy #186':112.842592,'Lynxy #187':177.5050028,'Lynxy #188':134.834694,'Lynxy #189':211.8890451,'Lynxy #190':175.53475,'Lynxy #191':158.2172719,'Lynxy #192':106.7404317,'Lynxy #193':131.1161267,'Lynxy #194':104.286177,'Lynxy #195':120.0266429,'Lynxy #196':220.3042832,'Lynxy #197':149.3886035,'Lynxy #198':183.6165577,'Lynxy #199':187.9785034,'Lynxy #200':141.5185535,'Lynxy #201':207.6860609,'Lynxy #202':178.2670174,'Lynxy #203':117.7796108,'Lynxy #204':151.3681635,'Lynxy #205':183.1540408,'Lynxy #206':135.989906,'Lynxy #207':130.7106175,'Lynxy #208':106.7924426,'Lynxy #209':126.6653204,'Lynxy #210':131.8372391,'Lynxy #211':139.1976176,'Lynxy #212':204.3909489,'Lynxy #213':186.4592234,'Lynxy #214':2219.22367,'Lynxy #215':180.5754348,'Lynxy #216':119.9742406,'Lynxy #217':238.2049627,'Lynxy #218':198.6696817,'Lynxy #219':120.784919,'Lynxy #220':157.3713966,'Lynxy #221':221.5874468,'Lynxy #222':143.3780505,'Lynxy #223':123.7141849,'Lynxy #224':288.7863376,'Lynxy #225':567.1243763,'Lynxy #226':179.638256,'Lynxy #227':97.51279263,'Lynxy 228':140.1515653,'Lynxy #229':238.2132743,'Lynxy #230':131.3787516,'Lynxy #231':133.5353435,'Lynxy #232':176.5564492,'Lynxy #233':154.0383516,'Lynxy #234':215.6980561,'Lynxy #235':151.5295278,'Lynxy #236':170.2531043,'Lynxy #237':164.1175234,'Lynxy #238':106.8457784,'Lynxy #239':209.652011,'Lynxy #240':137.5552265,'Lynxy #241':590.9192061,'Lynxy #242':160.7341558,'Lynxy #243':146.5502207,'Lynxy #244':208.068792,'Lynxy #245':232.3692895,'Lynxy #246':436.5484429,'Lynxy #247':96.19547323,'Lynxy #248':109.3722921,'Lynxy #249':149.6098026,'Lynxy #250':144.8274819,'Lynxy #251':144.048492,'Lynxy #252':192.8651748,'Lynxy #253':146.5959384,'Lynxy #254':266.3822423,'Lynxy #255':176.9459182,'Lynxy #256':142.8154761,'Lynxy #257':244.1307595,'Lynxy #258':125.3030475,'Lynxy #259':140.3805338,'Lynxy #260':109.3610869,'Lynxy #261':134.3218823,'Lynxy #262':213.4193573,'Lynxy #263':151.593864,'Lynxy #264':192.5770541,'Lynxy #265':368.7146776,'Lynxy #266':147.4195308,'Lynxy #267':131.6323787,'Lynxy #268':145.8894778,'Lynxy #269':273.3705649,'Lynxy #270':120.5690759,'Lynxy #271':154.4717966,'Lynxy #272':201.9305587,'Lynxy #273':185.9371297,'Lynxy #274':95.70517469,'Lynxy #275':128.2712891,'Lynxy #276':1339.291488,'Lynxy #277':244.6026951,'Lynxy #278':169.8327739,'Lynxy #279':192.8642828,'Lynxy #280':132.8429364,'Lynxy #281':135.01914,'Lynxy #282':201.8184421,'Lynxy #283':133.346735,'Lynxy #284':175.3949582,'Lynxy #285':134.8085796,'Lynxy #286':189.6020568,'Lynxy #287':214.1242288,'Lynxy #288':207.7999496,'Lynxy #289':263.6904797,'Lynxy #290':156.9561486,'Lynxy #291':155.1512394,'Lynxy #292':98.82976795,'Lynxy #293':238.8516296,'Lynxy #294':121.2767806,'Lynxy #295':155.2917744,'Lynxy #296':184.4032736,'Lynxy #297':226.6435037,'Lynxy #298':200.4722301,'Lynxy #299':133.2007483,'Lynxy #300':162.532077,'Lynxy #301':133.7559756,'Lynxy #302':115.8996587,'Lynxy #303':319.3280796,'Lynxy #304':156.8571912,'Lynxy #305':204.8163147,'Lynxy #306':159.8007475,'Lynxy #307':441.2838414,'Lynxy #308':108.1341586,'Lynxy #309':166.2127841,'Lynxy #310':98.26855706,'Lynxy #311':126.1479068,'Lynxy #312':212.336894,'Lynxy #313':203.7246381,'Lynxy #314':127.8544904,'Lynxy #315':126.3930771,'Lynxy #316':291.9692349,'Lynxy #317':133.20754,'Lynxy #318':224.4643955,'Lynxy #319':220.3157693,'Lynxy #320':141.8152491,'Lynxy #321':112.9440117,'Lynxy #322':130.8759454,'Lynxy #323':145.488081,'Lynxy #324':132.1051902,'Lynxy #325':117.240577,'Lynxy #326':169.8329182,'Lynxy #327':341.8877179,'Lynxy #328':125.3021546,'Lynxy #329':121.406381,'Lynxy #330':168.1503598,'Lynxy #331':189.710239,'Lynxy #332':142.7030239,'Lynxy #333':177.3527776,'Lynxy #334':151.4614953,'Lynxy #335':170.1767848,'Lynxy #336':114.5532319,'Lynxy #337':139.8544038,'Lynxy #338':151.0862652,'Lynxy #339':919.830262,'Lynxy #340':248.4336982,'Lynxy #341':259.5583429,'Lynxy #342':149.9131168,'Lynxy #343':180.7896919,'Lynxy #344':179.2627759,'Lynxy #345':163.507016,'Lynxy #346':139.4320525,'Lynxy #347':100.7661703,'Lynxy #348':188.6186322,'Lynxy #349':130.0010759,'Lynxy #350':236.7052031,'Lynxy #351':132.0338167,'Lynxy #352':206.2831336,'Lynxy #353':134.2165039,'Lynxy #354':124.0018556,'Lynxy #355':188.2057804,'Lynxy #356':154.6713556,'Lynxy #357':278.2240238,'Lynxy #358':152.6249002,'Lynxy #359':115.0273441,'Lynxy #360':299.4867767,'Lynxy #361':425.2996925,'Lynxy #362':141.3259304,'Lynxy #363':154.6858441,'Lynxy #364':180.9541712,'Lynxy #365':412.4665669,'Lynxy #366':147.079403,'Lynxy #367':138.7258086,'Lynxy #368':223.0040679,'Lynxy #369':129.4500949,'Lynxy #370':96.22685063,'Lynxy #371':559.9808907,'Lynxy #372':179.4230438,'Lynxy #373':227.6097802,'Lynxy #374':89.73334414,'Lynxy #375':186.7348378,'Lynxy #376':136.8479861,'Lynxy #377':114.2288503,'Lynxy #378':2396.345002,'Lynxy #379':2039.327339,'Lynxy #380':113.1157652,'Lynxy #381':121.7560209,'Lynxy #382':127.8293176,'Lynxy #383':92.24135903,'Lynxy #384':111.8069172,'Lynxy #385':125.3660403,'Lynxy #386':139.9495875,'Lynxy #387':124.0048224,'Lynxy #388':87.79597593,'Lynxy #389':186.7333932,'Lynxy #390':206.278916,'Lynxy #391':98.99517166,'Lynxy #392':129.4848122,'Lynxy #393':155.7290055,'Lynxy #394':155.0846461,'Lynxy #395':125.6706082,'Lynxy #396':94.94229617,'Lynxy #397':107.1924786,'Lynxy #398':154.3299672,'Lynxy #399':133.9164027,'Lynxy #400':136.5546886,'Lynxy #401':109.1605303,'Lynxy #402':208.283616,'Lynxy #403':260.7325616,'Lynxy #404':99.21775018,'Lynxy #405':225.9758406,'Lynxy #406':233.1034231,'Lynxy #407':151.2790036,'Lynxy #408':169.8329182,'Lynxy #409':112.0949502,'Lynxy #410':204.7416363,'Lynxy #411':95.68205447,'Lynxy #412':176.7177991,'Lynxy #413':162.2949463,'Lynxy #414':106.8167317,'Lynxy #415':122.4405096,'Lynxy #416':100.6339274,'Lynxy #417':95.5220424,'Lynxy #418':604.7353065,'Lynxy #419':143.8990454,'Lynxy #420':192.167986,'Lynxy #421':234.471356,'Lynxy #422':139.9139707,'Lynxy #423':163.7659814,'Lynxy #424':193.5321062,'Lynxy #425':142.479304,'Lynxy #426':170.6546248,'Lynxy 427':126.9337918,'Lynxy #428':113.3168924,'Lynxy #429':108.3931312,'Lynxy #430':178.032927,'Lynxy #431':124.9573019,'Lynxy #432':164.6843951,'Lynxy #433':150.278032,'Lynxy #434':170.0431891,'Lynxy #435':254.2512056,'Lynxy #436':177.0148287,'Lynxy #437':194.1034987,'Lynxy #438':144.9339781,'Lynxy #439':1072.497613,'Lynxy #440':238.1539216,'Lynxy #441':226.7436334,'Lynxy #442':139.3523285,'Lynxy #443':185.4095771,'Lynxy #444':142.4784548,'Lynxy #445':209.3105275,'Lynxy #446':92.74833437,'Lynxy #447':243.5987863,'Lynxy #448':140.1757114,'Lynxy #449':218.2172955,'Lynxy #450':450.8598845,'Lynxy #451':230.1155574,'Lynxy #452':142.5333939,'Lynxy #453':222.4097897,'Lynxy #454':1717.776257,'Lynxy #455':96.90308048,'Lynxy #456':1472.974227,'Lynxy #457':152.5043952,'Lynxy #458':426.594672,'Lynxy #459':179.6827096,'Lynxy #460':98.83345722,'Lynxy #461':133.096781,'Lynxy #462':185.9392995,'Lynxy #463':198.4010421,'Lynxy #464':202.5700132,'Lynxy #465':141.7850214,'Lynxy #466':120.0026526,'Lynxy #467':195.2488893,'Lynxy #468':1184.843774,'Lynxy #469':133.2269977,'Lynxy #470':99.17401643,'Lynxy #471':158.962075,'Lynxy #472':154.2065191,'Lynxy #473':168.8725864,'Lynxy #474':116.6688356,'Lynxy #475':252.3004677,'Lynxy #476':181.3184427,'Lynxy #477':111.728704,'Lynxy #478':283.2782827,'Lynxy #479':179.3492087,'Lynxy #480':287.2858705,'Lynxy #481':128.8394262,'Lynxy #482':187.531137,'Lynxy #483':147.6849046,'Lynxy #484':123.6556979,'Lynxy #485':145.4861841,'Lynxy #486':300.9653561,'Lynxy #487':994.5762492,'Lynxy #488':173.619994,'Lynxy #489':164.2244167,'Lynxy #490':140.9770916,'Lynxy #491':115.5572973,'Lynxy #492':292.0157914,'Lynxy #493':183.7947144,'Lynxy #494':128.2153747,'Lynxy #495':196.9604746,'Lynxy #496':111.5661994,'Lynxy #497':191.5714766,'Lynxy #498':122.42818,'Lynxy #499':122.6112112,'Lynxy #500':103.7404605,'Lynxy #501':97.50481011,'Lynxy #502':129.9504244,'Lynxy #503':283.8096439,'Lynxy #504':806.3840951,'Lynxy #505':112.6144789,'Lynxy #506':207.9493317,'Lynxy #507':140.423689,'Lynxy #508':137.6703986,'Lynxy #509':139.1674595,'Lynxy #510':195.0438293,'Lynxy #511':181.9105162,'Lynxy #512':194.6282269,'Lynxy #513':89.841379,'Lynxy #514':129.6122074,'Lynxy #515':157.6922643,'Lynxy #516':169.9885824,'Lynxy #517':181.4046448,'Lynxy #518':109.3081727,'Lynxy #519':230.7461277,'Lynxy #520':141.6949781,'Lynxy #521':134.8827982,'Lynxy #522':133.2203138,'Lynxy #523':118.7511812,'Lynxy #524':1398.345665,'Lynxy #525':269.6458385,'Lynxy #526':177.663013,'Lynxy #527':119.205973,'Lynxy #528':109.0014644,'Lynxy #529':170.4497442,'Lynxy #530':91.86988243,'Lynxy #531':192.9356363,'Lynxy #532':101.3099865,'Lynxy #533':1869.543481,'Lynxy #534':123.0261597,'Lynxy #535':121.9964577,'Lynxy #536':225.1762823,'Lynxy #537':127.6123599,'Lynxy #538':187.040263,'Lynxy #539':118.5350245,'Lynxy #540':1401.532923,'Lynxy #541':194.632466,'Lynxy #542':125.2786187,'Lynxy #543':134.6638453,'Lynxy #544':152.4120268,'Lynxy #545':287.3412609,'Lynxy #546':341.1983851,'Lynxy #547':110.7377675,'Lynxy #548':204.3611181,'Lynxy #549':805.0586835,'Lynxy #550':195.6348436,'Lynxy #551':138.265074,'Lynxy #552':214.8325729,'Lynxy #553':108.3857988,'Lynxy #554':136.7584877,'Lynxy #555':101.03094,'Lynxy #556':721.6929016,'Lynxy #557':152.9898728,'Lynxy #558':231.2177745,'Lynxy #559':128.4605281,'Lynxy #560':133.8645216,'Lynxy #561':117.047584,'Lynxy #562':153.9195139,'Lynxy #563':130.5731941,'Lynxy #564':125.7674371,'Lynxy #565':119.931373,'Lynxy #566':672.823968,'Lynxy #567':129.916028,'Lynxy #568':102.0311895,'Lynxy #569':269.8073836,'Lynxy #570':194.9404172,'Lynxy #571':194.7755092,'Lynxy #572':176.0132471,'Lynxy #573':1649.832737,'Lynxy #574':168.113747,'Lynxy #575':203.1005514,'Lynxy #576':205.6330246,'Lynxy #577':204.2816745,'Lynxy #578':130.7738429,'Lynxy #579':172.9400736,'Lynxy #580':489.5278364,'Lynxy #581':190.7308917,'Lynxy #582':218.8316294,'Lynxy #583':175.8458228,'Lynxy #584':234.7636151,'Lynxy #585':120.362884,'Lynxy #586':156.6497754,'Lynxy #587':179.4473871,'Lynxy #588':224.0998054,'Lynxy #589':172.0705811,'Lynxy #590':238.2480858,'Lynxy #591':118.5202736,'Lynxy #592':163.0510761,'Lynxy #593':322.4254641,'Lynxy #594':202.4076754,'Lynxy #595':114.741449,'Lynxy #596':662.9953177,'Lynxy #597':129.8882178,'Lynxy #598':125.1789336,'Lynxy #599':186.101857,'Lynxy #600':197.1767896,'Lynxy #601':161.9867147,'Lynxy #602':158.9893619,'Lynxy #603':188.8441584,'Lynxy #604':135.6938737,'Lynxy #605':135.1741974,'Lynxy #606':218.8493762,'Lynxy #607':138.8487895,'Lynxy #608':1020.137892,'Lynxy #609':127.1572254,'Lynxy #610':1080.988753,'Lynxy #611':130.1190769,'Lynxy #612':123.8884159,'Lynxy #613':142.7788012,'Lynxy #614':123.6714471,'Lynxy #615':121.2060582,'Lynxy #616':133.8848452,'Lynxy #617':127.8904609,'Lynxy #618':174.6105506,'Lynxy #619':210.4662036,'Lynxy #620':127.6511798,'Lynxy #621':144.1284557,'Lynxy #622':197.6974091,'Lynxy #623':147.5089912,'Lynxy #624':905.1645036,'Lynxy #625':102.0875271,'Lynxy #626':175.8616212,'Lynxy #627':1580.387497,'Lynxy #628':145.7313406,'Lynxy #629':136.5966018,'Lynxy #630':130.5827219,'Lynxy #631':125.6501486,'Lynxy #632':93.11966551,'Lynxy #633':1358.411409,'Lynxy #634':247.8619233,'Lynxy #635':194.4878675,'Lynxy #636':272.3157159,'Lynxy #637':178.9133699,'Lynxy #638':127.1939089,'Lynxy #639':228.9816726,'Lynxy #640':126.7736333,'Lynxy #641':220.8714162,'Lynxy #642':142.679846,'Lynxy #643':216.7031185,'Lynxy #644':199.6766732,'Lynxy #645':200.0635124,'Lynxy #646':120.9647354,'Lynxy #647':134.4346513,'Lynxy #648':151.4705309,'Lynxy #649':161.4060826,'Lynxy #650':1658.534188,'Lynxy #651':187.3623307,'Lynxy #652':181.0267919,'Lynxy #653':173.4175764,'Lynxy #654':98.9205714,'Lynxy #655':430.5599193,'Lynxy #656':129.6415105,'Lynxy #657':442.0946828,'Lynxy #658':228.3698522,'Lynxy #659':125.6731111,'Lynxy #660':279.9583434,'Lynxy #661':191.2807582,'Lynxy #662':116.6862258,'Lynxy #663':124.6928263,'Lynxy #664':254.8366687,'Lynxy #665':93.83471523,'Lynxy #666':166.2342888,'Lynxy #667':127.7362001,'Lynxy #668':274.3148572,'Lynxy #669':192.2810096,'Lynxy #670':196.481664,'Lynxy #671':162.9779438,'Lynxy #672':202.0669456,'Lynxy #673':164.7620583,'Lynxy #674':159.2148918,'Lynxy #675':126.0725645,'Lynxy #676':143.926489,'Lynxy #677':131.2746813,'Lynxy #678':134.2575633,'Lynxy #679':203.3919241,'Lynxy #680':150.0734395,'Lynxy #681':163.2188952,'Lynxy #682':152.0734145,'Lynxy #683':174.9289711,'Lynxy #684':139.8836552,'Lynxy #685':126.528966,'Lynxy #686':143.7584772,'Lynxy #687':104.2365122,'Lynxy #688':433.3898187,'Lynxy #689':129.7734438,'Lynxy #690':159.4842223,'Lynxy #691':210.1963088,'Lynxy #692':690.3187535,'Lynxy #693':235.3390185,'Lynxy #694':130.2074694,'Lynxy #695':175.9380081,'Lynxy #696':175.150235,'Lynxy #697':118.0316524,'Lynxy #698':145.2859583,'Lynxy #699':238.1575603,'Lynxy #700':210.0229672,'Lynxy #701':103.0479359,'Lynxy #702':131.4295716,'Lynxy #703':138.0322246,'Lynxy #704':128.5557118,'Lynxy #705':311.533575,'Lynxy #706':164.1936002,'Lynxy #707':227.7587638,'Lynxy #708':104.2830809,'Lynxy #709':247.7618203,'Lynxy #710':247.0588032,'Lynxy #711':158.7303697,'Lynxy #712':121.8875214,'Lynxy #713':128.9539698,'Lynxy #714':122.8818721,'Lynxy #715':176.3562801,'Lynxy #716':279.6862781,'Lynxy #717':155.5352199,'Lynxy #718':176.5283206,'Lynxy #719':171.3398778,'Lynxy #720':179.5658491,'Lynxy #721':192.8414261,'Lynxy #722':121.2676851,'Lynxy #723':110.014305,'Lynxy #724':194.8545869,'Lynxy #725':194.3153745,'Lynxy #726':192.7462424,'Lynxy #727':117.9491687,'Lynxy #728':247.6699454,'Lynxy #729':147.7788823,'Lynxy #730':298.5225409,'Lynxy #731':170.3066257,'Lynxy #732':156.2720615,'Lynxy #733':134.3440922,'Lynxy #734':129.8212186,'Lynxy #735':181.4312211,'Lynxy #736':169.9219584,'Lynxy #737':126.8177984,'Lynxy #738':222.846205,'Lynxy #739':117.2627838,'Lynxy #740':168.5833556,'Lynxy #741':192.9866102,'Lynxy #742':156.7754042,'Lynxy #743':141.29226,'Lynxy #744':242.5237251,'Lynxy #745':212.8119253,'Lynxy #746':275.337137,'Lynxy #747':148.5355683,'Lynxy #748':175.3678817,'Lynxy #749':184.9095332,'Lynxy #750':144.0668959,'Lynxy #751':124.9986537,'Lynxy #752':87.94345495,'Lynxy #753':158.2803859,'Lynxy #754':135.18478,'Lynxy #755':214.501285,'Lynxy #756':258.2146007,'Lynxy #757':175.1466195,'Lynxy #758':445.488118,'Lynxy #759':122.3796218,'Lynxy #760':184.4673946,'Lynxy #761':158.28859,'Lynxy #762':184.215672,'Lynzy #763':144.2936085,'Lynxy #764':162.7469932,'Lynxy #765':119.2619899,'Lynxy #766':142.1265143,'Lynxy #767':260.3263336,'Lynxy #768':165.4082635,'Lynxy #769':118.9505062,'Lynxy #770':163.9107031,'Lynxy #771':117.6290368,'Lynxy #772':127.6702353,'Lynxy #773':101.3731877,'Lynxy #774':278.9792011,'Lynxy #775':227.1519003,'Lynxy #776':144.4114319,'Lynxy #777':147.0682541,'Lynxy #778':137.1474683,'Lynxy #779':173.4136337,'Lynxy #780':154.7625149,'Lynxy #781':135.150873,'Lynxy #782':167.5452703,'Lynxy #783':164.2067256,'Lynxy #784':122.9437449,'Lynxy #785':137.6297581,'Lynxy #786':117.1789555,'Lynxy #787':429.6880717,'Lynxy #788':118.858799,'Lynxy #789':171.4052174,'Lynxy #790':153.0912446,'Lynxy #791':116.5935818,'Lynxy #792':186.2636506,'Lynxy #793':138.1657339,'Lynxy #794':133.0554272,'Lynxy #795':136.4805684,'Lynxy #796':100.1607206,'Lynxy #797':126.8721796,'Lynxy #798':124.5029194,'Lynxy #799':157.7827718,'Lynxy #800':298.1950978,'Lynxy #801':141.7978936,'Lynxy #802':121.8116521,'Lynxy #803':135.9875432,'Lynxy #804':126.2822733,'Lynxy #805':272.2202148,'Lynxy #806':131.1989808,'Lynxy #807':132.8920973,'Lynxy #808':198.1652067,'Lynxy #809':163.8573196,'Lynxy #810':112.5856715,'Lynxy #811':165.183703,'Lynxy #812':178.03082,'Lynxy #813':210.2449437,'Lynxy #814':119.1769794,'Lynxy #815':1329.151208,'Lynxy #816':129.7794649,'Lynxy #817':103.6782336,'Lynxy #818':174.1814621,'Lynxy #819':119.5316214,'Lynxy #820':125.2261338,'Lynxy #821':103.6184042,'Lynxy #822':197.7607033,'Lynxy #823':129.6969031,'Lynxy #824':1071.480867,'Lynxy #825':125.6950442,'Lynxy #826':163.123901,'Lynxy #827':177.297682,'Lynxy #828':118.5750131,'Lynxy #829':167.0568455,'Lynxy #830':190.6761674,'Lynxy #831':137.5237373,'Lynxy #832':127.9023187,'Lynxy #833':169.8922428,'Lynxy #834':811.1759048,'Lynxy #835':209.7209333,'Lynxy #836':133.7384104,'Lynxy #837':1863.912598,'Lynxy #838':126.4115505,'Lynxy #839':166.5634411,'Lynxy #840':210.2449437,'Lynxy #841':266.9632885,'Lynxy #842':142.0778185,'Lynxy #843':144.8490965,'Lynxy #844':129.5766523,'Lynxy #845':299.4592443,'Lynxy #846':201.3557403,'Lynxy #847':119.8049491,'Lynxy #848':195.8487348,'Lynxy #849':198.3034638,'Lynxy #850':166.5667874,'Lynxy #851':196.5244687,'Lynxy #852':1025.339282,'Lynxy #853':351.8184421,'Lynxy #854':170.9820411,'Lynxy #855':262.6079341,'Lynxy #856':248.5751003,'Lynxy #857':208.175303,'Lynxy #858':194.8908798,'Lynxy #859':105.1813983,'Lynxy #860':135.9447427,'Lynxy #861':191.8860931,'Lynxy #862':1579.99084,'Lynxy #863':123.466304,'Lynxy #864':177.9768576,'Lynxy #865':102.6632855,'Lynxy #866':293.2928787,'Lynxy #867':116.3843319,'Lynxy #868':168.9409259,'Lynxy #869':182.202266,'Lynxy #870':172.0770799,'Lynxy #871':174.7930282,'Lynxy #872':253.6088859,'Lynxy #873':137.7095812,'Lynxy #874':179.5780274,'Lynxy #875':100.3245251,'Lynxy #876':193.6437776,'Lynxy #877':104.0226296,'Lynxy #878':138.1991659,'Lynxy #879':157.4849197,'Lynxy #880':165.8443612,'Lynxy #881':138.9974131,'Lynxy #882':268.2224402,'Lynxy #883':171.798327,'Lynxy #884':124.1471069,'Lynxy #885':171.041732,'Lynxy #886':167.8876222,'Lynxy #887':236.5953712,'Lynxy #888':174.2653363,'Lynxy #889':1030.270305,'Lynxy #890':214.5612504,'Lynxy #891':185.4890925,'Lynxy #892':116.9276373,'Lynxy #893':119.7532036,'Lynxy #894':195.3838617,'Lynxy #895':152.4894604,'Lynxy #896':102.9794588,'Lynxy #897':100.4327695,'Lynxy #898':130.9623235,'Lynxy #899':555.5653741,'Lynxy #900':129.6600709,'Lynxy #901':146.5428077,'Lynxy #902':196.8442101,'Lynxy #903':142.0881174,'Lynxy #904':1081.328988,'Lynxy #905':149.1597373,'Lynxy #906':131.2393683,'Lynxy #907':314.521425,'Lynxy #908':155.713805,'Lynxy #909':156.8663241,'Lynxy #910':189.2274856,'Lynxy #911':122.975747,'Lynxy #912':281.4764797,'Lynxy #913':1395.696236,'Lynxy #914':119.0978493,'Lynxy #915':116.8213072,'Lynxy #916':190.3163357,'Lynxy #917':231.5348663,'Lynxy #918':160.4321784,'Lynxy #919':122.5192717,'Lynxy #920':220.8266923,'Lynxy #921':181.7128871,'Lynxy #922':153.103887,'Lynxy #923':129.6686686,'Lynxy #924':143.2518638,'Lynxy #925':182.1125384,'Lynxy #926':298.2818503,'Lynxy #927':178.7045756,'Lynxy #928':163.1373858,'Lynxy #929':128.3998901,'Lynxy #930':173.5034523,'Lynxy #931':200.6350905,'Lynxy #932':153.6094181,'Lynxy #933':125.9849139,'Lynxy #934':712.8565281,'Lynxy #935':550.1393679,'Lynxy #936':170.9583506,'Lynxy #937':440.6672427,'Lynxy #938':147.6926165,'Lynxy #939':106.4105382,'Lynxy #940':97.04299794,'Lynxy #941':122.2804171,'Lynxy #942':146.3659075,'Lynxy #943':173.8213242,'Lynxy #944':243.8330049,'Lynxy #945':139.3870115,'Lynxy #946':132.0769235,'Lynxy #947':120.1408504,'Lynxy #948':122.7179875,'Lynxy #949':160.1826807,'Lynxy #950':224.5103973,'Lynxy #951':189.9155183,'Lynxy #952':157.1665681,'Lynxy #953':239.1587392,'Lynxy #954':121.2401516,'Lynxy #955':264.0225176,'Lynxy #956':134.4684693,'Lynxy #957':187.8647368,'Lynxy #958':145.8776501,'Lynxy #959':120.0603723,'Lynxy #960':185.9224261,'Lynxy #961':134.834694,'Lynxy #962':219.4128752,'Lynxy #963':137.5276701,'Lynxy #964':147.9054598,'Lynxy #965':187.1181828,'Lynxy #966':93.83311077,'Lynxy #967':143.1571422,'Lynxy #968':190.5159023,'Lynxy #969':95.33264846,'Lynxy #970':232.2044112,'Lynxy #971':173.0546316,'Lynxy #972':149.9207345,'Lynxy #973':142.440964,'Lynxy #974':185.7236148,'Lynxy #975':173.3547922,'Lynxy #976':107.3056335,'Lynxy #977':125.804763,'Lynxy #978':204.908442,'Lynxy #979':179.4346643,'Lynxy #980':192.971571,'Lynxy #981':102.153966,'Lynxy #982':193.2291554,'Lynxy #983':183.9615714,'Lynxy #984':161.9756459,'Lynxy #985':128.9887606,'Lynxy #986':185.3827584,'Lynxy #987':242.3305932,'Lynxy #988':225.6901086,'Lynxy #989':300.1150635,'Lynxy #990':105.4695947,'Lynxy #991':98.54312479,'Lynxy #992':142.9055796,'Lynxy #993':163.7257854,'Lynxy #994':98.55418158,'Lynxy #995':227.9429396,'Lynxy #996':199.6711875,'Lynxy #997':130.1704601,'Lynxy #998':208.5360734,'Lynxy #999':173.409929,'Lynxy #1000':126.9559877,}
                                ws3.cell(row=1,column=3).value="Points"
                                ws3.cell(row=1,column=4).value="Rarity"
                                temp1=ws3.cell(row=rowz, column=3).value
                                temp2=ws3.cell(row=rowz, column=4).value
                                temp3=ws3.cell(row=rowz, column=5).value
                                if temp1=="0":
                                      ws3.cell(row=rowz, column=3).value = int(prices)+int(totalz)
                                else:
                                    if temp1=="0":
                                        ws3.cell(row=rowz, column=3).value = int(prices) + int(0)
                                    else:
                                        try:
                                            ws3.cell(row=rowz, column=3).value = int(prices) + int(temp1)
                                        except TypeError:
                                            ws3.cell(row=rowz, column=3).value = int(prices) + int(0)
                                try:
                                    ws3.cell(row=rowz, column=4).value = int(lookup[nft_name])+int(temp2)
                                except :
                                    try:
                                        ws3.cell(row=rowz, column=4).value = 0 + int(temp2)
                                    except:
                                        ws3.cell(row=rowz, column=4).value = 0 + int(0)


                                try:
                                    totalz = temp1 + prices
                                except:
                                    totalz = 0 + prices
                            else:
                                try:
                                    prices = data_info["prices"][0]["min"]
                                    prices = int(prices) / 1000000
                                except IndexError:
                                    prices = 0
                                ws3.cell(row=1,column=3).value="Points"
                                temp1=ws3.cell(row=rowz, column=3).value

                                if temp1=="0":
                                      ws3.cell(row=rowz, column=3).value = int(prices)+int(totalz)
                                else:
                                    if temp1=="0":
                                        ws3.cell(row=rowz, column=3).value = int(prices) + int(0)
                                    else:
                                        try:
                                            ws3.cell(row=rowz, column=3).value = int(prices) + int(temp1)
                                        except TypeError:
                                            ws3.cell(row=rowz, column=3).value = int(prices) + int(0)
                                try:
                                    totalz = temp1 + prices
                                except:
                                    totalz = 0 + prices
                    people = "https://proton.api.atomicassets.io/atomicmarket/v1/assets?collection_name={}&owner={}&page={}&limit=100&order=desc&sort=asset_id".format(
                    collection_name, checker,pages)
                    test = requests.get(people)
                    next = test.headers['X-RateLimit-Reset']
                    resset = test.headers['X-RateLimit-Remaining']
                    resset = int(resset)
                    next = int(next)
                    wait = next - time.time()
                    if resset < 3:
                        time.sleep(wait)
                    people_ = json.loads((test.text))
                    time.sleep(0.2)

            holders = ("https://proton.api.atomicassets.io/atomicassets/v1/accounts?collection_name={}"
                       "&page={}&limit=100&order=desc".format(collection_name,amount))
            holders = requests.get(holders).text
            holders_ = json.loads(holders)
        count=0




        for row in ws3.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
            for col, value in dims.items():
                ws3.column_dimensions[col].width = value + 5

        if (collection_name == '133523522522'):

            ws4 = wb.create_sheet("Parents")
            for data_info in parents_['data']:

                word = data_info['data']['desc']
                if word.find('Parent 1:') != -1:
                    start = word.find('Parent 1')
                    start += 10
                    end = word.find('Parent 2')
                    end -= 1
                    new = word[start:end]
                    start2 = word.find('Parent 2')
                    start2 += 10
                    end1 = len(word)
                    new1 = word[start2:end1]

                    parents_list.append([data_info['data']['name'], new, new1])
            s_df = pd.DataFrame(data=parents_list, columns=["Name ", "first parent", "second parent"])

            for r in dataframe_to_rows(s_df, index=False):
                ws4.append(r)
            dimsz = {}
            for row in ws4.rows:
                for cell in row:
                    if cell.value:
                        dimsz[cell.column_letter] = max((dimsz.get(cell.column_letter, 0), len(str(cell.value))))
                for col, value in dimsz.items():
                    ws4.column_dimensions[col].width = value
            excelsave = "".join(excelsheetname)
        excelsave = "".join(excelsheetname)
        wb.save(excelsave)
        print("Creating the excel file")
        wb.close()
        os.chdir(path.parent.absolute())

    normalServic(authers, all, resales, FirstSale, Holders, holders, *excelsheetname)


# collection(auther,collection_name,heading,excelsheetname1)

auther = 'mrfrankie'
universe = 'mrfrankies'
heading = "{} Collection".format(universe)
collection_name = '344241522322'
collection6 = 'Weirdos'
excelsheetname1 = "{}.xlsx".format(collection6)

collection(auther, collection_name, heading, excelsheetname1)

heading = "{} Collection".format(universe)
collection_name = '213523232313'
collection6 = 'Angels and Demons'
excelsheetname1 = "{}.xlsx".format(collection6)
time.sleep(4)
collection(auther, collection_name, heading, excelsheetname1)

heading = "{} Collection".format(universe)
collection_name = '531513514231'
collection6 = 'Proton Chimps'
excelsheetname1 = "{}.xlsx".format(collection6)
time.sleep(4)

collection(auther, collection_name, heading, excelsheetname1)
heading = "{} Collection".format(universe)
collection_name = '432543233152'
collection6 = 'Lynxy the Proton Lynx'
excelsheetname1 = "{}.xlsx".format(collection6)
time.sleep(4)

collection(auther, collection_name, heading, excelsheetname1)
